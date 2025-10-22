#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Análisis del archivo Excel de Administración General
"""

import json
from pathlib import Path

import openpyxl
import pandas as pd


def analyze_excel(file_path):
    """Analiza el archivo Excel y extrae toda la información relevante"""

    print("="*80)
    print("ANÁLISIS DEL ARCHIVO EXCEL - ADMINISTRACIÓN GENERAL")
    print("="*80)

    # Cargar el archivo Excel
    workbook = openpyxl.load_workbook(file_path, data_only=False)

    analysis = {
        'sheets': [],
        'formulas': {},
        'data_structure': {},
        'business_logic': {}
    }

    # Analizar cada hoja
    for sheet_name in workbook.sheetnames:
        print(f"\n{'='*80}")
        print(f"HOJA: {sheet_name}")
        print(f"{'='*80}\n")

        sheet = workbook[sheet_name]
        sheet_info = {
            'name': sheet_name,
            'dimensions': f"{sheet.min_row}:{sheet.max_row}, {sheet.min_column}:{sheet.max_column}",
            'formulas': [],
            'headers': [],
            'data_preview': []
        }

        # Leer con pandas para ver estructura
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            print(f"Dimensiones: {df.shape[0]} filas x {df.shape[1]} columnas")
            print(f"\nColumnas encontradas:")
            for col in df.columns:
                print(f"  - {col}")

            sheet_info['headers'] = list(df.columns)

            print(f"\nPrimeras filas de datos:")
            print(df.head(10).to_string())

            # Guardar preview de datos
            sheet_info['data_preview'] = df.head(20).to_dict('records')

        except Exception as e:
            print(f"No se pudo leer como tabla: {e}")

        # Extraer fórmulas
        print(f"\n{'='*80}")
        print("FÓRMULAS ENCONTRADAS:")
        print(f"{'='*80}\n")

        formula_count = 0
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = {
                        'cell': cell.coordinate,
                        'formula': cell.value,
                        'value': cell.value
                    }
                    sheet_info['formulas'].append(formula)
                    print(f"  {cell.coordinate}: {cell.value}")
                    formula_count += 1

        if formula_count == 0:
            print("  No se encontraron fórmulas en esta hoja")
        else:
            print(f"\nTotal de fórmulas: {formula_count}")

        # Buscar patrones de negocio
        print(f"\n{'='*80}")
        print("ANÁLISIS DE LÓGICA DE NEGOCIO:")
        print(f"{'='*80}\n")

        # Buscar celdas con palabras clave
        keywords = [
            'total', 'costo', 'precio', 'venta', 'compra', 'ganancia', 'margen',
            'capital', 'banco', 'utilidad', 'gasto', 'ingreso', 'adeudo', 'deuda',
            'inventario', 'stock', 'cantidad', 'producto', 'cliente', 'proveedor',
            'distribuidor', 'orden', 'transferencia', 'ventas', 'fletes'
        ]

        business_cells = []
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell_lower = str(cell.value).lower()
                    for keyword in keywords:
                        if keyword in cell_lower:
                            business_cells.append({
                                'cell': cell.coordinate,
                                'keyword': keyword,
                                'value': cell.value
                            })
                            break

        if business_cells:
            print("Celdas relevantes para lógica de negocio:")
            for bc in business_cells[:30]:  # Limitar a 30
                print(f"  {bc['cell']}: {bc['value']} (keyword: {bc['keyword']})")

        analysis['sheets'].append(sheet_info)

    # Guardar análisis en JSON
    output_file = Path(file_path).parent / 'excel_analysis.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(analysis, f, indent=2, ensure_ascii=False)

    print(f"\n{'='*80}")
    print(f"Análisis completo guardado en: {output_file}")
    print(f"{'='*80}\n")

    return analysis

if __name__ == "__main__":
    file_path = r"C:\Users\xpovo\Documents\premium-ecosystem\Administación_General.xlsx"
    analyze_excel(file_path)
