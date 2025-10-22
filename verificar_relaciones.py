import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\xpovo\Downloads\Copia de Administación_General.xlsx', data_only=True)

print('='*100)
print('VERIFICACIÓN DE RELACIONES ENTRE MONTOS Y VARIABLES')
print('='*100)

# 1. Analizar ventas en Control_Maestro
ws_control = wb['Control_Maestro']
ws_boveda = wb['Bóveda_Monte']
ws_almacen = wb['Almacen_Monte']
ws_clientes = wb['Clientes']

print('\n📊 ANÁLISIS 1: Verificar Ventas Pagadas vs Ingresos en Bóveda_Monte')
print('='*80)

# Obtener ventas pagadas de Control_Maestro
ventas_pagadas = []
ventas_pendientes = []
for row in range(4, 100):
    fecha = ws_control.cell(row, 1).value
    if fecha:
        cliente = ws_control.cell(row, 4).value
        boveda_monte = ws_control.cell(row, 5).value
        ingreso = ws_control.cell(row, 7).value
        estatus = ws_control.cell(row, 11).value
        cantidad = ws_control.cell(row, 3).value
        precio_venta = ws_control.cell(row, 6).value
        flete_utilidad = ws_control.cell(row, 9).value
        utilidad = ws_control.cell(row, 10).value

        venta = {
            'fila': row,
            'fecha': fecha,
            'cliente': cliente,
            'cantidad': cantidad,
            'precio_venta': precio_venta,
            'boveda_monte': boveda_monte,
            'ingreso': ingreso,
            'flete_utilidad': flete_utilidad,
            'utilidad': utilidad,
            'estatus': estatus
        }

        if estatus == 'Pagado':
            ventas_pagadas.append(venta)
        elif estatus == 'Pendiente':
            ventas_pendientes.append(venta)

print(f'\nVentas Pagadas encontradas: {len(ventas_pagadas)}')
print(f'Ventas Pendientes encontradas: {len(ventas_pendientes)}')

print('\nPRIMERAS 10 VENTAS PAGADAS (con detalle de montos):')
for v in ventas_pagadas[:10]:
    print(f'  Fila {v["fila"]}: {v["cliente"]}')
    print(f'    Cantidad: {v["cantidad"]} x ${v["precio_venta"]:,.0f} = ${v["ingreso"]:,.0f}')
    print(f'    Bóveda Monte (base): ${v["boveda_monte"]:,.0f}')
    print(f'    Flete Utilidad: ${v["flete_utilidad"] or 0:,.0f}')
    print(f'    Utilidad: ${v["utilidad"] or 0:,.0f}')
    print(f'    Ingreso Total: ${v["ingreso"]:,.0f}')
    print(f'    Status: {v["estatus"]}')
    print()

print('\nPRIMERAS 10 VENTAS PENDIENTES (con detalle de montos):')
for v in ventas_pendientes[:10]:
    print(f'  Fila {v["fila"]}: {v["cliente"]}')
    print(f'    Cantidad: {v["cantidad"]} x ${v["precio_venta"]:,.0f} = ${v["ingreso"]:,.0f}')
    print(f'    Bóveda Monte (base): ${v["boveda_monte"]:,.0f}')
    print(f'    Flete Utilidad: ${v["flete_utilidad"] or 0:,.0f}')
    print(f'    Utilidad: ${v["utilidad"] or 0:,.0f}')
    print(f'    Ingreso Total: ${v["ingreso"]:,.0f}')
    print(f'    Status: {v["estatus"]}')
    print()

# 2. Verificar ingresos en Bóveda_Monte
print('\n\n📊 ANÁLISIS 2: Ingresos registrados en Bóveda_Monte')
print('='*80)

ingresos_boveda = []
for row in range(4, 100):
    fecha = ws_boveda.cell(row, 1).value
    if fecha:
        cliente = ws_boveda.cell(row, 2).value
        ingreso = ws_boveda.cell(row, 3).value

        ingresos_boveda.append({
            'fila': row,
            'fecha': fecha,
            'cliente': cliente,
            'ingreso': ingreso
        })

print(f'\nIngresos en Bóveda_Monte: {len(ingresos_boveda)}')
print('PRIMEROS 15 INGRESOS:')
for ing in ingresos_boveda[:15]:
    print(f'  Fila {ing["fila"]}: {ing["cliente"]} - ${ing["ingreso"]:,.0f}')

# 3. Calcular totales
print('\n\n📊 ANÁLISIS 3: TOTALES Y RESUMEN')
print('='*80)

total_ventas_pagadas_boveda = sum(v['boveda_monte'] for v in ventas_pagadas if v['boveda_monte'])
total_ventas_pagadas_ingreso = sum(v['ingreso'] for v in ventas_pagadas if v['ingreso'])
total_ventas_pendientes_boveda = sum(v['boveda_monte'] for v in ventas_pendientes if v['boveda_monte'])
total_ventas_pendientes_ingreso = sum(v['ingreso'] for v in ventas_pendientes if v['ingreso'])
total_ingresos_boveda = sum(i['ingreso'] for i in ingresos_boveda if i['ingreso'])

print(f'\n💰 VENTAS PAGADAS (Control_Maestro):')
print(f'  Total Bóveda Monte (campo base): ${total_ventas_pagadas_boveda:,.2f}')
print(f'  Total Ingreso (precio venta * cantidad): ${total_ventas_pagadas_ingreso:,.2f}')

print(f'\n💰 VENTAS PENDIENTES (Control_Maestro):')
print(f'  Total Bóveda Monte (campo base): ${total_ventas_pendientes_boveda:,.2f}')
print(f'  Total Ingreso (precio venta * cantidad): ${total_ventas_pendientes_ingreso:,.2f}')

print(f'\n💰 INGRESOS (Bóveda_Monte):')
print(f'  Total Ingresos Registrados: ${total_ingresos_boveda:,.2f}')

print(f'\n🔍 VERIFICACIÓN:')
print(f'  ✓ Ventas Pagadas (Bóveda Monte) = ${total_ventas_pagadas_boveda:,.2f}')
print(f'  ✓ Ingresos en Bóveda_Monte = ${total_ingresos_boveda:,.2f}')
print(f'  ✓ Diferencia = ${total_ventas_pagadas_boveda - total_ingresos_boveda:,.2f}')
print(f'\n  NOTA: Los ingresos en Bóveda_Monte deben ser igual al campo "Bóveda Monte" de ventas PAGADAS')

# 4. Analizar clientes con deuda
print('\n\n📊 ANÁLISIS 4: Clientes con Deuda (hoja Clientes)')
print('='*80)

clientes_con_deuda = []
clientes_todos = []
for row in range(4, 50):
    cliente = ws_clientes.cell(row, 5).value
    if cliente:
        deuda = ws_clientes.cell(row, 7).value or 0
        abonos = ws_clientes.cell(row, 8).value or 0
        pendiente = ws_clientes.cell(row, 9).value or 0

        cliente_data = {
            'nombre': cliente,
            'deuda': deuda,
            'abonos': abonos,
            'pendiente': pendiente
        }

        clientes_todos.append(cliente_data)

        if pendiente and pendiente != 0:
            clientes_con_deuda.append(cliente_data)

print(f'\nTotal clientes: {len(clientes_todos)}')
print(f'Clientes con deuda pendiente: {len(clientes_con_deuda)}')

print('\n📋 TODOS LOS CLIENTES CON DEUDA PENDIENTE:')
total_deuda = 0
total_abonos = 0
total_pendiente = 0

for c in clientes_con_deuda:
    print(f'\n  {c["nombre"]}:')
    print(f'    Deuda Total: ${c["deuda"]:,.0f}')
    print(f'    Abonos: ${c["abonos"]:,.0f}')
    print(f'    Pendiente: ${c["pendiente"]:,.0f}')

    total_deuda += c["deuda"]
    total_abonos += c["abonos"]
    total_pendiente += c["pendiente"]

print(f'\n💰 TOTALES DE CLIENTES:')
print(f'  Total Deuda: ${total_deuda:,.2f}')
print(f'  Total Abonos: ${total_abonos:,.2f}')
print(f'  Total Pendiente: ${total_pendiente:,.2f}')

# 5. Análisis de estructura de montos
print('\n\n📊 ANÁLISIS 5: ENTENDIMIENTO DE LA ESTRUCTURA DE MONTOS')
print('='*80)

print(f'\nRelación entre campos en Control_Maestro:')
print(f'  • Bóveda Monte = Cantidad × Costo Base (generalmente 6300)')
print(f'  • Ingreso = Cantidad × Precio De Venta')
print(f'  • Flete Utilidad = Cantidad × 500 (cuando aplica flete)')
print(f'  • Utilidad = Ingreso - Bóveda Monte - Flete Utilidad')
print(f'  • Estatus = "Pagado" o "Pendiente"')

print(f'\nFlujo de dinero:')
print(f'  1. Venta se registra en Control_Maestro con Estatus="Pendiente"')
print(f'  2. Cliente paga → Estatus cambia a "Pagado"')
print(f'  3. Cuando Estatus="Pagado" → Se registra ingreso en Bóveda_Monte')
print(f'  4. El monto que entra a Bóveda_Monte es el campo "Bóveda Monte" (no el Ingreso total)')

print(f'\nDeuda de clientes:')
print(f'  • Deuda = SUMA de todas las ventas Pendientes del cliente (campo Bóveda Monte)')
print(f'  • Abonos = SUMA de abonos registrados en GYA para ese cliente')
print(f'  • Pendiente = Deuda - Abonos')

print('\n' + '='*100)
print('ANÁLISIS COMPLETADO')
print('='*100)
