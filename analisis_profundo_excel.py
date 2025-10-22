#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ANÁLISIS ULTRA PROFUNDO DEL EXCEL
Extrae TODA la estructura, datos y lógica de negocio
"""

import json
from datetime import datetime

import openpyxl


def analizar_excel_completo():
    wb = openpyxl.load_workbook('Administación_General.xlsx', data_only=False)

    print('=' * 80)
    print('ANÁLISIS ULTRA PROFUNDO DEL EXCEL - EXTRACCIÓN COMPLETA DE LÓGICA')
    print('=' * 80)

    resultado = {
        'clientes': [],
        'ventas': [],
        'gastos_abonos': [],
        'bancos': [],
        'formulas_criticas': [],
        'estructura': {}
    }

    # 1. Analizar estructura de CLIENTES
    print('\n### 1. ESTRUCTURA DE CLIENTES ###')
    if 'Clientes' in wb.sheetnames:
        ws = wb['Clientes']
        print(f'Rango usado: {ws.dimensions}')

        # Leer headers
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(1, col).value
            if val:
                headers.append(val)
        print(f'Columnas: {headers}')
        resultado['estructura']['clientes'] = headers

        # Leer TODOS los clientes
        print(f'\nExtrayendo {ws.max_row - 1} clientes...')
        for row in range(2, ws.max_row + 1):
            cliente = {}
            for idx, header in enumerate(headers, 1):
                cell = ws.cell(row, idx)
                if cell.data_type == 'f':
                    cliente[header] = {'formula': cell.value, 'valor': cell.value}
                else:
                    cliente[header] = cell.value

            if any(cliente.values()):  # Solo agregar si tiene datos
                resultado['clientes'].append(cliente)

        print(f'Clientes extraídos: {len(resultado["clientes"])}')
        if resultado['clientes']:
            print(f'Ejemplo: {json.dumps(resultado["clientes"][0], default=str, ensure_ascii=False)}')

    # 2. Analizar V_Monte (Ventas)
    print('\n### 2. ESTRUCTURA DE VENTAS ###')
    for sheet_name in wb.sheetnames:
        if 'V_' in sheet_name:
            ws_ventas = wb[sheet_name]
            print(f'Sheet encontrado: {sheet_name}')
            print(f'Rango usado: {ws_ventas.dimensions}')

            # Headers
            headers_ventas = []
            for col in range(1, ws_ventas.max_column + 1):
                val = ws_ventas.cell(1, col).value
                if val:
                    headers_ventas.append(val)
            print(f'Columnas: {headers_ventas}')
            resultado['estructura']['ventas'] = headers_ventas

            # Leer TODAS las ventas
            print(f'\nExtrayendo ventas de {sheet_name}...')
            for row in range(2, min(ws_ventas.max_row + 1, 100)):  # Primeras 100
                venta = {'sheet': sheet_name}
                for idx, header in enumerate(headers_ventas, 1):
                    cell = ws_ventas.cell(row, idx)
                    if cell.data_type == 'f':
                        venta[header] = {'formula': cell.value, 'tipo': 'formula'}
                    else:
                        venta[header] = cell.value

                if any(v for k, v in venta.items() if k != 'sheet'):
                    resultado['ventas'].append(venta)

            print(f'Ventas extraídas de {sheet_name}: {len([v for v in resultado["ventas"] if v.get("sheet") == sheet_name])}')

    # 3. Analizar G_Monte (Gastos/Abonos)
    print('\n### 3. ESTRUCTURA DE GASTOS/ABONOS ###')
    for sheet_name in wb.sheetnames:
        if 'G_' in sheet_name:
            ws_gastos = wb[sheet_name]
            print(f'Sheet encontrado: {sheet_name}')
            print(f'Rango usado: {ws_gastos.dimensions}')

            # Headers
            headers_gastos = []
            for col in range(1, ws_gastos.max_column + 1):
                val = ws_gastos.cell(1, col).value
                if val:
                    headers_gastos.append(val)
            print(f'Columnas: {headers_gastos}')
            resultado['estructura']['gastos'] = headers_gastos

            # Leer TODOS los gastos/abonos
            print(f'\nExtrayendo gastos/abonos de {sheet_name}...')
            for row in range(2, min(ws_gastos.max_row + 1, 100)):  # Primeras 100
                gasto = {'sheet': sheet_name}
                for idx, header in enumerate(headers_gastos, 1):
                    cell = ws_gastos.cell(row, idx)
                    if cell.data_type == 'f':
                        gasto[header] = {'formula': cell.value, 'tipo': 'formula'}
                    else:
                        gasto[header] = cell.value

                if any(v for k, v in gasto.items() if k != 'sheet'):
                    resultado['gastos_abonos'].append(gasto)

            print(f'Gastos/Abonos extraídos de {sheet_name}: {len([g for g in resultado["gastos_abonos"] if g.get("sheet") == sheet_name])}')

    # 4. Analizar BANCOS desde DATA
    print('\n### 4. IDENTIFICAR BANCOS ###')
    if 'DATA' in wb.sheetnames:
        ws_data = wb['DATA']
        bancos_keywords = ['bóveda', 'utilidades', 'flete', 'azteca', 'leftie', 'profit', 'banco']

        for row in ws_data.iter_rows(max_row=50):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for keyword in bancos_keywords:
                        if keyword in cell.value.lower():
                            resultado['bancos'].append({
                                'nombre': cell.value,
                                'celda': cell.coordinate,
                                'keyword': keyword
                            })

    print(f'Bancos identificados: {len(resultado["bancos"])}')
    for banco in resultado['bancos']:
        print(f'  - {banco["nombre"]} [{banco["celda"]}]')

    # 5. EXTRAER FÓRMULAS CRÍTICAS
    print('\n### 5. FÓRMULAS CRÍTICAS ###')
    keywords_formulas = ['SUMIF', 'SUMIFS', 'ADEUDO', 'DEUDA', 'ESTATUS', 'PENDIENTE', 'PAGADO', 'ABONO']

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value:
                    formula = str(cell.value).upper()
                    if any(keyword in formula for keyword in keywords_formulas):
                        resultado['formulas_criticas'].append({
                            'sheet': sheet.title,
                            'celda': cell.coordinate,
                            'formula': cell.value
                        })

    print(f'Fórmulas críticas encontradas: {len(resultado["formulas_criticas"])}')
    print('\nPrimeras 15 fórmulas más relevantes:')
    for f in resultado['formulas_criticas'][:15]:
        print(f'  [{f["sheet"]}!{f["celda"]}]')
        print(f'    {f["formula"][:120]}...' if len(f["formula"]) > 120 else f'    {f["formula"]}')

    # 6. Guardar resultado
    print('\n### 6. GUARDANDO ANÁLISIS COMPLETO ###')
    with open('ANALISIS_EXCEL_COMPLETO.json', 'w', encoding='utf-8') as f:
        # Convertir a formato serializable
        resultado_serializable = {
            'metadata': {
                'fecha_analisis': datetime.now().isoformat(),
                'total_sheets': len(wb.sheetnames),
                'sheets': wb.sheetnames
            },
            'estructura': resultado['estructura'],
            'totales': {
                'clientes': len(resultado['clientes']),
                'ventas': len(resultado['ventas']),
                'gastos_abonos': len(resultado['gastos_abonos']),
                'bancos': len(resultado['bancos']),
                'formulas_criticas': len(resultado['formulas_criticas'])
            },
            'muestras': {
                'clientes': resultado['clientes'][:5],
                'ventas': resultado['ventas'][:5],
                'gastos_abonos': resultado['gastos_abonos'][:5],
                'bancos': resultado['bancos'],
                'formulas_criticas': resultado['formulas_criticas'][:20]
            }
        }

        json.dump(resultado_serializable, f, indent=2, ensure_ascii=False, default=str)

    print('✅ Análisis guardado en: ANALISIS_EXCEL_COMPLETO.json')

    # 7. Crear archivo de mapeo de lógica
    print('\n### 7. CREANDO MAPEO DE LÓGICA ###')
    with open('MAPEO_LOGICA_EXCEL_A_SISTEMA.md', 'w', encoding='utf-8') as f:
        f.write('# 🎯 MAPEO DE LÓGICA: EXCEL → SISTEMA FLOWDISTRIBUTOR\n\n')
        f.write('## 1. ESTRUCTURA DE CLIENTES\n\n')
        f.write(f'**Columnas Excel**: {", ".join(resultado["estructura"].get("clientes", []))}\n\n')
        f.write('**Mapeo al Sistema**:\n')
        f.write('```javascript\n')
        f.write('{\n')
        for col in resultado["estructura"].get("clientes", [])[:10]:
            f.write(f'  {col.lower().replace(" ", "_")}: valor,\n')
        f.write('}\n```\n\n')

        f.write('## 2. ESTRUCTURA DE VENTAS\n\n')
        f.write(f'**Columnas Excel**: {", ".join(resultado["estructura"].get("ventas", []))}\n\n')
        f.write('**Mapeo al Sistema**:\n')
        f.write('```javascript\n')
        f.write('{\n')
        for col in resultado["estructura"].get("ventas", [])[:10]:
            f.write(f'  {col.lower().replace(" ", "_")}: valor,\n')
        f.write('}\n```\n\n')

        f.write('## 3. FÓRMULAS CRÍTICAS A IMPLEMENTAR\n\n')
        for i, formula in enumerate(resultado['formulas_criticas'][:10], 1):
            f.write(f'### Fórmula {i}: {formula["sheet"]}!{formula["celda"]}\n')
            f.write(f'```excel\n{formula["formula"]}\n```\n\n')

    print('✅ Mapeo guardado en: MAPEO_LOGICA_EXCEL_A_SISTEMA.md')

    print('\n' + '=' * 80)
    print('✅ ANÁLISIS ULTRA PROFUNDO COMPLETADO')
    print('=' * 80)
    print(f'\n📊 RESUMEN:')
    print(f'  - Clientes: {len(resultado["clientes"])}')
    print(f'  - Ventas: {len(resultado["ventas"])}')
    print(f'  - Gastos/Abonos: {len(resultado["gastos_abonos"])}')
    print(f'  - Bancos: {len(resultado["bancos"])}')
    print(f'  - Fórmulas críticas: {len(resultado["formulas_criticas"])}')

    return resultado

if __name__ == '__main__':
    analizar_excel_completo()
