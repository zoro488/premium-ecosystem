#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Análisis detallado de las 2 tablas en la hoja Distribuidores
"""
import openpyxl
import json
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

# Cargar Excel
wb = openpyxl.load_workbook(r'C:\Users\xpovo\Documents\premium-ecosystem\Administación_General.xlsx', data_only=True)
ws = wb['Distribuidores']

print('=' * 120)
print('📊 ANÁLISIS COMPLETO: HOJA DISTRIBUIDORES - 2 TABLAS')
print('=' * 120)

print(f'\n✅ Hoja: "Distribuidores"')
print(f'   Dimensiones: {ws.max_row} filas x {ws.max_column} columnas')

# TABLA 1: ÓRDENES DE COMPRA (Columnas A-K)
print('\n' + '=' * 120)
print('📋 TABLA 1: ÓRDENES DE COMPRA')
print('=' * 120)

tabla1_headers = []
for col in range(1, 12):  # A-K
    valor = ws.cell(3, col).value
    if valor:
        tabla1_headers.append(valor)

print(f'\n📋 Encabezados ({len(tabla1_headers)}):')
for i, h in enumerate(tabla1_headers, 1):
    print(f'   {i}. {h}')

# Leer datos de Tabla 1
tabla1_datos = []
for fila in range(4, 20):  # Leer primeras filas
    if ws.cell(fila, 1).value:  # Si hay OC
        registro = {}
        for i, header in enumerate(tabla1_headers, 1):
            valor = ws.cell(fila, i).value
            if isinstance(valor, datetime):
                registro[header] = valor.strftime('%Y-%m-%d')
            else:
                registro[header] = valor
        tabla1_datos.append(registro)

print(f'\n📊 Total registros encontrados: {len(tabla1_datos)}')
print(f'\n💾 PRIMEROS 5 REGISTROS:\n')

for i, reg in enumerate(tabla1_datos[:5], 1):
    print(f'Registro {i}:')
    for key, val in reg.items():
        print(f'   {key}: {val}')
    print()

# TABLA 2: DISTRIBUIDORES CONSOLIDADO (Columnas M-P)
print('\n' + '=' * 120)
print('📋 TABLA 2: DISTRIBUIDORES CONSOLIDADO')
print('=' * 120)

tabla2_headers = []
for col in range(13, 17):  # M-P (13-16)
    valor = ws.cell(3, col).value
    if valor:
        tabla2_headers.append(valor)

print(f'\n📋 Encabezados ({len(tabla2_headers)}):')
for i, h in enumerate(tabla2_headers, 1):
    print(f'   {i}. {h}')

# Leer datos de Tabla 2
tabla2_datos = []
for fila in range(4, 20):
    if ws.cell(fila, 13).value:  # Si hay nombre de distribuidor
        registro = {}
        for i, header in enumerate(tabla2_headers):
            valor = ws.cell(fila, 13 + i).value
            registro[header] = valor
        tabla2_datos.append(registro)

print(f'\n📊 Total registros encontrados: {len(tabla2_datos)}')
print(f'\n💾 TODOS LOS REGISTROS:\n')

for i, reg in enumerate(tabla2_datos, 1):
    print(f'Distribuidor {i}:')
    for key, val in reg.items():
        print(f'   {key}: {val}')
    print()

# COMPARACIÓN CON PANELES
print('\n' + '=' * 120)
print('🔍 COMPARACIÓN: DATOS EXCEL vs PANELES')
print('=' * 120)

# Definir campos esperados en cada panel
campos_panel_ordenes = [
    'OC', 'Fecha', 'Origen', 'Cantidad',
    'Costo Distribuidor', 'Costo Transporte', 'Costo Por Unidad',
    'Stock Actual', 'Costo Total', 'Pago a Distribuidor', 'Deuda'
]

campos_panel_distribuidores = [
    'Distribuidores', 'Costo total', 'Abonos', 'Pendiente'
]

print('\n📊 PANEL ÓRDENES DE COMPRA')
print('-' * 120)
print(f'✓ Campos en Excel: {len(tabla1_headers)}')
print(f'✓ Campos esperados en Panel: {len(campos_panel_ordenes)}')

campos_faltantes_oc = []
for campo in tabla1_headers:
    if campo not in campos_panel_ordenes:
        campos_faltantes_oc.append(campo)

if campos_faltantes_oc:
    print(f'\n⚠️  Campos en Excel pero no esperados en panel: {", ".join(campos_faltantes_oc)}')
else:
    print('\n✅ Todos los campos del Excel coinciden con los esperados')

# Verificar campos faltantes en el panel
print('\n🔍 Verificando implementación en PanelOrdenesCompra.jsx...')

import os
ruta_panel_oc = r'C:\Users\xpovo\Documents\premium-ecosystem\src\apps\FlowDistributor\components\PanelOrdenesCompra.jsx'

if os.path.exists(ruta_panel_oc):
    with open(ruta_panel_oc, 'r', encoding='utf-8') as f:
        contenido_oc = f.read()

    campos_usados_oc = []
    campos_no_usados_oc = []

    for campo in tabla1_headers:
        # Convertir a camelCase
        camel = campo.split(' ')
        if len(camel) > 1:
            camel_case = camel[0].lower() + ''.join(word.capitalize() for word in camel[1:])
        else:
            camel_case = campo.lower()

        # Buscar en el código
        if f'.{camel_case}' in contenido_oc or f'"{campo}"' in contenido_oc or f"'{campo}'" in contenido_oc:
            campos_usados_oc.append(campo)
        else:
            campos_no_usados_oc.append(campo)

    print(f'\n✅ CAMPOS USADOS en PanelOrdenesCompra ({len(campos_usados_oc)}/{len(tabla1_headers)}):')
    for campo in campos_usados_oc:
        print(f'   ✓ {campo}')

    if campos_no_usados_oc:
        print(f'\n❌ CAMPOS NO USADOS en PanelOrdenesCompra ({len(campos_no_usados_oc)}):')
        for campo in campos_no_usados_oc:
            print(f'   ✗ {campo}')

print('\n' + '-' * 120)
print('📊 PANEL DISTRIBUIDORES')
print('-' * 120)
print(f'✓ Campos en Excel: {len(tabla2_headers)}')
print(f'✓ Campos esperados en Panel: {len(campos_panel_distribuidores)}')

# Verificar PanelDistribuidores
ruta_panel_dist = r'C:\Users\xpovo\Documents\premium-ecosystem\src\apps\FlowDistributor\components\PanelDistribuidores.jsx'

if os.path.exists(ruta_panel_dist):
    with open(ruta_panel_dist, 'r', encoding='utf-8') as f:
        contenido_dist = f.read()

    campos_usados_dist = []
    campos_no_usados_dist = []

    for campo in tabla2_headers:
        # Convertir a camelCase
        camel = campo.split(' ')
        if len(camel) > 1:
            camel_case = camel[0].lower() + ''.join(word.capitalize() for word in camel[1:])
        else:
            camel_case = campo.lower()

        # Buscar en el código
        if f'.{camel_case}' in contenido_dist or f'"{campo}"' in contenido_dist or f"'{campo}'" in contenido_dist:
            campos_usados_dist.append(campo)
        else:
            campos_no_usados_dist.append(campo)

    print(f'\n✅ CAMPOS USADOS en PanelDistribuidores ({len(campos_usados_dist)}/{len(tabla2_headers)}):')
    for campo in campos_usados_dist:
        print(f'   ✓ {campo}')

    if campos_no_usados_dist:
        print(f'\n❌ CAMPOS NO USADOS en PanelDistribuidores ({len(campos_no_usados_dist)}):')
        for campo in campos_no_usados_dist:
            print(f'   ✗ {campo}')

# RESUMEN FINAL
print('\n' + '=' * 120)
print('📈 RESUMEN EJECUTIVO')
print('=' * 120)

print(f'\n📊 TABLA 1 - ÓRDENES DE COMPRA:')
print(f'   Total columnas en Excel: {len(tabla1_headers)}')
print(f'   Total registros: {len(tabla1_datos)}')
print(f'   Campos usados en Panel: {len(campos_usados_oc)} ({round(len(campos_usados_oc)/len(tabla1_headers)*100)}%)')
print(f'   Campos NO usados: {len(campos_no_usados_oc)}')

print(f'\n📊 TABLA 2 - DISTRIBUIDORES:')
print(f'   Total columnas en Excel: {len(tabla2_headers)}')
print(f'   Total registros: {len(tabla2_datos)}')
print(f'   Campos usados en Panel: {len(campos_usados_dist)} ({round(len(campos_usados_dist)/len(tabla2_headers)*100)}%)')
print(f'   Campos NO usados: {len(campos_no_usados_dist)}')

# Guardar resultados
resultados = {
    'tabla_ordenes_compra': {
        'headers': tabla1_headers,
        'total_registros': len(tabla1_datos),
        'registros': tabla1_datos,
        'campos_usados': campos_usados_oc,
        'campos_no_usados': campos_no_usados_oc,
        'porcentaje_uso': round(len(campos_usados_oc)/len(tabla1_headers)*100)
    },
    'tabla_distribuidores': {
        'headers': tabla2_headers,
        'total_registros': len(tabla2_datos),
        'registros': tabla2_datos,
        'campos_usados': campos_usados_dist,
        'campos_no_usados': campos_no_usados_dist,
        'porcentaje_uso': round(len(campos_usados_dist)/len(tabla2_headers)*100)
    }
}

with open('scripts/analisis-distribuidores-2-tablas.json', 'w', encoding='utf-8') as f:
    json.dump(resultados, f, ensure_ascii=False, indent=2)

print('\n💾 Resultados guardados en: scripts/analisis-distribuidores-2-tablas.json')
print('\n' + '=' * 120 + '\n')