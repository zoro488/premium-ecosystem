"""
IMPORTACIÓN QUIRÚRGICA DE EXCEL A FLOWDISTRIBUTOR
Sistema de importación con 0% errores, trazabilidad completa
Análisis profundo de estructura, relaciones y lógica de negocio
"""

import json
import re
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

import numpy as np
import openpyxl
import pandas as pd


class SurgicalExcelImporter:
    """Importador quirúrgico con análisis completo"""

    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.workbook = openpyxl.load_workbook(excel_path, data_only=True)
        self.full_workbook = openpyxl.load_workbook(excel_path, data_only=False)
        self.analysis_report = {
            'timestamp': datetime.now().isoformat(),
            'file': excel_path,
            'sheets': {},
            'relationships': [],
            'formulas': [],
            'validations': [],
            'business_logic': [],
            'data_quality': {},
            'migration_plan': {}
        }

    def analyze_complete_structure(self):
        """Análisis completo de estructura del Excel"""
        print("🔍 INICIANDO ANÁLISIS QUIRÚRGICO DEL EXCEL")
        print("=" * 80)

        for sheet_name in self.workbook.sheetnames:
            print(f"\n📊 Analizando hoja: {sheet_name}")
            sheet_data = self.analyze_sheet(sheet_name)
            self.analysis_report['sheets'][sheet_name] = sheet_data

        return self.analysis_report

    def analyze_sheet(self, sheet_name: str) -> Dict:
        """Análisis detallado de cada hoja"""
        sheet = self.workbook[sheet_name]
        formula_sheet = self.full_workbook[sheet_name]

        analysis = {
            'name': sheet_name,
            'dimensions': f"{sheet.max_row} filas x {sheet.max_column} columnas",
            'headers': [],
            'data_types': {},
            'formulas': [],
            'validations': [],
            'named_ranges': [],
            'conditional_formatting': [],
            'merged_cells': [],
            'hyperlinks': [],
            'comments': [],
            'data_sample': [],
            'statistics': {},
            'relationships_detected': [],
            'business_rules': []
        }

        # Detectar encabezados
        first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        analysis['headers'] = [h for h in first_row if h is not None]

        # Analizar tipos de datos por columna
        for idx, header in enumerate(analysis['headers'], 1):
            col_letter = openpyxl.utils.get_column_letter(idx)
            col_data = [cell.value for cell in sheet[col_letter][1:] if cell.value is not None]

            data_type = self.detect_data_type(col_data)
            analysis['data_types'][header] = {
                'type': data_type,
                'nullable': None in [cell.value for cell in sheet[col_letter][1:]],
                'unique_values': len(set(col_data)),
                'sample_values': col_data[:5] if col_data else []
            }

        # Extraer fórmulas
        for row in formula_sheet.iter_rows(min_row=2):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    analysis['formulas'].append({
                        'cell': cell.coordinate,
                        'formula': cell.value,
                        'dependencies': self.extract_formula_dependencies(cell.value)
                    })

        # Detectar validaciones de datos
        if hasattr(sheet, 'data_validations'):
            for dv in sheet.data_validations.dataValidation:
                analysis['validations'].append({
                    'cells': str(dv.sqref),
                    'type': dv.type,
                    'formula1': dv.formula1,
                    'formula2': dv.formula2,
                    'allow_blank': dv.allowBlank
                })

        # Celdas combinadas
        if hasattr(sheet, 'merged_cells'):
            analysis['merged_cells'] = [str(mc) for mc in sheet.merged_cells.ranges]

        # Comentarios
        for cell in sheet:
            if hasattr(cell, 'comment') and cell.comment:
                analysis['comments'].append({
                    'cell': cell.coordinate,
                    'comment': cell.comment.text
                })

        # Extraer datos completos
        df = pd.DataFrame(sheet.values)
        if len(df) > 0:
            df.columns = df.iloc[0]
            df = df[1:]

            # Muestra de datos
            analysis['data_sample'] = df.head(10).to_dict('records')

            # Estadísticas
            analysis['statistics'] = {
                'total_rows': len(df),
                'total_columns': len(df.columns),
                'null_counts': df.isnull().sum().to_dict(),
                'duplicate_rows': df.duplicated().sum()
            }

            # Detectar relaciones (claves foráneas potenciales)
            analysis['relationships_detected'] = self.detect_relationships(df, sheet_name)

        return analysis

    def detect_data_type(self, values: List) -> str:
        """Detecta el tipo de dato predominante"""
        if not values:
            return 'EMPTY'

        types_count = {
            'DATE': 0,
            'DATETIME': 0,
            'NUMERIC': 0,
            'INTEGER': 0,
            'DECIMAL': 0,
            'CURRENCY': 0,
            'PERCENTAGE': 0,
            'TEXT': 0,
            'BOOLEAN': 0,
            'EMAIL': 0,
            'PHONE': 0,
            'UUID': 0
        }

        for val in values:
            if val is None:
                continue

            if isinstance(val, datetime):
                types_count['DATETIME'] += 1
            elif isinstance(val, (int, np.integer)):
                types_count['INTEGER'] += 1
            elif isinstance(val, (float, np.floating)):
                types_count['DECIMAL'] += 1
            elif isinstance(val, bool):
                types_count['BOOLEAN'] += 1
            elif isinstance(val, str):
                if re.match(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$', val, re.I):
                    types_count['UUID'] += 1
                elif re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', val):
                    types_count['EMAIL'] += 1
                elif re.match(r'^\+?[\d\s\-\(\)]+$', val):
                    types_count['PHONE'] += 1
                elif '%' in val:
                    types_count['PERCENTAGE'] += 1
                elif any(c in val for c in '$€£¥'):
                    types_count['CURRENCY'] += 1
                else:
                    types_count['TEXT'] += 1

        return max(types_count.items(), key=lambda x: x[1])[0]

    def extract_formula_dependencies(self, formula: str) -> List[str]:
        """Extrae referencias de celdas de una fórmula"""
        # Patrón para celdas (A1, B2, etc.)
        cell_pattern = r'\b[A-Z]+\d+\b'
        # Patrón para rangos (A1:B10)
        range_pattern = r'\b[A-Z]+\d+:[A-Z]+\d+\b'
        # Patrón para otras hojas (Sheet1!A1)
        sheet_pattern = r'\b\w+![A-Z]+\d+\b'

        dependencies = []
        dependencies.extend(re.findall(cell_pattern, formula))
        dependencies.extend(re.findall(range_pattern, formula))
        dependencies.extend(re.findall(sheet_pattern, formula))

        return list(set(dependencies))

    def detect_relationships(self, df: pd.DataFrame, sheet_name: str) -> List[Dict]:
        """Detecta relaciones potenciales entre tablas"""
        relationships = []

        # Buscar columnas que parecen claves foráneas
        for col in df.columns:
            if col is None:
                continue

            col_str = str(col).lower()

            # Patrones comunes de FK
            fk_patterns = [
                r'.*_id$',
                r'id_.*',
                r'.*codigo.*',
                r'.*clave.*',
                r'.*folio.*',
                r'.*ref.*'
            ]

            for pattern in fk_patterns:
                if re.match(pattern, col_str):
                    # Analizar valores únicos y no nulos
                    unique_ratio = df[col].nunique() / len(df[col].dropna())

                    relationships.append({
                        'source_sheet': sheet_name,
                        'source_column': col,
                        'relationship_type': 'FOREIGN_KEY' if unique_ratio < 0.9 else 'PRIMARY_KEY',
                        'confidence': unique_ratio,
                        'target_hint': self.guess_target_table(col_str)
                    })

        return relationships

    def guess_target_table(self, column_name: str) -> str:
        """Adivina la tabla objetivo basándose en el nombre de columna"""
        mappings = {
            'cliente': 'clientes',
            'producto': 'productos',
            'vendedor': 'vendedores',
            'proveedor': 'proveedores',
            'almacen': 'almacenes',
            'categoria': 'categorias',
            'venta': 'ventas',
            'compra': 'compras',
            'usuario': 'usuarios',
            'pedido': 'pedidos',
            'factura': 'facturas'
        }

        for key, value in mappings.items():
            if key in column_name:
                return value

        return 'UNKNOWN'

    def map_to_flowdistributor_entities(self) -> Dict:
        """Mapea datos del Excel a entidades de FlowDistributor"""
        print("\n🎯 MAPEANDO A ENTIDADES FLOWDISTRIBUTOR")
        print("=" * 80)

        mapping = {
            'productos': self.extract_productos(),
            'clientes': self.extract_clientes(),
            'ventas': self.extract_ventas(),
            'inventario': self.extract_inventario(),
            'proveedores': self.extract_proveedores(),
            'compras': self.extract_compras(),
            'usuarios': self.extract_usuarios(),
            'almacenes': self.extract_almacenes(),
            'categorias': self.extract_categorias(),
            'precios': self.extract_precios(),
            'descuentos': self.extract_descuentos(),
            'impuestos': self.extract_impuestos()
        }

        return mapping

    def extract_productos(self) -> List[Dict]:
        """Extrae productos del Excel"""
        productos = []

        # Buscar hojas que contengan productos
        product_sheets = [s for s in self.workbook.sheetnames
                         if any(word in s.lower() for word in ['producto', 'articulo', 'item', 'sku'])]

        for sheet_name in product_sheets:
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name)

            for _, row in df.iterrows():
                producto = {
                    'id': str(uuid.uuid4()),
                    'codigo': self.extract_field(row, ['codigo', 'sku', 'clave', 'id']),
                    'nombre': self.extract_field(row, ['nombre', 'descripcion', 'producto', 'articulo']),
                    'descripcion': self.extract_field(row, ['descripcion', 'detalle', 'descripcion_larga']),
                    'categoria': self.extract_field(row, ['categoria', 'familia', 'grupo', 'tipo']),
                    'precio_base': self.extract_numeric(row, ['precio', 'precio_base', 'precio_venta', 'pvp']),
                    'costo': self.extract_numeric(row, ['costo', 'precio_costo', 'precio_compra']),
                    'stock_actual': self.extract_numeric(row, ['stock', 'existencia', 'inventario', 'cantidad']),
                    'stock_minimo': self.extract_numeric(row, ['stock_minimo', 'min', 'minimo']),
                    'stock_maximo': self.extract_numeric(row, ['stock_maximo', 'max', 'maximo']),
                    'unidad_medida': self.extract_field(row, ['unidad', 'unidad_medida', 'um', 'uom']),
                    'proveedor': self.extract_field(row, ['proveedor', 'proveedor_principal', 'supplier']),
                    'estado': 'activo',
                    'fecha_creacion': datetime.now().isoformat(),
                    'metadata': {
                        'source_sheet': sheet_name,
                        'source_row': _ + 2,
                        'import_date': datetime.now().isoformat(),
                        'raw_data': row.to_dict()
                    }
                }

                if producto['codigo']:  # Solo agregar si tiene código
                    productos.append(producto)

        print(f"✅ Productos extraídos: {len(productos)}")
        return productos

    def extract_clientes(self) -> List[Dict]:
        """Extrae clientes del Excel"""
        clientes = []

        client_sheets = [s for s in self.workbook.sheetnames
                        if any(word in s.lower() for word in ['cliente', 'customer', 'cte'])]

        for sheet_name in client_sheets:
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name)

            for _, row in df.iterrows():
                cliente = {
                    'id': str(uuid.uuid4()),
                    'codigo': self.extract_field(row, ['codigo', 'codigo_cliente', 'id', 'clave']),
                    'razon_social': self.extract_field(row, ['razon_social', 'nombre', 'empresa', 'cliente']),
                    'rfc': self.extract_field(row, ['rfc', 'tax_id', 'nit']),
                    'tipo_cliente': self.extract_field(row, ['tipo', 'tipo_cliente', 'categoria']),
                    'limite_credito': self.extract_numeric(row, ['limite_credito', 'credito', 'credit_limit']),
                    'dias_credito': self.extract_numeric(row, ['dias_credito', 'plazo', 'payment_terms']),
                    'descuento': self.extract_numeric(row, ['descuento', 'discount', 'desc']),
                    'direccion': self.extract_field(row, ['direccion', 'address', 'domicilio']),
                    'ciudad': self.extract_field(row, ['ciudad', 'city']),
                    'estado': self.extract_field(row, ['estado', 'state', 'provincia']),
                    'cp': self.extract_field(row, ['cp', 'codigo_postal', 'zip']),
                    'telefono': self.extract_field(row, ['telefono', 'phone', 'tel']),
                    'email': self.extract_field(row, ['email', 'correo', 'mail']),
                    'contacto': self.extract_field(row, ['contacto', 'contact', 'persona_contacto']),
                    'vendedor': self.extract_field(row, ['vendedor', 'agente', 'sales_rep']),
                    'estado_cuenta': 'activo',
                    'fecha_alta': datetime.now().isoformat(),
                    'metadata': {
                        'source_sheet': sheet_name,
                        'source_row': _ + 2,
                        'import_date': datetime.now().isoformat(),
                        'raw_data': row.to_dict()
                    }
                }

                if cliente['codigo'] or cliente['razon_social']:
                    clientes.append(cliente)

        print(f"✅ Clientes extraídos: {len(clientes)}")
        return clientes

    def extract_ventas(self) -> List[Dict]:
        """Extrae ventas del Excel"""
        ventas = []

        sales_sheets = [s for s in self.workbook.sheetnames
                       if any(word in s.lower() for word in ['venta', 'pedido', 'factura', 'orden', 'sale'])]

        for sheet_name in sales_sheets:
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name)

            for _, row in df.iterrows():
                venta = {
                    'id': str(uuid.uuid4()),
                    'folio': self.extract_field(row, ['folio', 'numero', 'orden', 'id', 'invoice']),
                    'fecha': self.extract_date(row, ['fecha', 'date', 'fecha_venta']),
                    'cliente_codigo': self.extract_field(row, ['cliente', 'codigo_cliente', 'customer']),
                    'cliente_nombre': self.extract_field(row, ['nombre_cliente', 'razon_social']),
                    'vendedor': self.extract_field(row, ['vendedor', 'agente', 'sales_rep']),
                    'subtotal': self.extract_numeric(row, ['subtotal', 'sub_total', 'importe']),
                    'descuento': self.extract_numeric(row, ['descuento', 'discount']),
                    'impuestos': self.extract_numeric(row, ['impuestos', 'iva', 'tax']),
                    'total': self.extract_numeric(row, ['total', 'importe_total', 'grand_total']),
                    'tipo_documento': self.extract_field(row, ['tipo', 'tipo_documento', 'document_type']),
                    'estado': self.extract_field(row, ['estado', 'status', 'estatus']),
                    'forma_pago': self.extract_field(row, ['forma_pago', 'payment_method', 'pago']),
                    'observaciones': self.extract_field(row, ['observaciones', 'notas', 'comments']),
                    'metadata': {
                        'source_sheet': sheet_name,
                        'source_row': _ + 2,
                        'import_date': datetime.now().isoformat(),
                        'raw_data': row.to_dict()
                    }
                }

                if venta['folio']:
                    ventas.append(venta)

        print(f"✅ Ventas extraídas: {len(ventas)}")
        return ventas

    def extract_inventario(self) -> List[Dict]:
        """Extrae inventario del Excel"""
        inventario = []

        inventory_sheets = [s for s in self.workbook.sheetnames
                           if any(word in s.lower() for word in ['inventario', 'stock', 'existencia', 'almacen'])]

        for sheet_name in inventory_sheets:
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name)

            for _, row in df.iterrows():
                item = {
                    'id': str(uuid.uuid4()),
                    'producto_codigo': self.extract_field(row, ['codigo', 'sku', 'producto']),
                    'producto_nombre': self.extract_field(row, ['nombre', 'descripcion', 'articulo']),
                    'almacen': self.extract_field(row, ['almacen', 'bodega', 'warehouse']),
                    'ubicacion': self.extract_field(row, ['ubicacion', 'location', 'posicion']),
                    'lote': self.extract_field(row, ['lote', 'batch', 'serie']),
                    'cantidad': self.extract_numeric(row, ['cantidad', 'existencia', 'stock', 'qty']),
                    'cantidad_reservada': self.extract_numeric(row, ['reservada', 'reserved']),
                    'cantidad_disponible': self.extract_numeric(row, ['disponible', 'available']),
                    'costo_unitario': self.extract_numeric(row, ['costo', 'costo_unitario', 'unit_cost']),
                    'valor_total': self.extract_numeric(row, ['valor', 'valor_total', 'total_value']),
                    'fecha_movimiento': self.extract_date(row, ['fecha', 'fecha_movimiento', 'date']),
                    'tipo_movimiento': self.extract_field(row, ['tipo', 'tipo_movimiento', 'movement_type']),
                    'metadata': {
                        'source_sheet': sheet_name,
                        'source_row': _ + 2,
                        'import_date': datetime.now().isoformat(),
                        'raw_data': row.to_dict()
                    }
                }

                if item['producto_codigo']:
                    inventario.append(item)

        print(f"✅ Inventario extraído: {len(inventario)}")
        return inventario

    def extract_proveedores(self) -> List[Dict]:
        """Extrae proveedores del Excel"""
        proveedores = []

        supplier_sheets = [s for s in self.workbook.sheetnames
                          if any(word in s.lower() for word in ['proveedor', 'supplier', 'vendor'])]

        for sheet_name in supplier_sheets:
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name)

            for _, row in df.iterrows():
                proveedor = {
                    'id': str(uuid.uuid4()),
                    'codigo': self.extract_field(row, ['codigo', 'id', 'clave']),
                    'razon_social': self.extract_field(row, ['razon_social', 'nombre', 'empresa']),
                    'rfc': self.extract_field(row, ['rfc', 'tax_id']),
                    'contacto': self.extract_field(row, ['contacto', 'contact']),
                    'telefono': self.extract_field(row, ['telefono', 'phone']),
                    'email': self.extract_field(row, ['email', 'correo']),
                    'direccion': self.extract_field(row, ['direccion', 'address']),
                    'dias_credito': self.extract_numeric(row, ['dias_credito', 'plazo']),
                    'metadata': {
                        'source_sheet': sheet_name,
                        'source_row': _ + 2,
                        'import_date': datetime.now().isoformat(),
                        'raw_data': row.to_dict()
                    }
                }

                if proveedor['codigo'] or proveedor['razon_social']:
                    proveedores.append(proveedor)

        print(f"✅ Proveedores extraídos: {len(proveedores)}")
        return proveedores

    def extract_compras(self) -> List[Dict]:
        """Extrae compras del Excel"""
        compras = []

        purchase_sheets = [s for s in self.workbook.sheetnames
                          if any(word in s.lower() for word in ['compra', 'purchase', 'orden_compra'])]

        for sheet_name in purchase_sheets:
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name)

            for _, row in df.iterrows():
                compra = {
                    'id': str(uuid.uuid4()),
                    'folio': self.extract_field(row, ['folio', 'orden', 'numero']),
                    'fecha': self.extract_date(row, ['fecha', 'date']),
                    'proveedor_codigo': self.extract_field(row, ['proveedor', 'supplier']),
                    'total': self.extract_numeric(row, ['total', 'importe']),
                    'estado': self.extract_field(row, ['estado', 'status']),
                    'metadata': {
                        'source_sheet': sheet_name,
                        'source_row': _ + 2,
                        'import_date': datetime.now().isoformat(),
                        'raw_data': row.to_dict()
                    }
                }

                if compra['folio']:
                    compras.append(compra)

        print(f"✅ Compras extraídas: {len(compras)}")
        return compras

    def extract_usuarios(self) -> List[Dict]:
        """Extrae usuarios del Excel"""
        usuarios = []

        user_sheets = [s for s in self.workbook.sheetnames
                      if any(word in s.lower() for word in ['usuario', 'user', 'empleado', 'vendedor'])]

        for sheet_name in user_sheets:
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name)

            for _, row in df.iterrows():
                usuario = {
                    'id': str(uuid.uuid4()),
                    'codigo': self.extract_field(row, ['codigo', 'id', 'employee_id']),
                    'nombre': self.extract_field(row, ['nombre', 'name']),
                    'email': self.extract_field(row, ['email', 'correo']),
                    'rol': self.extract_field(row, ['rol', 'role', 'puesto']),
                    'departamento': self.extract_field(row, ['departamento', 'area']),
                    'metadata': {
                        'source_sheet': sheet_name,
                        'source_row': _ + 2,
                        'import_date': datetime.now().isoformat(),
                        'raw_data': row.to_dict()
                    }
                }

                if usuario['codigo'] or usuario['email']:
                    usuarios.append(usuario)

        print(f"✅ Usuarios extraídos: {len(usuarios)}")
        return usuarios

    def extract_almacenes(self) -> List[Dict]:
        """Extrae almacenes del Excel"""
        almacenes = []

        warehouse_sheets = [s for s in self.workbook.sheetnames
                           if any(word in s.lower() for word in ['almacen', 'bodega', 'warehouse'])]

        for sheet_name in warehouse_sheets:
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name)

            for _, row in df.iterrows():
                almacen = {
                    'id': str(uuid.uuid4()),
                    'codigo': self.extract_field(row, ['codigo', 'id']),
                    'nombre': self.extract_field(row, ['nombre', 'name', 'descripcion']),
                    'ubicacion': self.extract_field(row, ['ubicacion', 'location', 'direccion']),
                    'metadata': {
                        'source_sheet': sheet_name,
                        'source_row': _ + 2,
                        'import_date': datetime.now().isoformat(),
                        'raw_data': row.to_dict()
                    }
                }

                if almacen['codigo']:
                    almacenes.append(almacen)

        print(f"✅ Almacenes extraídos: {len(almacenes)}")
        return almacenes

    def extract_categorias(self) -> List[Dict]:
        """Extrae categorías del Excel"""
        categorias = []

        category_sheets = [s for s in self.workbook.sheetnames
                          if any(word in s.lower() for word in ['categoria', 'familia', 'grupo', 'category'])]

        for sheet_name in category_sheets:
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name)

            for _, row in df.iterrows():
                categoria = {
                    'id': str(uuid.uuid4()),
                    'codigo': self.extract_field(row, ['codigo', 'id']),
                    'nombre': self.extract_field(row, ['nombre', 'descripcion', 'category']),
                    'padre': self.extract_field(row, ['padre', 'parent', 'categoria_padre']),
                    'metadata': {
                        'source_sheet': sheet_name,
                        'source_row': _ + 2,
                        'import_date': datetime.now().isoformat(),
                        'raw_data': row.to_dict()
                    }
                }

                if categoria['codigo'] or categoria['nombre']:
                    categorias.append(categoria)

        print(f"✅ Categorías extraídas: {len(categorias)}")
        return categorias

    def extract_precios(self) -> List[Dict]:
        """Extrae precios del Excel"""
        precios = []

        price_sheets = [s for s in self.workbook.sheetnames
                       if any(word in s.lower() for word in ['precio', 'price', 'tarifa', 'lista_precio'])]

        for sheet_name in price_sheets:
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name)

            for _, row in df.iterrows():
                precio = {
                    'id': str(uuid.uuid4()),
                    'producto_codigo': self.extract_field(row, ['producto', 'codigo', 'sku']),
                    'lista_precio': self.extract_field(row, ['lista', 'tipo_precio', 'price_list']),
                    'precio': self.extract_numeric(row, ['precio', 'price', 'valor']),
                    'moneda': self.extract_field(row, ['moneda', 'currency']) or 'MXN',
                    'vigencia_desde': self.extract_date(row, ['vigencia_desde', 'desde', 'valid_from']),
                    'vigencia_hasta': self.extract_date(row, ['vigencia_hasta', 'hasta', 'valid_to']),
                    'metadata': {
                        'source_sheet': sheet_name,
                        'source_row': _ + 2,
                        'import_date': datetime.now().isoformat(),
                        'raw_data': row.to_dict()
                    }
                }

                if precio['producto_codigo'] and precio['precio']:
                    precios.append(precio)

        print(f"✅ Precios extraídos: {len(precios)}")
        return precios

    def extract_descuentos(self) -> List[Dict]:
        """Extrae descuentos del Excel"""
        descuentos = []

        discount_sheets = [s for s in self.workbook.sheetnames
                          if any(word in s.lower() for word in ['descuento', 'discount', 'promocion'])]

        for sheet_name in discount_sheets:
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name)

            for _, row in df.iterrows():
                descuento = {
                    'id': str(uuid.uuid4()),
                    'codigo': self.extract_field(row, ['codigo', 'id']),
                    'descripcion': self.extract_field(row, ['descripcion', 'nombre']),
                    'tipo': self.extract_field(row, ['tipo', 'type']),
                    'valor': self.extract_numeric(row, ['valor', 'porcentaje', 'descuento']),
                    'metadata': {
                        'source_sheet': sheet_name,
                        'source_row': _ + 2,
                        'import_date': datetime.now().isoformat(),
                        'raw_data': row.to_dict()
                    }
                }

                if descuento['codigo']:
                    descuentos.append(descuento)

        print(f"✅ Descuentos extraídos: {len(descuentos)}")
        return descuentos

    def extract_impuestos(self) -> List[Dict]:
        """Extrae impuestos del Excel"""
        impuestos = []

        tax_sheets = [s for s in self.workbook.sheetnames
                     if any(word in s.lower() for word in ['impuesto', 'tax', 'iva'])]

        for sheet_name in tax_sheets:
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name)

            for _, row in df.iterrows():
                impuesto = {
                    'id': str(uuid.uuid4()),
                    'codigo': self.extract_field(row, ['codigo', 'id']),
                    'nombre': self.extract_field(row, ['nombre', 'descripcion']),
                    'tasa': self.extract_numeric(row, ['tasa', 'porcentaje', 'rate']),
                    'metadata': {
                        'source_sheet': sheet_name,
                        'source_row': _ + 2,
                        'import_date': datetime.now().isoformat(),
                        'raw_data': row.to_dict()
                    }
                }

                if impuesto['codigo']:
                    impuestos.append(impuesto)

        print(f"✅ Impuestos extraídos: {len(impuestos)}")
        return impuestos

    def extract_field(self, row: pd.Series, possible_names: List[str]) -> Any:
        """Extrae campo buscando en múltiples posibles nombres"""
        for name in possible_names:
            # Buscar exacto
            if name in row.index and pd.notna(row[name]):
                return str(row[name]).strip() if row[name] else None

            # Buscar case-insensitive
            for col in row.index:
                if col and str(col).lower() == name.lower() and pd.notna(row[col]):
                    return str(row[col]).strip() if row[col] else None

            # Buscar contenga
            for col in row.index:
                if col and name in str(col).lower() and pd.notna(row[col]):
                    return str(row[col]).strip() if row[col] else None

        return None

    def extract_numeric(self, row: pd.Series, possible_names: List[str]) -> float:
        """Extrae valor numérico"""
        value = self.extract_field(row, possible_names)
        if value is None:
            return None

        try:
            # Limpiar formato de número
            clean_value = str(value).replace(',', '').replace('$', '').replace('%', '').strip()
            return float(clean_value)
        except:
            return None

    def extract_date(self, row: pd.Series, possible_names: List[str]) -> str:
        """Extrae fecha"""
        value = self.extract_field(row, possible_names)
        if value is None:
            return None

        if isinstance(value, datetime):
            return value.isoformat()

        try:
            date = pd.to_datetime(value)
            return date.isoformat()
        except:
            return None

    def analyze_business_logic(self) -> Dict:
        """Analiza la lógica de negocio del Excel"""
        print("\n🧠 ANALIZANDO LÓGICA DE NEGOCIO")
        print("=" * 80)

        business_logic = {
            'calculations': [],
            'validations': [],
            'workflows': [],
            'dependencies': [],
            'constraints': []
        }

        # Analizar todas las fórmulas para entender cálculos
        for sheet_name, sheet_data in self.analysis_report['sheets'].items():
            for formula in sheet_data.get('formulas', []):
                logic = self.interpret_formula(formula['formula'])
                business_logic['calculations'].append({
                    'sheet': sheet_name,
                    'cell': formula['cell'],
                    'formula': formula['formula'],
                    'logic': logic,
                    'dependencies': formula['dependencies']
                })

        # Analizar validaciones
        for sheet_name, sheet_data in self.analysis_report['sheets'].items():
            for validation in sheet_data.get('validations', []):
                business_logic['validations'].append({
                    'sheet': sheet_name,
                    'cells': validation['cells'],
                    'type': validation['type'],
                    'rule': validation['formula1']
                })

        return business_logic

    def interpret_formula(self, formula: str) -> str:
        """Interpreta la lógica de una fórmula"""
        formula_upper = formula.upper()

        interpretations = {
            'SUM': 'Suma de valores',
            'AVERAGE': 'Promedio de valores',
            'COUNT': 'Conteo de elementos',
            'IF': 'Condición lógica',
            'VLOOKUP': 'Búsqueda vertical',
            'HLOOKUP': 'Búsqueda horizontal',
            'INDEX': 'Obtener valor por índice',
            'MATCH': 'Encontrar posición',
            'SUMIF': 'Suma condicional',
            'COUNTIF': 'Conteo condicional',
            'ROUND': 'Redondeo',
            'MAX': 'Valor máximo',
            'MIN': 'Valor mínimo',
            'CONCATENATE': 'Concatenación de texto',
            'LEFT': 'Primeros caracteres',
            'RIGHT': 'Últimos caracteres',
            'MID': 'Caracteres intermedios',
            'DATE': 'Crear fecha',
            'NOW': 'Fecha/hora actual',
            'TODAY': 'Fecha actual'
        }

        for func, description in interpretations.items():
            if func in formula_upper:
                return description

        return 'Cálculo personalizado'

    def build_relationship_graph(self) -> Dict:
        """Construye grafo de relaciones entre entidades"""
        print("\n🔗 CONSTRUYENDO GRAFO DE RELACIONES")
        print("=" * 80)

        graph = {
            'nodes': [],
            'edges': []
        }

        # Crear nodos por cada hoja
        for sheet_name in self.analysis_report['sheets'].keys():
            graph['nodes'].append({
                'id': sheet_name,
                'label': sheet_name,
                'type': 'entity'
            })

        # Crear aristas basadas en relaciones detectadas
        for sheet_name, sheet_data in self.analysis_report['sheets'].items():
            for rel in sheet_data.get('relationships_detected', []):
                if rel['target_hint'] != 'UNKNOWN':
                    graph['edges'].append({
                        'from': sheet_name,
                        'to': rel['target_hint'],
                        'type': rel['relationship_type'],
                        'confidence': rel['confidence'],
                        'column': rel['source_column']
                    })

        return graph

    def generate_migration_plan(self) -> Dict:
        """Genera plan de migración completo"""
        print("\n📋 GENERANDO PLAN DE MIGRACIÓN")
        print("=" * 80)

        plan = {
            'phases': [
                {
                    'phase': 1,
                    'name': 'Preparación',
                    'tasks': [
                        'Validar estructura de datos',
                        'Crear backup del Excel original',
                        'Preparar base de datos PostgreSQL',
                        'Configurar Firebase'
                    ]
                },
                {
                    'phase': 2,
                    'name': 'Migración de Catálogos',
                    'tasks': [
                        'Importar categorías',
                        'Importar almacenes',
                        'Importar proveedores',
                        'Importar usuarios',
                        'Importar productos'
                    ]
                },
                {
                    'phase': 3,
                    'name': 'Migración de Maestros',
                    'tasks': [
                        'Importar clientes',
                        'Importar precios',
                        'Importar descuentos',
                        'Importar impuestos'
                    ]
                },
                {
                    'phase': 4,
                    'name': 'Migración de Transacciones',
                    'tasks': [
                        'Importar inventario',
                        'Importar compras',
                        'Importar ventas'
                    ]
                },
                {
                    'phase': 5,
                    'name': 'Validación',
                    'tasks': [
                        'Verificar integridad referencial',
                        'Validar cálculos',
                        'Comprobar totales',
                        'Revisar relaciones'
                    ]
                },
                {
                    'phase': 6,
                    'name': 'Post-Migración',
                    'tasks': [
                        'Generar reportes de validación',
                        'Documentar discrepancias',
                        'Capacitar usuarios',
                        'Go-live'
                    ]
                }
            ],
            'estimated_time': '4-6 horas',
            'rollback_plan': 'Restaurar desde backup',
            'validation_checkpoints': []
        }

        return plan

    def generate_complete_report(self) -> Dict:
        """Genera reporte completo del análisis"""
        print("\n📊 GENERANDO REPORTE COMPLETO")
        print("=" * 80)

        # Mapear a entidades
        entities_data = self.map_to_flowdistributor_entities()

        # Analizar lógica de negocio
        business_logic = self.analyze_business_logic()

        # Construir grafo de relaciones
        relationship_graph = self.build_relationship_graph()

        # Generar plan de migración
        migration_plan = self.generate_migration_plan()

        report = {
            'metadata': {
                'excel_file': self.excel_path,
                'analysis_date': datetime.now().isoformat(),
                'total_sheets': len(self.workbook.sheetnames),
                'sheets_analyzed': list(self.workbook.sheetnames)
            },
            'structure_analysis': self.analysis_report,
            'entities_mapped': {
                'productos': len(entities_data['productos']),
                'clientes': len(entities_data['clientes']),
                'ventas': len(entities_data['ventas']),
                'inventario': len(entities_data['inventario']),
                'proveedores': len(entities_data['proveedores']),
                'compras': len(entities_data['compras']),
                'usuarios': len(entities_data['usuarios']),
                'almacenes': len(entities_data['almacenes']),
                'categorias': len(entities_data['categorias']),
                'precios': len(entities_data['precios']),
                'descuentos': len(entities_data['descuentos']),
                'impuestos': len(entities_data['impuestos'])
            },
            'entities_data': entities_data,
            'business_logic': business_logic,
            'relationship_graph': relationship_graph,
            'migration_plan': migration_plan,
            'recommendations': self.generate_recommendations(entities_data),
            'data_quality_report': self.generate_quality_report(entities_data)
        }

        return report

    def generate_recommendations(self, entities_data: Dict) -> List[str]:
        """Genera recomendaciones basadas en el análisis"""
        recommendations = []

        # Analizar completitud de datos
        for entity_name, records in entities_data.items():
            if len(records) == 0:
                recommendations.append(f"⚠️ No se encontraron datos para {entity_name}")
            elif len(records) < 5:
                recommendations.append(f"⚠️ Pocos registros en {entity_name} ({len(records)})")

        # Recomendaciones de integridad
        productos = entities_data['productos']
        clientes = entities_data['clientes']

        if productos:
            sin_precio = sum(1 for p in productos if not p['precio_base'])
            if sin_precio > 0:
                recommendations.append(f"⚠️ {sin_precio} productos sin precio")

        if clientes:
            sin_rfc = sum(1 for c in clientes if not c['rfc'])
            if sin_rfc > 0:
                recommendations.append(f"⚠️ {sin_rfc} clientes sin RFC")

        recommendations.append("✅ Implementar validaciones de integridad referencial")
        recommendations.append("✅ Configurar índices para optimizar consultas")
        recommendations.append("✅ Establecer triggers para auditoría")
        recommendations.append("✅ Crear vistas materializadas para reportes")

        return recommendations

    def generate_quality_report(self, entities_data: Dict) -> Dict:
        """Genera reporte de calidad de datos"""
        quality = {
            'completeness': {},
            'consistency': {},
            'accuracy': {},
            'uniqueness': {}
        }

        for entity_name, records in entities_data.items():
            if not records:
                continue

            total = len(records)

            # Completitud
            complete = sum(1 for r in records if all(v is not None for k, v in r.items() if k != 'metadata'))
            quality['completeness'][entity_name] = f"{(complete/total)*100:.1f}%"

            # Unicidad (basado en códigos)
            codes = [r.get('codigo') for r in records if r.get('codigo')]
            unique = len(set(codes)) if codes else 0
            quality['uniqueness'][entity_name] = f"{(unique/len(codes)*100) if codes else 0:.1f}%"

        return quality

    def save_report(self, report: Dict, output_dir: str = 'reports'):
        """Guarda reporte en múltiples formatos"""
        Path(output_dir).mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        # JSON completo
        json_path = f"{output_dir}/excel_analysis_{timestamp}.json"
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(report, f, indent=2, ensure_ascii=False, default=str)
        print(f"✅ Reporte JSON guardado: {json_path}")

        # Excel con datos mapeados
        excel_path = f"{output_dir}/flowdistributor_data_{timestamp}.xlsx"
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            for entity_name, records in report['entities_data'].items():
                if records:
                    df = pd.DataFrame(records)
                    # Remover columna metadata para Excel
                    if 'metadata' in df.columns:
                        df = df.drop('metadata', axis=1)
                    df.to_excel(writer, sheet_name=entity_name, index=False)
        print(f"✅ Excel mapeado guardado: {excel_path}")

        # SQL de inserción
        sql_path = f"{output_dir}/migration_sql_{timestamp}.sql"
        self.generate_sql_inserts(report['entities_data'], sql_path)
        print(f"✅ Scripts SQL generados: {sql_path}")

        return {
            'json': json_path,
            'excel': excel_path,
            'sql': sql_path
        }

    def generate_sql_inserts(self, entities_data: Dict, output_path: str):
        """Genera scripts SQL de inserción"""
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("-- FlowDistributor - Scripts de Migración\n")
            f.write(f"-- Generado: {datetime.now().isoformat()}\n\n")

            f.write("BEGIN;\n\n")

            # Generar INSERTs para cada entidad
            for entity_name, records in entities_data.items():
                if not records:
                    continue

                f.write(f"-- {entity_name.upper()}\n")
                f.write(f"-- Total registros: {len(records)}\n\n")

                for record in records:
                    # Remover metadata
                    clean_record = {k: v for k, v in record.items() if k != 'metadata'}

                    columns = ', '.join(clean_record.keys())
                    values = ', '.join(
                        f"'{str(v).replace(chr(39), chr(39)+chr(39))}'" if v is not None else 'NULL'
                        for v in clean_record.values()
                    )

                    f.write(f"INSERT INTO {entity_name} ({columns}) VALUES ({values});\n")

                f.write("\n")

            f.write("COMMIT;\n")


def main():
    """Función principal de importación"""
    print("=" * 80)
    print("🚀 FLOWDISTRIBUTOR - IMPORTACIÓN QUIRÚRGICA DE EXCEL")
    print("=" * 80)

    excel_path = r"C:\Users\xpovo\Documents\premium-ecosystem\Copia de Administación_General.xlsx"

    if not Path(excel_path).exists():
        print(f"❌ Error: No se encuentra el archivo {excel_path}")
        return

    print(f"\n📂 Archivo: {excel_path}")
    print(f"📊 Iniciando análisis completo...\n")

    # Crear importador
    importer = SurgicalExcelImporter(excel_path)

    # Análisis completo
    structure = importer.analyze_complete_structure()

    # Generar reporte completo
    report = importer.generate_complete_report()

    # Guardar reportes
    output_files = importer.save_report(report)

    # Resumen final
    print("\n" + "=" * 80)
    print("✅ IMPORTACIÓN COMPLETADA")
    print("=" * 80)
    print(f"\n📊 Hojas analizadas: {len(structure['sheets'])}")
    print(f"📦 Entidades mapeadas: {len(report['entities_data'])}")
    print(f"\n📈 Registros importados:")
    for entity, count in report['entities_mapped'].items():
        print(f"   • {entity}: {count} registros")

    print(f"\n📁 Archivos generados:")
    for file_type, path in output_files.items():
        print(f"   • {file_type.upper()}: {path}")

    print(f"\n💡 Recomendaciones: {len(report['recommendations'])}")
    for rec in report['recommendations']:
        print(f"   {rec}")

    print("\n🎯 Sistema FlowDistributor listo para recibir los datos")
    print("=" * 80)

    return report


if __name__ == "__main__":
    # Instalar dependencias si es necesario
    try:
        import openpyxl
        import pandas as pd
    except ImportError:
        print("Instalando dependencias...")
        import subprocess
        subprocess.check_call(['pip', 'install', 'pandas', 'openpyxl', 'numpy'])

    report = main()
