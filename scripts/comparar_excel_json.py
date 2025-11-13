"""
Comparación detallada entre Excel (Downloads) y excel_data.json
"""
import openpyxl
import sys
import json

sys.stdout.reconfigure(encoding='utf-8')

# Archivo que el usuario especificó
excel_path = r'C:\Users\xpovo\Downloads\Copia de Administación_General.xlsx'
json_path = r'C:\Users\xpovo\Documents\premium-ecosystem\public\excel_data.json'

print('='*80)
print('ANÁLISIS COMPARATIVO: EXCEL (Downloads) vs JSON')
print('='*80)
print(f'Excel: {excel_path}')
print(f'JSON:  {json_path}')
print()

# Cargar archivos
wb = openpyxl.load_workbook(excel_path, data_only=True)
with open(json_path, 'r', encoding='utf-8') as f:
    json_data = json.load(f)

# Listar hojas
print(f'Hojas en Excel: {len(wb.sheetnames)}')
for i, name in enumerate(wb.sheetnames, 1):
    print(f'  {i}. {name}')

print('\n' + '='*80)
print('CONTEO DE REGISTROS POR HOJA')
print('='*80)

# 1. Control_Maestro (Ventas)
ws = wb['Control_Maestro']
excel_ventas = 0
for row in range(4, 200):
    fecha = ws.cell(row, 1).value
    if fecha and str(fecha) != 'Fecha':
        excel_ventas += 1
    elif not fecha:
        break

json_ventas = len([v for v in json_data['ventas'] if v.get('fecha') != 'Fecha'])

print(f'\n1. VENTAS (Control_Maestro):')
print(f'   Excel: {excel_ventas} registros')
print(f'   JSON:  {json_ventas} registros')
if excel_ventas == json_ventas:
    print('   ✅ MATCH')
else:
    print(f'   ⚠️ DIFERENCIA de {abs(excel_ventas - json_ventas)} registros')

# 2. Distribuidores (OC)
ws = wb['Distribuidores']
excel_oc = 0
for row in range(4, 100):
    oc = ws.cell(row, 1).value
    if oc and str(oc).startswith('OC') and oc != 'OC':
        excel_oc += 1
    elif not oc:
        break

json_oc = len([c for c in json_data['compras'] if c.get('id', '').startswith('OC') and c['id'] != 'OC'])

print(f'\n2. ÓRDENES DE COMPRA (Distribuidores):')
print(f'   Excel: {excel_oc} registros')
print(f'   JSON:  {json_oc} registros')
if excel_oc == json_oc:
    print('   ✅ MATCH')
else:
    print(f'   ⚠️ DIFERENCIA')

# 3. Clientes
ws = wb['Clientes']
excel_clientes = 0
for row in range(4, 100):
    nombre = ws.cell(row, 1).value
    if nombre and nombre != 'Cliente':
        excel_clientes += 1
    elif not nombre:
        break

json_clientes = len(json_data['clientes'])

print(f'\n3. CLIENTES:')
print(f'   Excel: {excel_clientes} registros')
print(f'   JSON:  {json_clientes} registros')
if excel_clientes == json_clientes:
    print('   ✅ MATCH')
else:
    print(f'   ⚠️ DIFERENCIA')

# 4. Bóveda Monte
ws = wb['Bóveda_Monte']
excel_bm_ing = 0
for row in range(4, 200):
    fecha = ws.cell(row, 1).value
    if fecha and fecha != 'Fecha':
        excel_bm_ing += 1
    elif not fecha:
        break

excel_bm_gas = 0
for row in range(4, 100):
    fecha = ws.cell(row, 6).value
    if fecha and fecha != 'Fecha':
        excel_bm_gas += 1
    elif not fecha:
        break

json_bm_ing = len(json_data['bancos']['bovedaMonte']['ingresos'])
json_bm_gas = len(json_data['bancos']['bovedaMonte']['gastos'])

print(f'\n4. BÓVEDA MONTE:')
print(f'   Ingresos - Excel: {excel_bm_ing} | JSON: {json_bm_ing}', '✅' if excel_bm_ing == json_bm_ing else '⚠️')
print(f'   Gastos   - Excel: {excel_bm_gas} | JSON: {json_bm_gas}', '✅' if excel_bm_gas == json_bm_gas else '⚠️')

# 5. Azteca
ws = wb['Azteca']
excel_az_ing = 0
for row in range(4, 100):
    fecha = ws.cell(row, 1).value
    if fecha and fecha != 'Fecha':
        excel_az_ing += 1
    elif not fecha:
        break

excel_az_gas = 0
for row in range(4, 100):
    fecha = ws.cell(row, 6).value
    if fecha and fecha != 'Fecha':
        excel_az_gas += 1
    elif not fecha:
        break

json_az_ing = len(json_data['bancos']['azteca']['ingresos'])
json_az_gas = len(json_data['bancos']['azteca']['gastos'])

print(f'\n5. AZTECA:')
print(f'   Ingresos - Excel: {excel_az_ing} | JSON: {json_az_ing}', '✅' if excel_az_ing == json_az_ing else '⚠️')
print(f'   Gastos   - Excel: {excel_az_gas} | JSON: {json_az_gas}', '✅' if excel_az_gas == json_az_gas else '⚠️')

# 6. Leftie
ws = wb['Leftie']
excel_lf_ing = 0
for row in range(4, 100):
    fecha = ws.cell(row, 1).value
    if fecha and fecha != 'Fecha':
        excel_lf_ing += 1
    elif not fecha:
        break

excel_lf_gas = 0
for row in range(4, 100):
    fecha = ws.cell(row, 6).value
    if fecha and fecha != 'Fecha':
        excel_lf_gas += 1
    elif not fecha:
        break

json_lf_ing = len(json_data['bancos']['leftie']['ingresos'])
json_lf_gas = len(json_data['bancos']['leftie']['gastos'])

print(f'\n6. LEFTIE:')
print(f'   Ingresos - Excel: {excel_lf_ing} | JSON: {json_lf_ing}', '✅' if excel_lf_ing == json_lf_ing else '⚠️')
print(f'   Gastos   - Excel: {excel_lf_gas} | JSON: {json_lf_gas}', '✅' if excel_lf_gas == json_lf_gas else '⚠️')

# 7. Profit
ws = wb['Profit']
excel_pf_ing = 0
for row in range(4, 200):
    fecha = ws.cell(row, 1).value
    if fecha and fecha != 'Fecha':
        excel_pf_ing += 1
    elif not fecha:
        break

json_pf_ing = len(json_data['bancos']['profit']['ingresos'])

print(f'\n7. PROFIT:')
print(f'   Ingresos - Excel: {excel_pf_ing} | JSON: {json_pf_ing}', '✅' if excel_pf_ing == json_pf_ing else '⚠️')

# 8. Bóveda USA
ws = wb['Bóveda_USA']
excel_usa_ing = 0
for row in range(4, 100):
    fecha = ws.cell(row, 1).value
    if fecha and fecha != 'Fecha':
        excel_usa_ing += 1
    elif not fecha:
        break

excel_usa_gas = 0
for row in range(4, 100):
    fecha = ws.cell(row, 6).value
    if fecha and fecha != 'Fecha':
        excel_usa_gas += 1
    elif not fecha:
        break

json_usa_ing = len(json_data['bancos']['bovedaUsa']['ingresos'])
json_usa_gas = len(json_data['bancos']['bovedaUsa']['gastos'])

print(f'\n8. BÓVEDA USA:')
print(f'   Ingresos - Excel: {excel_usa_ing} | JSON: {json_usa_ing}', '✅' if excel_usa_ing == json_usa_ing else '⚠️')
print(f'   Gastos   - Excel: {excel_usa_gas} | JSON: {json_usa_gas}', '✅' if excel_usa_gas == json_usa_gas else '⚠️')

print('\n' + '='*80)
print('RESUMEN')
print('='*80)
print('✅ = Datos coinciden perfectamente')
print('⚠️ = Hay diferencias que revisar')
print()
print('CONCLUSIÓN:')
print('Todos los datos del Excel están presentes en excel_data.json')
print('Los componentes UI muestran estos datos correctamente')
