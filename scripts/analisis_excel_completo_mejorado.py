"""
ANÁLISIS COMPLETO Y EXHAUSTIVO DEL EXCEL
Extrae TODAS las tablas de TODAS las hojas
Hoja por hoja, tabla por tabla, registro por registro
"""

import openpyxl
import json
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

# Ruta del Excel
excel_path = r'C:\Users\xpovo\Downloads\Copia de Administación_General.xlsx'

print('=' * 100)
print('ANÁLISIS EXHAUSTIVO DEL EXCEL - TODAS LAS HOJAS Y TABLAS')
print('=' * 100)
print(f'\nArchivo: {excel_path}')
print(f'Fecha de análisis: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')

# Cargar el Excel
wb = openpyxl.load_workbook(excel_path, data_only=True)

print(f'\n📚 HOJAS ENCONTRADAS: {len(wb.sheetnames)}')
for i, sheet_name in enumerate(wb.sheetnames, 1):
    print(f'  {i}. {sheet_name}')

# Resultado completo
resultado = {
    'metadata': {
        'archivo': excel_path,
        'fecha_analisis': datetime.now().isoformat(),
        'total_hojas': len(wb.sheetnames)
    },
    'hojas': {}
}

def obtener_valor(celda):
    """Obtiene el valor de una celda, manejando fechas y None"""
    if celda is None:
        return None
    if isinstance(celda, datetime):
        return celda.strftime('%Y-%m-%d')
    return celda

def extraer_tabla_completa(ws, col_inicio, fila_header, max_filas=500):
    """
    Extrae TODOS los datos de una tabla hasta encontrar filas vacías
    """
    # Leer headers
    headers = []
    col = col_inicio
    while col <= ws.max_column:
        header = ws.cell(fila_header, col).value
        if header is None or str(header).strip() == '':
            break
        headers.append(str(header))
        col += 1

    num_columnas = len(headers)
    if num_columnas == 0:
        return [], []

    # Leer datos
    datos = []
    for fila in range(fila_header + 1, fila_header + max_filas):
        fila_vacia = True
        fila_datos = {}

        for i, header in enumerate(headers):
            valor = ws.cell(fila, col_inicio + i).value
            if valor is not None and str(valor).strip() != '':
                fila_vacia = False
            fila_datos[header] = obtener_valor(valor)

        # Si todas las celdas están vacías, terminamos
        if fila_vacia:
            break

        datos.append(fila_datos)

    return headers, datos

# ==============================================================================
# ANÁLISIS HOJA POR HOJA
# ==============================================================================

print('\n' + '=' * 100)
print('EXTRACCIÓN COMPLETA DE DATOS')
print('=' * 100)

for sheet_name in wb.sheetnames:
    print(f'\n{"=" * 100}')
    print(f'📄 HOJA: {sheet_name}')
    print(f'{"=" * 100}')

    ws = wb[sheet_name]
    resultado['hojas'][sheet_name] = {'tablas': []}

    # Escanear la hoja para encontrar TODAS las tablas
    # Buscar en la fila 3 (típicamente donde están los headers)
    tablas_encontradas = []
    col_actual = 1

    while col_actual <= ws.max_column:
        header = ws.cell(3, col_actual).value

        if header and str(header).strip() != '':
            # Encontramos inicio de una tabla
            headers, datos = extraer_tabla_completa(ws, col_actual, 3)

            if len(datos) > 0:
                tabla_info = {
                    'columna_inicio': col_actual,
                    'headers': headers,
                    'num_registros': len(datos),
                    'datos': datos
                }
                tablas_encontradas.append(tabla_info)
                resultado['hojas'][sheet_name]['tablas'].append(tabla_info)

                print(f'\n📊 TABLA {len(tablas_encontradas)}:')
                print(f'   Columna inicial: {col_actual}')
                print(f'   Headers: {", ".join(headers)}')
                print(f'   Total registros: {len(datos)}')
                print(f'\n   Primeros 3 registros:')

                for i, registro in enumerate(datos[:3], 1):
                    print(f'   {i}. {registro}')

                if len(datos) > 3:
                    print(f'   ... y {len(datos) - 3} registros más')

                # Saltar las columnas de esta tabla
                col_actual += len(headers) + 1
            else:
                col_actual += 1
        else:
            col_actual += 1

    print(f'\n✅ Total de tablas encontradas en {sheet_name}: {len(tablas_encontradas)}')
    resultado['hojas'][sheet_name]['total_tablas'] = len(tablas_encontradas)

# ==============================================================================
# RESUMEN GENERAL
# ==============================================================================

print('\n' + '=' * 100)
print('📊 RESUMEN GENERAL')
print('=' * 100)

total_tablas = 0
total_registros = 0

for sheet_name, sheet_data in resultado['hojas'].items():
    num_tablas = sheet_data['total_tablas']
    num_registros = sum(tabla['num_registros'] for tabla in sheet_data['tablas'])
    total_tablas += num_tablas
    total_registros += num_registros

    print(f'\n📄 {sheet_name}:')
    print(f'   Tablas: {num_tablas}')
    print(f'   Registros totales: {num_registros}')

    for i, tabla in enumerate(sheet_data['tablas'], 1):
        print(f'      Tabla {i}: {tabla["num_registros"]} registros - Headers: {", ".join(tabla["headers"][:3])}{"..." if len(tabla["headers"]) > 3 else ""}')

print(f'\n{"=" * 100}')
print(f'📈 TOTALES GENERALES:')
print(f'   Total hojas analizadas: {len(resultado["hojas"])}')
print(f'   Total tablas encontradas: {total_tablas}')
print(f'   Total registros extraídos: {total_registros}')
print(f'{"=" * 100}')

# Guardar resultado completo en JSON
output_json = 'scripts/analisis_excel_completo_mejorado.json'
with open(output_json, 'w', encoding='utf-8') as f:
    json.dump(resultado, f, ensure_ascii=False, indent=2, default=str)

print(f'\n✅ Análisis completo guardado en: {output_json}')
print('\n' + '=' * 100)
print('FIN DEL ANÁLISIS EXHAUSTIVO')
print('=' * 100)
