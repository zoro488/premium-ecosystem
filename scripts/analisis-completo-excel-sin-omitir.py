"""
🔍 ANÁLISIS COMPLETO DE EXCEL - SIN OMITIR NINGÚN DATO
========================================================
Este script analiza TODAS las hojas del Excel y extrae ABSOLUTAMENTE TODOS los datos
sin omitir ninguna columna ni registro.

Funciones:
- Extrae TODAS las columnas de cada hoja
- Cuenta TODOS los registros
- Identifica columnas vacías vs con datos
- Genera reporte completo en JSON
- Compara con datos actuales en FlowDistributorData.js
"""

import openpyxl
import json
import sys
import os
from datetime import datetime

# Fix encoding para Windows
sys.stdout.reconfigure(encoding='utf-8')

# Rutas
EXCEL_PATH = r'C:\Users\xpovo\Documents\premium-ecosystem\Administación_General.xlsx'
DATA_JS_PATH = r'C:\Users\xpovo\Documents\premium-ecosystem\src\apps\FlowDistributor\data\FlowDistributorData.js'
OUTPUT_JSON = r'C:\Users\xpovo\Documents\premium-ecosystem\analisis_excel_completo_sin_omitir.json'

def leer_valor_celda(celda):
    """Lee el valor de una celda manejando todos los tipos de datos"""
    if celda.value is None:
        return None
    if isinstance(celda.value, (int, float)):
        return celda.value
    if isinstance(celda.value, datetime):
        return celda.value.strftime('%Y-%m-%d')
    return str(celda.value).strip()

def analizar_hoja_completa(ws, nombre_hoja):
    """Analiza una hoja completa extrayendo TODOS los datos"""
    print(f'\n📄 Analizando hoja: {nombre_hoja}')

    # Determinar rango de datos
    max_row = ws.max_row
    max_col = ws.max_column

    print(f'  📏 Dimensiones: {max_row} filas x {max_col} columnas')

    # Leer TODAS las cabeceras (fila 3)
    cabeceras = []
    for col_idx in range(1, max_col + 1):
        header = leer_valor_celda(ws.cell(3, col_idx))
        cabeceras.append(header if header else f'Columna_{col_idx}')

    print(f'  📋 Cabeceras encontradas: {len(cabeceras)}')
    for i, h in enumerate(cabeceras, 1):
        print(f'      {i}. {h}')

    # Leer TODOS los datos desde fila 4
    registros = []
    for row_idx in range(4, max_row + 1):
        # Verificar si la fila está vacía
        valores_fila = [leer_valor_celda(ws.cell(row_idx, col_idx)) for col_idx in range(1, max_col + 1)]

        # Si todos los valores son None, saltar fila
        if all(v is None for v in valores_fila):
            continue

        # Crear registro con TODOS los campos
        registro = {}
        for col_idx, header in enumerate(cabeceras, 1):
            valor = leer_valor_celda(ws.cell(row_idx, col_idx))
            registro[header] = valor

        registros.append(registro)

    print(f'  ✅ Registros extraídos: {len(registros)}')

    # Analizar columnas con datos vs vacías
    columnas_con_datos = {}
    columnas_vacias = []

    for header in cabeceras:
        valores_no_nulos = sum(1 for r in registros if r.get(header) is not None)
        if valores_no_nulos > 0:
            columnas_con_datos[header] = valores_no_nulos
        else:
            columnas_vacias.append(header)

    print(f'  📊 Columnas con datos: {len(columnas_con_datos)}')
    print(f'  🚫 Columnas vacías: {len(columnas_vacias)}')

    return {
        'nombre': nombre_hoja,
        'dimensiones': {
            'filas': max_row,
            'columnas': max_col
        },
        'cabeceras': cabeceras,
        'registros_totales': len(registros),
        'columnas_con_datos': columnas_con_datos,
        'columnas_vacias': columnas_vacias,
        'registros': registros
    }

def analizar_excel_completo():
    """Analiza TODAS las hojas del Excel sin omitir ningún dato"""
    print('='*80)
    print('🔍 ANÁLISIS COMPLETO DE EXCEL - SIN OMITIR NINGÚN DATO')
    print('='*80)

    try:
        # Cargar workbook
        print(f'\n📂 Cargando Excel: {EXCEL_PATH}')
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

        print(f'✅ Excel cargado exitosamente')
        print(f'📑 Hojas encontradas: {len(wb.sheetnames)}')
        for i, sheet in enumerate(wb.sheetnames, 1):
            print(f'   {i}. {sheet}')

        # Analizar TODAS las hojas
        analisis_completo = {
            'fecha_analisis': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'archivo': EXCEL_PATH,
            'total_hojas': len(wb.sheetnames),
            'hojas': {}
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            resultado = analizar_hoja_completa(ws, sheet_name)
            analisis_completo['hojas'][sheet_name] = resultado

        # Resumen general
        total_registros = sum(h['registros_totales'] for h in analisis_completo['hojas'].values())
        total_columnas = sum(len(h['cabeceras']) for h in analisis_completo['hojas'].values())

        analisis_completo['resumen'] = {
            'total_registros_todos_paneles': total_registros,
            'total_columnas_todos_paneles': total_columnas,
            'hojas_analizadas': len(analisis_completo['hojas'])
        }

        # Guardar en JSON
        print(f'\n💾 Guardando análisis completo en: {OUTPUT_JSON}')
        with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
            json.dump(analisis_completo, f, ensure_ascii=False, indent=2)

        print(f'✅ Análisis guardado exitosamente')

        # Mostrar resumen
        print('\n' + '='*80)
        print('📊 RESUMEN GENERAL')
        print('='*80)
        print(f'Total de hojas analizadas: {analisis_completo["total_hojas"]}')
        print(f'Total de registros (todos los paneles): {total_registros}')
        print(f'Total de columnas (todas las hojas): {total_columnas}')

        print('\n📋 DESGLOSE POR HOJA:')
        print('-'*80)
        for nombre, datos in analisis_completo['hojas'].items():
            print(f'\n📄 {nombre}')
            print(f'   Registros: {datos["registros_totales"]}')
            print(f'   Columnas totales: {len(datos["cabeceras"])}')
            print(f'   Columnas con datos: {len(datos["columnas_con_datos"])}')
            print(f'   Columnas vacías: {len(datos["columnas_vacias"])}')

            # Mostrar primeras 5 columnas con datos
            if datos["columnas_con_datos"]:
                print(f'   Principales columnas:')
                for col, count in list(datos["columnas_con_datos"].items())[:5]:
                    print(f'      • {col}: {count} registros')

        return analisis_completo

    except Exception as e:
        print(f'\n❌ Error durante el análisis: {e}')
        import traceback
        traceback.print_exc()
        return None

def comparar_con_flowdistributor_data(analisis):
    """Compara el análisis con los datos actuales en FlowDistributorData.js"""
    print('\n' + '='*80)
    print('🔍 COMPARACIÓN CON FLOWDISTRIBUTOR DATA')
    print('='*80)

    try:
        # Leer archivo JS
        with open(DATA_JS_PATH, 'r', encoding='utf-8') as f:
            contenido_js = f.read()

        # Mapeo de hojas Excel a exports JS
        mapeo_hojas = {
            'Distribuidores': 'DISTRIBUIDORES',
            'Ordenes de Compra': 'ORDENES_COMPRA',
            'Venta Local': 'VENTAS_LOCAL',
            'Almacen Monte': 'ALMACEN_MONTE',
            'Clientes': 'CLIENTES',
            'Boveda Monte': 'BOVEDA_MONTE',
            'Boveda Usa': 'BOVEDA_USA',
            'Flete Sur': 'FLETE_SUR',
            'Banco Azteca': 'AZTECA',
            'Banco Utilidades': 'UTILIDADES',
            'Banco Leftie': 'LEFTIE',
            'Banco Profit': 'PROFIT'
        }

        print('\n📊 VERIFICACIÓN DE EXPORTS:')
        print('-'*80)

        faltantes = []
        presentes = []

        for hoja_excel, export_js in mapeo_hojas.items():
            busqueda = f'export const {export_js} ='
            if busqueda in contenido_js:
                # Contar registros en Excel
                registros_excel = analisis['hojas'].get(hoja_excel, {}).get('registros_totales', 0)
                presentes.append({
                    'hoja': hoja_excel,
                    'export': export_js,
                    'registros_excel': registros_excel
                })
                print(f'✅ {export_js} - Presente ({registros_excel} registros en Excel)')
            else:
                faltantes.append({
                    'hoja': hoja_excel,
                    'export': export_js
                })
                print(f'❌ {export_js} - NO ENCONTRADO')

        # Resumen
        print('\n' + '='*80)
        print('📈 RESUMEN DE COMPARACIÓN:')
        print('='*80)
        print(f'Exports presentes: {len(presentes)}/{len(mapeo_hojas)}')
        print(f'Exports faltantes: {len(faltantes)}')

        if faltantes:
            print('\n⚠️  EXPORTS FALTANTES:')
            for item in faltantes:
                print(f'   • {item["export"]} (hoja: {item["hoja"]})')

        return {
            'presentes': presentes,
            'faltantes': faltantes,
            'total_verificado': len(presentes),
            'total_faltante': len(faltantes)
        }

    except Exception as e:
        print(f'\n❌ Error en comparación: {e}')
        return None

def generar_reporte_markdown(analisis, comparacion):
    """Genera un reporte en Markdown con TODOS los detalles"""
    reporte_path = r'C:\Users\xpovo\Documents\premium-ecosystem\REPORTE_DATOS_COMPLETO.md'

    with open(reporte_path, 'w', encoding='utf-8') as f:
        f.write('# 📊 REPORTE COMPLETO DE DATOS - SIN OMISIONES\n\n')
        f.write(f'**Fecha de análisis**: {analisis["fecha_analisis"]}\n\n')
        f.write('---\n\n')

        # Resumen ejecutivo
        f.write('## 🎯 RESUMEN EJECUTIVO\n\n')
        f.write(f'- **Total de hojas analizadas**: {analisis["total_hojas"]}\n')
        f.write(f'- **Total de registros**: {analisis["resumen"]["total_registros_todos_paneles"]}\n')
        f.write(f'- **Total de columnas**: {analisis["resumen"]["total_columnas_todos_paneles"]}\n')
        f.write(f'- **Exports presentes**: {comparacion["total_verificado"]}/{comparacion["total_verificado"] + comparacion["total_faltante"]}\n\n')

        f.write('---\n\n')

        # Detalles por hoja
        f.write('## 📋 DETALLES POR HOJA\n\n')
        for nombre, datos in analisis['hojas'].items():
            f.write(f'### 📄 {nombre}\n\n')
            f.write(f'**Estadísticas**:\n')
            f.write(f'- Registros totales: {datos["registros_totales"]}\n')
            f.write(f'- Columnas totales: {len(datos["cabeceras"])}\n')
            f.write(f'- Columnas con datos: {len(datos["columnas_con_datos"])}\n')
            f.write(f'- Columnas vacías: {len(datos["columnas_vacias"])}\n\n')

            f.write(f'**Columnas** ({len(datos["cabeceras"])}):\n')
            for i, col in enumerate(datos["cabeceras"], 1):
                registros_con_datos = datos["columnas_con_datos"].get(col, 0)
                if registros_con_datos > 0:
                    f.write(f'{i}. ✅ **{col}** - {registros_con_datos} registros\n')
                else:
                    f.write(f'{i}. 🚫 {col} - (vacía)\n')

            f.write('\n---\n\n')

        # Estado de exports
        f.write('## 🔍 ESTADO DE EXPORTS EN FLOWDISTRIBUTORDATA.JS\n\n')
        f.write('### ✅ Exports Presentes\n\n')
        for item in comparacion['presentes']:
            f.write(f'- **{item["export"]}** - {item["registros_excel"]} registros\n')

        if comparacion['faltantes']:
            f.write('\n### ❌ Exports Faltantes\n\n')
            for item in comparacion['faltantes']:
                f.write(f'- **{item["export"]}** (hoja: {item["hoja"]})\n')

        f.write('\n---\n\n')
        f.write('**Generado automáticamente** - ' + analisis["fecha_analisis"] + '\n')

    print(f'\n📄 Reporte Markdown generado: {reporte_path}')

if __name__ == '__main__':
    # Ejecutar análisis completo
    analisis = analizar_excel_completo()

    if analisis:
        # Comparar con FlowDistributorData.js
        comparacion = comparar_con_flowdistributor_data(analisis)

        if comparacion:
            # Generar reporte
            generar_reporte_markdown(analisis, comparacion)

            print('\n' + '='*80)
            print('✅ ANÁLISIS COMPLETO FINALIZADO')
            print('='*80)
            print(f'📄 JSON: analisis_excel_completo_sin_omitir.json')
            print(f'📄 Reporte: REPORTE_DATOS_COMPLETO.md')
            print('='*80)
