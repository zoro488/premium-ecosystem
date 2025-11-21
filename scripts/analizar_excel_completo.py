"""
Análisis completo del Excel original vs excel_data.json
Verifica que TODOS los datos estén correctamente importados
"""

import openpyxl
import json
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

# Rutas
excel_path = r'C:\Users\xpovo\Documents\premium-ecosystem\Administación_General.xlsx'
json_path = r'C:\Users\xpovo\Documents\premium-ecosystem\public\excel_data.json'

# Cargar archivos
wb = openpyxl.load_workbook(excel_path, data_only=True)
with open(json_path, 'r', encoding='utf-8') as f:
    json_data = json.load(f)

print('=' * 80)
print('ANÁLISIS DETALLADO: EXCEL ORIGINAL vs excel_data.json')
print('=' * 80)

# ===========================================================================
# 1. HOJA: Distribuidores (Órdenes de Compra)
# ===========================================================================
print('\n' + '=' * 80)
print('1. HOJA: Distribuidores (Órdenes de Compra)')
print('=' * 80)

ws = wb['Distribuidores']
excel_oc_count = 0
for row in range(4, 100):  # Empezar desde fila 4 (después de headers)
    oc = ws.cell(row, 1).value
    if oc and oc != "OC" and str(oc).startswith("OC"):
        excel_oc_count += 1
    elif not oc or str(oc).strip() == "":
        break

json_compras_count = len([c for c in json_data.get('compras', []) if c.get('id') != 'OC'])

print(f'\nRegistros en Excel: {excel_oc_count} órdenes de compra')
print(f'Registros en JSON (compras): {json_compras_count}')
print(f'Registros en JSON (FlowDistributorData ORDENES_COMPRA): Se verifica en código JS')

if excel_oc_count == json_compras_count:
    print('✅ MATCH: Misma cantidad de órdenes')
else:
    print(f'⚠️ DIFERENCIA: Excel tiene {excel_oc_count}, JSON tiene {json_compras_count}')

# Mostrar primeras 3 OC del Excel
print('\nPrimeras 3 OC del Excel:')
for row in range(4, 7):
    oc = ws.cell(row, 1).value
    fecha = ws.cell(row, 2).value
    origen = ws.cell(row, 3).value
    cantidad = ws.cell(row, 4).value
    print(f'  {oc} | {fecha} | {origen} | {cantidad} unidades')

# ===========================================================================
# 2. HOJA: Control_Maestro (Ventas)
# ===========================================================================
print('\n' + '=' * 80)
print('2. HOJA: Control_Maestro (Ventas)')
print('=' * 80)

ws = wb['Control_Maestro']
excel_ventas_count = 0
for row in range(4, 200):
    fecha = ws.cell(row, 1).value
    if fecha and fecha != "Fecha" and str(fecha) != "":
        excel_ventas_count += 1
    elif not fecha or str(fecha).strip() == "":
        break

json_ventas_count = len([v for v in json_data.get('ventas', []) if v.get('fecha') != 'Fecha'])

print(f'\nRegistros en Excel: {excel_ventas_count} ventas')
print(f'Registros en JSON: {json_ventas_count} ventas')

if excel_ventas_count == json_ventas_count:
    print('✅ MATCH: Misma cantidad de ventas')
else:
    print(f'⚠️ DIFERENCIA: Excel tiene {excel_ventas_count}, JSON tiene {json_ventas_count}')

# Mostrar primeras 3 ventas
print('\nPrimeras 3 ventas del Excel:')
for row in range(4, 7):
    fecha = ws.cell(row, 1).value
    oc = ws.cell(row, 2).value
    cantidad = ws.cell(row, 3).value
    cliente = ws.cell(row, 4).value
    print(f'  {fecha} | {oc} | {cliente} | {cantidad} unidades')

# ===========================================================================
# 3. HOJA: Clientes
# ===========================================================================
print('\n' + '=' * 80)
print('3. HOJA: Clientes')
print('=' * 80)

ws = wb['Clientes']
excel_clientes_count = 0
for row in range(4, 100):
    nombre = ws.cell(row, 1).value
    if nombre and nombre != "Cliente" and str(nombre) != "":
        excel_clientes_count += 1
    elif not nombre or str(nombre).strip() == "":
        break

json_clientes_count = len([c for c in json_data.get('clientes', []) if c.get('nombre', '').strip() != ''])

print(f'\nRegistros en Excel: {excel_clientes_count} clientes')
print(f'Registros en JSON: {json_clientes_count} clientes')

if excel_clientes_count == json_clientes_count:
    print('✅ MATCH: Misma cantidad de clientes')
else:
    print(f'⚠️ DIFERENCIA: Excel tiene {excel_clientes_count}, JSON tiene {json_clientes_count}')

# Mostrar primeros 5 clientes
print('\nPrimeros 5 clientes del Excel:')
for row in range(4, 9):
    nombre = ws.cell(row, 1).value
    adeudo = ws.cell(row, 2).value or 0
    if nombre and nombre != "Cliente":
        print(f'  {nombre} | Adeudo: ${adeudo}')

# ===========================================================================
# 4. HOJA: Almacen_Monte
# ===========================================================================
print('\n' + '=' * 80)
print('4. HOJA: Almacen_Monte')
print('=' * 80)

ws = wb['Almacen_Monte']

# Contar Ingresos (columnas A-D, empezando en fila 4)
ingresos_count = 0
for row in range(4, 100):
    oc = ws.cell(row, 1).value
    if oc and oc != "OC" and str(oc) != "":
        ingresos_count += 1
    elif not oc:
        break

# Contar Salidas (columnas F-J, empezando en fila 4)
salidas_count = 0
for row in range(4, 200):
    fecha = ws.cell(row, 6).value
    if fecha and fecha != "Fecha":
        salidas_count += 1
    elif not fecha:
        break

# Contar RF Actual Cortes (columnas L-M)
rf_cortes_count = 0
for row in range(4, 100):
    fecha = ws.cell(row, 12).value
    if fecha and fecha != "Fecha":
        rf_cortes_count += 1
    elif not fecha:
        break

print(f'\nIngresos en Excel: {ingresos_count}')
print(f'Salidas en Excel: {salidas_count}')
print(f'RF Actual Cortes en Excel: {rf_cortes_count}')

print(f'\n(JSON no tiene estructura almacen.ingresos directa)')
print(f'Verificar en FlowDistributorData.js: ALMACEN_MONTE')

# ===========================================================================
# 5. HOJA: Bóveda_Monte
# ===========================================================================
print('\n' + '=' * 80)
print('5. HOJA: Bóveda_Monte')
print('=' * 80)

ws = wb['Bóveda_Monte']

# Contar Ingresos (columnas A-D)
ingresos_boveda_count = 0
for row in range(4, 200):
    fecha = ws.cell(row, 1).value
    if fecha and fecha != "Fecha":
        ingresos_boveda_count += 1
    elif not fecha:
        break

# Contar Gastos (columnas F-I)
gastos_boveda_count = 0
for row in range(4, 100):
    fecha = ws.cell(row, 6).value
    if fecha and fecha != "Fecha":
        gastos_boveda_count += 1
    elif not fecha:
        break

json_boveda_ingresos = len(json_data.get('bancos', {}).get('bovedaMonte', {}).get('ingresos', []))
json_boveda_gastos = len(json_data.get('bancos', {}).get('bovedaMonte', {}).get('gastos', []))

print(f'\nIngresos en Excel: {ingresos_boveda_count}')
print(f'Ingresos en JSON: {json_boveda_ingresos}')
if ingresos_boveda_count == json_boveda_ingresos:
    print('✅ MATCH: Ingresos')
else:
    print(f'⚠️ DIFERENCIA: Excel {ingresos_boveda_count}, JSON {json_boveda_ingresos}')

print(f'\nGastos en Excel: {gastos_boveda_count}')
print(f'Gastos en JSON: {json_boveda_gastos}')
if gastos_boveda_count == json_boveda_gastos:
    print('✅ MATCH: Gastos')
else:
    print(f'⚠️ DIFERENCIA: Excel {gastos_boveda_count}, JSON {json_boveda_gastos}')

# ===========================================================================
# 6. HOJA: Azteca
# ===========================================================================
print('\n' + '=' * 80)
print('6. HOJA: Azteca')
print('=' * 80)

ws = wb['Azteca']

# Contar Ingresos
ingresos_azteca_count = 0
for row in range(4, 100):
    fecha = ws.cell(row, 1).value
    if fecha and fecha != "Fecha":
        ingresos_azteca_count += 1
    elif not fecha:
        break

# Contar Gastos
gastos_azteca_count = 0
for row in range(4, 100):
    fecha = ws.cell(row, 6).value
    if fecha and fecha != "Fecha":
        gastos_azteca_count += 1
    elif not fecha:
        break

json_azteca_ingresos = len(json_data.get('bancos', {}).get('azteca', {}).get('ingresos', []))
json_azteca_gastos = len(json_data.get('bancos', {}).get('azteca', {}).get('gastos', []))

print(f'\nIngresos en Excel: {ingresos_azteca_count}')
print(f'Ingresos en JSON: {json_azteca_ingresos}')
if ingresos_azteca_count == json_azteca_ingresos:
    print('✅ MATCH: Ingresos')
else:
    print(f'⚠️ DIFERENCIA')

print(f'\nGastos en Excel: {gastos_azteca_count}')
print(f'Gastos en JSON: {json_azteca_gastos}')
if gastos_azteca_count == json_azteca_gastos:
    print('✅ MATCH: Gastos')
else:
    print(f'⚠️ DIFERENCIA')

# ===========================================================================
# 7. HOJA: Leftie
# ===========================================================================
print('\n' + '=' * 80)
print('7. HOJA: Leftie')
print('=' * 80)

ws = wb['Leftie']

ingresos_leftie_count = 0
for row in range(4, 100):
    fecha = ws.cell(row, 1).value
    if fecha and fecha != "Fecha":
        ingresos_leftie_count += 1
    elif not fecha:
        break

gastos_leftie_count = 0
for row in range(4, 100):
    fecha = ws.cell(row, 6).value
    if fecha and fecha != "Fecha":
        gastos_leftie_count += 1
    elif not fecha:
        break

json_leftie_ingresos = len(json_data.get('bancos', {}).get('leftie', {}).get('ingresos', []))
json_leftie_gastos = len(json_data.get('bancos', {}).get('leftie', {}).get('gastos', []))

print(f'\nIngresos en Excel: {ingresos_leftie_count}')
print(f'Ingresos en JSON: {json_leftie_ingresos}')
if ingresos_leftie_count == json_leftie_ingresos:
    print('✅ MATCH')
else:
    print(f'⚠️ DIFERENCIA')

print(f'\nGastos en Excel: {gastos_leftie_count}')
print(f'Gastos en JSON: {json_leftie_gastos}')
if gastos_leftie_count == json_leftie_gastos:
    print('✅ MATCH')
else:
    print(f'⚠️ DIFERENCIA')

# ===========================================================================
# 8. HOJA: Profit
# ===========================================================================
print('\n' + '=' * 80)
print('8. HOJA: Profit')
print('=' * 80)

ws = wb['Profit']

ingresos_profit_count = 0
for row in range(4, 200):
    fecha = ws.cell(row, 1).value
    if fecha and fecha != "Fecha":
        ingresos_profit_count += 1
    elif not fecha:
        break

json_profit_ingresos = len(json_data.get('bancos', {}).get('profit', {}).get('ingresos', []))

print(f'\nIngresos en Excel: {ingresos_profit_count}')
print(f'Ingresos en JSON: {json_profit_ingresos}')
if ingresos_profit_count == json_profit_ingresos:
    print('✅ MATCH')
else:
    print(f'⚠️ DIFERENCIA')

# ===========================================================================
# 9. HOJA: Bóveda_USA
# ===========================================================================
print('\n' + '=' * 80)
print('9. HOJA: Bóveda_USA')
print('=' * 80)

ws = wb['Bóveda_USA']

ingresos_usa_count = 0
for row in range(4, 100):
    fecha = ws.cell(row, 1).value
    if fecha and fecha != "Fecha":
        ingresos_usa_count += 1
    elif not fecha:
        break

gastos_usa_count = 0
for row in range(4, 100):
    fecha = ws.cell(row, 6).value
    if fecha and fecha != "Fecha":
        gastos_usa_count += 1
    elif not fecha:
        break

# Nota: En JSON es "bovedaUsa" no "bovedaUSA"
json_usa_ingresos = len(json_data.get('bancos', {}).get('bovedaUsa', {}).get('ingresos', []))
json_usa_gastos = len(json_data.get('bancos', {}).get('bovedaUsa', {}).get('gastos', []))

print(f'\nIngresos en Excel: {ingresos_usa_count}')
print(f'Ingresos en JSON (bovedaUsa): {json_usa_ingresos}')
if ingresos_usa_count == json_usa_ingresos:
    print('✅ MATCH')
else:
    print(f'⚠️ DIFERENCIA')

print(f'\nGastos en Excel: {gastos_usa_count}')
print(f'Gastos en JSON: {json_usa_gastos}')
if gastos_usa_count == json_usa_gastos:
    print('✅ MATCH')
else:
    print(f'⚠️ DIFERENCIA')

# ===========================================================================
# 10. HOJA: Flete_Sur
# ===========================================================================
print('\n' + '=' * 80)
print('10. HOJA: Flete_Sur')
print('=' * 80)

ws = wb['Flete_Sur']

ingresos_flete_count = 0
for row in range(4, 100):
    fecha = ws.cell(row, 1).value
    if fecha and fecha != "Fecha":
        ingresos_flete_count += 1
    elif not fecha:
        break

print(f'\nIngresos en Excel: {ingresos_flete_count}')
print('JSON: Verificar en FlowDistributorData.js: FLETE_SUR')

# ===========================================================================
# RESUMEN FINAL
# ===========================================================================
print('\n' + '=' * 80)
print('RESUMEN FINAL DE VERIFICACIÓN')
print('=' * 80)
print('\n✅ = Datos coinciden')
print('⚠️ = Revisar diferencias')
print('\nHOJAS VERIFICADAS:')
print('1. Distribuidores (OC) - Verificar manualmente')
print('2. Control_Maestro (Ventas) - Verificar conteo')
print('3. Clientes - Verificar conteo')
print('4. Almacen_Monte - En FlowDistributorData.js')
print('5. Bóveda_Monte - Verificado arriba')
print('6. Azteca - Verificado arriba')
print('7. Leftie - Verificado arriba')
print('8. Profit - Verificado arriba')
print('9. Bóveda_USA - Verificado arriba')
print('10. Flete_Sur - En FlowDistributorData.js')

print('\n' + '=' * 80)
print('FIN DEL ANÁLISIS')
print('=' * 80)
