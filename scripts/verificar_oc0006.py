#!/usr/bin/env python3
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\xpovo\Documents\premium-ecosystem\Administación_General.xlsx', data_only=True)
ws = wb['Distribuidores']

print('REVISANDO OC0006 EN EXCEL')
print('='*80)

# Buscar OC0006
for row_idx in range(4, 20):
    oc = ws.cell(row_idx, 1).value
    if str(oc) == 'OC0006':
        print(f'\nFila {row_idx} - OC0006:')
        print(f'  OC: {ws.cell(row_idx, 1).value}')
        print(f'  Fecha: {ws.cell(row_idx, 2).value}')
        print(f'  Origen: {ws.cell(row_idx, 3).value}')
        print(f'  Cantidad: {ws.cell(row_idx, 4).value}')
        print(f'  Costo Distribuidor: {ws.cell(row_idx, 5).value}')
        print(f'  Costo Transporte: {ws.cell(row_idx, 6).value}')
        print(f'  Costo Por Unidad: {ws.cell(row_idx, 7).value}')
        print(f'  Stock Actual: {ws.cell(row_idx, 8).value}')
        print(f'  Costo Total (col 9): {ws.cell(row_idx, 9).value}')
        print(f'  Pago a Distribuidor (col 10): {ws.cell(row_idx, 10).value}')
        print(f'  Deuda (col 11): {ws.cell(row_idx, 11).value}')

        # Verificar lógica
        costo_total = ws.cell(row_idx, 9).value or 0
        pago = ws.cell(row_idx, 10).value or 0
        deuda_excel = ws.cell(row_idx, 11).value or 0

        print(f'\n  ANALISIS:')
        print(f'    Costo Total: ${costo_total:,.0f}')
        print(f'    Pago: ${pago:,.0f}')
        print(f'    Deuda en Excel: ${deuda_excel:,.0f}')
        print(f'    Deuda calculada (total-pago): ${costo_total-pago:,.0f}')

        if deuda_excel == 0 and pago == 0 and costo_total > 0:
            print(f'  PROBLEMA: OC pagada completamente deberia tener pago={costo_total}')
        elif deuda_excel == 0 and pago == costo_total:
            print(f'  OK: OC pagada completamente')
        elif deuda_excel > 0:
            print(f'  OK: OC con deuda pendiente')

        break
