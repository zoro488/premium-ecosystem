"""
🔍 ANÁLISIS DETALLADO COMPLETO DEL EXCEL
=========================================
Analiza TODAS las hojas y muestra datos reales línea por línea
"""

import openpyxl
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r'C:\Users\xpovo\Documents\premium-ecosystem\Administación_General.xlsx'

def leer_valor(celda):
    if celda.value is None:
        return None
    if isinstance(celda.value, (int, float)):
        return celda.value
    if isinstance(celda.value, datetime):
        return celda.value.strftime('%Y-%m-%d')
    return str(celda.value).strip()

def analizar_hoja_detallada(ws, nombre):
    """Analiza una hoja y muestra datos reales"""
    print(f'\n{"="*80}')
    print(f'📋 HOJA: {nombre}')
    print(f'{"="*80}')

    # Leer encabezados (fila 3)
    max_col = ws.max_column
    encabezados = []
    for col in range(1, max_col + 1):
        header = leer_valor(ws.cell(3, col))
        encabezados.append(header if header else f'Col_{col}')

    print(f'\nENCabezados ({len(encabezados)} columnas):')
    for i, h in enumerate(encabezados, 1):
        print(f'  {i}. {h}')

    # Leer datos (desde fila 4)
    registros_con_datos = 0
    registros_vacios = 0
    datos_muestra = []

    for row_idx in range(4, min(ws.max_row + 1, 25)):  # Primeras 20 filas
        tiene_datos = False
        registro = {}

        for col_idx, header in enumerate(encabezados, 1):
            valor = leer_valor(ws.cell(row_idx, col_idx))
            if valor:
                tiene_datos = True
            registro[header] = valor

        if tiene_datos:
            registros_con_datos += 1
            if len(datos_muestra) < 3:  # Guardar primeros 3
                datos_muestra.append((row_idx, registro))
        else:
            registros_vacios += 1
            if registros_vacios > 5:  # Si hay 5 filas vacías seguidas, parar
                break

    print(f'\n📊 ESTADÍSTICAS:')
    print(f'  Registros con datos: {registros_con_datos}')
    print(f'  Filas revisadas: {row_idx - 3}')

    if datos_muestra:
        print(f'\n📝 MUESTRA DE DATOS (primeros 3 registros):')
        for fila, datos in datos_muestra:
            print(f'\n  Fila {fila}:')
            for key, val in datos.items():
                if val:
                    print(f'    {key}: {val}')

    return registros_con_datos

print('='*80)
print('🔍 ANÁLISIS DETALLADO COMPLETO - TODAS LAS HOJAS')
print('='*80)

try:
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    hojas_analizar = [
        'Distribuidores',
        'Control_Maestro',
        'Almacen_Monte',
        'Bóveda_Monte',
        'Bóveda_USA',
        'Flete_Sur',
        'Utilidades',
        'Azteca',
        'Leftie',
        'Profit',
        'Clientes',
        'DATA'
    ]

    totales = {}

    for nombre_hoja in hojas_analizar:
        if nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
            total = analizar_hoja_detallada(ws, nombre_hoja)
            totales[nombre_hoja] = total
        else:
            print(f'\n⚠️  Hoja "{nombre_hoja}" no encontrada')

    print(f'\n{"="*80}')
    print('📊 RESUMEN GENERAL')
    print(f'{"="*80}')

    total_general = 0
    for hoja, total in totales.items():
        print(f'{hoja.ljust(25)} {str(total).rjust(5)} registros')
        total_general += total

    print(f'{"="*80}')
    print(f'TOTAL GENERAL'.ljust(25) + str(total_general).rjust(5))
    print(f'{"="*80}')

except Exception as e:
    print(f'\n❌ Error: {e}')
    import traceback
    traceback.print_exc()
