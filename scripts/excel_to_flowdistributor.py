#!/usr/bin/env python3
"""
Parser de Excel Administración General a formato FlowDistributor
Convierte todos los datos del Excel a JSON compatible con el sistema
"""

import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

def safe_value(cell):
    """Extrae valor de celda de forma segura"""
    if cell.value is None:
        return None
    if isinstance(cell.value, datetime):
        return cell.value.strftime('%Y-%m-%d')
    if isinstance(cell.value, (int, float)):
        return cell.value
    return str(cell.value).strip()

def parse_control_maestro(ws):
    """
    Parsea la hoja Control_Maestro - Ventas principales
    Estructura: Fecha | OC Relacionada | Cantidad | Cliente | Bóveda Monte | Precio De Venta |
                Ingreso | Flete | Flete Utilidad | Utilidad | Estatus | Concepto
    """
    ventas = []

    # Datos comienzan en fila 4 (después de headers en fila 3)
    for row_idx in range(4, ws.max_row + 1):
        fecha = safe_value(ws.cell(row_idx, 1))
        oc = safe_value(ws.cell(row_idx, 2))
        cantidad = safe_value(ws.cell(row_idx, 3))
        cliente = safe_value(ws.cell(row_idx, 4))
        costo_boveda = safe_value(ws.cell(row_idx, 5))
        precio_venta = safe_value(ws.cell(row_idx, 6))
        ingreso = safe_value(ws.cell(row_idx, 7))
        flete_aplica = safe_value(ws.cell(row_idx, 8))
        flete_utilidad = safe_value(ws.cell(row_idx, 9))
        utilidad = safe_value(ws.cell(row_idx, 10))
        estatus = safe_value(ws.cell(row_idx, 11))
        concepto = safe_value(ws.cell(row_idx, 12))

        # Solo procesar filas con datos válidos
        if not fecha or not cliente:
            continue

        # Convertir valores a números
        try:
            cantidad = float(cantidad) if cantidad else 0
            precio_venta = float(precio_venta) if precio_venta else 0
            costo_boveda = float(costo_boveda) if costo_boveda else 0
            flete_utilidad = float(flete_utilidad) if flete_utilidad else 0
            utilidad = float(utilidad) if utilidad else 0
            ingreso = float(ingreso) if ingreso else 0
        except:
            continue

        # Calcular valores
        total_venta = precio_venta * cantidad
        fletes = flete_utilidad if flete_aplica == "Aplica" else 0

        # Determinar estado de pago (LÓGICA DEL EXCEL)
        estado_pago = 'completo' if estatus == 'Pagado' else 'pendiente'
        estatus_texto = 'Pagado' if estatus == 'Pagado' else 'Pendiente'
        adeudo = 0 if estado_pago == 'completo' else total_venta
        monto_pagado = total_venta if estado_pago == 'completo' else 0

        venta = {
            'id': f'VENTA-{fecha}-{cliente}-{row_idx}',
            'fecha': fecha,
            'ocRelacionada': oc or 'N/A',
            'cliente': cliente,
            'cantidad': cantidad,
            'precioVenta': precio_venta,
            'totalVenta': total_venta,
            'costoBoveda': costo_boveda,
            'fletes': fletes,
            'utilidades': utilidad,
            # ⭐ CAMPOS NUEVOS DEL EXCEL
            'estadoPago': estado_pago,        # 'completo' | 'pendiente' (para compatibilidad)
            'estatus': estatus_texto,         # 'Pagado' | 'Pendiente' (usado en filtros)
            'destino': 'bovedaMonte',         # Banco donde se acredita el pago
            'montoPagado': monto_pagado,      # Monto ya pagado
            'adeudo': adeudo,                 # Saldo pendiente
            'concepto': concepto or '',
            'tipo': 'venta',                  # Tipo de transacción
            'productos': [{
                'nombre': f'Producto OC {oc}',
                'cantidad': cantidad,
                'precio': precio_venta
            }]
        }

        ventas.append(venta)

    return ventas

def parse_clientes(ws):
    """
    Parsea la hoja Clientes
    Estructura: Cliente | Actual | Deuda | Abonos | Pendiente | Observaciones
    """
    clientes = []

    # Datos comienzan en fila 4
    for row_idx in range(4, ws.max_row + 1):
        nombre = safe_value(ws.cell(row_idx, 5))  # Col E
        actual = safe_value(ws.cell(row_idx, 6))  # Col F
        deuda = safe_value(ws.cell(row_idx, 7))   # Col G
        abonos = safe_value(ws.cell(row_idx, 8))  # Col H
        pendiente = safe_value(ws.cell(row_idx, 9)) # Col I
        observaciones = safe_value(ws.cell(row_idx, 10)) # Col J

        if not nombre or (isinstance(nombre, str) and nombre.startswith('=')):
            continue

        # Convertir a string si es necesario
        nombre = str(nombre) if not isinstance(nombre, str) else nombre

        try:
            deuda = float(deuda) if deuda else 0
            abonos = float(abonos) if abonos else 0
            pendiente = float(pendiente) if pendiente else 0
        except:
            deuda = 0
            abonos = 0
            pendiente = 0

        cliente = {
            'id': f'CLI-{nombre}',
            'nombre': nombre,
            'adeudo': pendiente,
            'totalComprado': deuda,
            'totalAbonado': abonos,
            'estado': actual if actual else 'activo',
            'observaciones': observaciones or '',
            'ventas': []
        }

        clientes.append(cliente)

    return clientes

def parse_distribuidores(ws):
    """
    Parsea la hoja Distribuidores - Órdenes de Compra
    Estructura: OC | Fecha | Origen | Cantidad | Costo Distribuidor | Costo Transporte |
                Costo Por Unidad | Stock Actual | Costo Total | Pago a Distribuidor | Deuda
    """
    ordenes = []
    distribuidores_map = {}

    # Datos comienzan en fila 4
    for row_idx in range(4, ws.max_row + 1):
        oc = safe_value(ws.cell(row_idx, 1))
        fecha = safe_value(ws.cell(row_idx, 2))
        origen = safe_value(ws.cell(row_idx, 3))
        cantidad = safe_value(ws.cell(row_idx, 4))
        costo_dist = safe_value(ws.cell(row_idx, 5))
        costo_trans = safe_value(ws.cell(row_idx, 6))
        costo_unidad = safe_value(ws.cell(row_idx, 7))
        stock = safe_value(ws.cell(row_idx, 8))
        costo_total = safe_value(ws.cell(row_idx, 9))
        pago = safe_value(ws.cell(row_idx, 10))
        deuda = safe_value(ws.cell(row_idx, 11))

        if not oc or not origen:
            continue

        try:
            cantidad = float(cantidad) if cantidad else 0
            costo_dist = float(costo_dist) if costo_dist else 0
            costo_trans = float(costo_trans) if costo_trans else 0
            costo_total = float(costo_total) if costo_total else 0
            pago = float(pago) if pago else 0
            deuda = float(deuda) if deuda else 0
        except:
            continue

        # Calcular costo por unidad si no está
        if not costo_unidad and cantidad > 0:
            costo_unidad = costo_total / cantidad
        else:
            try:
                costo_unidad = float(costo_unidad) if costo_unidad else 0
            except:
                costo_unidad = 0

        # Calcular adeudo si no está especificado
        # Si deuda es 0 y pago también, asumir que se debe el total
        if deuda == 0 and pago == 0 and costo_total > 0:
            deuda = costo_total
        elif deuda == 0:
            # Si hay pago pero deuda es 0, calcular deuda
            deuda = max(0, costo_total - pago)

        orden = {
            'id': oc,
            'fecha': fecha or '',
            'distribuidor': origen,
            'cantidad': cantidad,
            'costoDistribuidor': costo_dist,
            'costoTransporte': costo_trans,
            'costoPorUnidad': costo_unidad,
            'costoTotal': costo_total,
            'pagado': pago,
            'adeudo': deuda,
            'productos': [{
                'nombre': f'Producto {oc}',
                'cantidad': cantidad,
                'precio': costo_unidad  # ✅ Corregido: precio en vez de costo
            }]
        }

        ordenes.append(orden)

        # Agregar al mapa de distribuidores
        if origen not in distribuidores_map:
            distribuidores_map[origen] = {
                'nombre': origen,
                'totalComprado': 0,
                'totalPagado': 0,
                'adeudo': 0,
                'ordenes': []
            }

        distribuidores_map[origen]['totalComprado'] += costo_total
        distribuidores_map[origen]['totalPagado'] += pago
        distribuidores_map[origen]['adeudo'] += deuda
        distribuidores_map[origen]['ordenes'].append(oc)

    distribuidores = list(distribuidores_map.values())

    return ordenes, distribuidores

def parse_almacen(ws, ordenes_compra=None, ventas=None):
    """
    Parsea la hoja Almacen_Monte
    Estructura: Ingresos (OC | Cliente | Distribuidor | Cantidad) y Salidas (Fecha | Cliente | Cantidad | Concepto)
    Enriquece entradas con datos de OCs y salidas con datos de ventas
    """
    entradas = []
    salidas = []
    stock = []

    # Crear mapas para búsqueda rápida
    oc_map = {oc['id']: oc for oc in (ordenes_compra or [])}

    # Ingresos - empiezan en fila 4, columnas A-D
    for row_idx in range(4, min(100, ws.max_row + 1)):  # Limitar a 100 para no procesar toda la hoja
        oc = safe_value(ws.cell(row_idx, 1))
        fecha = safe_value(ws.cell(row_idx, 2))
        distribuidor = safe_value(ws.cell(row_idx, 3))
        cantidad = safe_value(ws.cell(row_idx, 4))

        if not oc or (isinstance(oc, str) and oc.startswith('=')):
            continue

        oc = str(oc) if not isinstance(oc, str) else oc

        try:
            cantidad = float(cantidad) if cantidad else 0
        except:
            continue

        # Buscar datos de la OC relacionada
        oc_data = oc_map.get(oc, {})
        costo_unitario = oc_data.get('costoPorUnidad', 0)
        costo_total = cantidad * costo_unitario if costo_unitario else oc_data.get('costoTotal', 0)
        proveedor = oc_data.get('distribuidor', distribuidor or '')

        entrada = {
            'id': f'ENT-{oc}-{row_idx}',
            'fecha': fecha or '',
            'ocRelacionada': oc,
            'distribuidor': distribuidor or '',
            'proveedor': proveedor,
            'cantidad': cantidad,
            'costoUnitario': costo_unitario,
            'costoTotal': costo_total,
            'numeroFactura': oc,  # Usar OC como número de factura
            'nombre': f'Producto {oc}',
            'productos': [{
                'nombre': f'Producto {oc}',
                'cantidad': cantidad
            }]
        }

        entradas.append(entrada)

    # Salidas - columnas G-J
    for row_idx in range(4, min(100, ws.max_row + 1)):
        fecha = safe_value(ws.cell(row_idx, 7))
        cliente = safe_value(ws.cell(row_idx, 8))
        cantidad = safe_value(ws.cell(row_idx, 9))
        concepto = safe_value(ws.cell(row_idx, 10))

        if not cliente or (isinstance(cliente, str) and cliente.startswith('=')):
            continue

        cliente = str(cliente) if not isinstance(cliente, str) else cliente

        try:
            cantidad = float(cantidad) if cantidad else 0
        except:
            continue

        # Buscar venta relacionada para obtener precio y valor total
        venta_relacionada = None
        if ventas:
            for venta in ventas:
                if venta.get('cliente') == cliente and venta.get('fecha') == fecha:
                    venta_relacionada = venta
                    break

        precio_venta = venta_relacionada.get('precioVenta', 0) if venta_relacionada else 0
        valor_total = venta_relacionada.get('totalVenta', cantidad * precio_venta) if venta_relacionada else 0

        salida = {
            'id': f'SAL-{fecha}-{cliente}-{row_idx}',
            'fecha': fecha or '',
            'cliente': cliente,
            'cantidad': cantidad,
            'concepto': concepto or '',
            'motivoSalida': concepto or 'Venta',
            'precioVenta': precio_venta,
            'valorTotal': valor_total,
            'nombre': venta_relacionada.get('productos', [{}])[0].get('nombre', 'Producto General') if venta_relacionada else 'Producto General',
            'productos': [{
                'nombre': 'Producto General',
                'cantidad': cantidad
            }]
        }

        salidas.append(salida)

    return {
        'stock': stock,
        'entradas': entradas,
        'salidas': salidas
    }

def parse_banco(ws, banco_nombre):
    """
    Parsea hojas de bancos (Bóveda_Monte, Utilidades, Flete_Sur, etc.)
    Estructura: Ingresos (Fecha | Cliente | Ingreso | Concepto) y Gastos (Fecha | Origen | Gasto | TC | Pesos)
    """
    ingresos = []
    gastos = []
    transferencias = []

    # RF Actual está en E2
    try:
        rf_actual = ws.cell(2, 5).value
        rf_actual = float(rf_actual) if rf_actual else 0
    except:
        rf_actual = 0

    # Ingresos - columnas A-D, empiezan en fila 4
    for row_idx in range(4, min(200, ws.max_row + 1)):
        fecha = safe_value(ws.cell(row_idx, 1))
        cliente = safe_value(ws.cell(row_idx, 2))
        ingreso = safe_value(ws.cell(row_idx, 3))
        concepto = safe_value(ws.cell(row_idx, 4))

        if not fecha or (isinstance(fecha, str) and fecha.startswith('=')) or not ingreso:
            continue

        fecha = str(fecha) if not isinstance(fecha, str) else fecha
        cliente = str(cliente) if cliente and not isinstance(cliente, str) else cliente

        try:
            ingreso = float(ingreso)
        except:
            continue

        registro = {
            'id': f'ING-{banco_nombre}-{fecha}-{row_idx}',
            'fecha': fecha,
            'cliente': cliente or 'N/A',
            'monto': ingreso,  # ✅ Corregido: monto en vez de cantidad
            'concepto': concepto or '',
            'tipo': 'Ingreso'  # ✅ Corregido: Ingreso capitalizado
        }

        ingresos.append(registro)

    # Gastos - columnas G-J (o K dependiendo de la hoja)
    for row_idx in range(4, min(200, ws.max_row + 1)):
        fecha = safe_value(ws.cell(row_idx, 7))
        origen = safe_value(ws.cell(row_idx, 8))
        gasto = safe_value(ws.cell(row_idx, 9))
        tc = safe_value(ws.cell(row_idx, 10))
        pesos = safe_value(ws.cell(row_idx, 11))

        if not fecha or (isinstance(fecha, str) and fecha.startswith('=')) or not gasto:
            continue

        fecha = str(fecha) if not isinstance(fecha, str) else fecha
        origen = str(origen) if origen and not isinstance(origen, str) else origen

        try:
            gasto = float(gasto)
            tc = float(tc) if tc else 0
            pesos = float(pesos) if pesos else 0
        except:
            continue

        registro = {
            'id': f'GAS-{banco_nombre}-{fecha}-{row_idx}',
            'fecha': fecha,
            'cliente': origen or 'N/A',
            'monto': gasto,  # ✅ Corregido: monto en vez de cantidad
            'tc': tc,
            'pesos': pesos,
            'concepto': '',
            'tipo': 'Egreso'  # ✅ Corregido: Egreso capitalizado
        }

        gastos.append(registro)

    return {
        'capitalActual': rf_actual,
        'historico': rf_actual,
        'registros': ingresos + gastos,
        'ingresos': ingresos,
        'gastos': gastos,
        'transferencias': transferencias
    }

def main():
    """Función principal de conversión"""

    excel_path = r'C:\Users\xpovo\Documents\premium-ecosystem\Administación_General.xlsx'

    print("🚀 INICIANDO CONVERSIÓN DE EXCEL A FLOWDISTRIBUTOR")
    print("=" * 80)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        print(f"✅ Excel cargado: {len(wb.sheetnames)} hojas encontradas")

        # Estructura completa para FlowDistributor
        flow_data = {
            'ventas': [],
            'clientes': [],
            'ordenesCompra': [],
            'distribuidores': [],
            'almacen': {'stock': [], 'entradas': [], 'salidas': []},
            'bancos': {
                'bovedaMonte': None,
                'utilidades': None,
                'fletes': None,
                'azteca': None,
                'leftie': None,
                'profit': None
            }
        }

        # 1. Control_Maestro - Ventas
        if 'Control_Maestro' in wb.sheetnames:
            print("\n📊 Procesando Control_Maestro (Ventas)...")
            flow_data['ventas'] = parse_control_maestro(wb['Control_Maestro'])
            print(f"   ✓ {len(flow_data['ventas'])} ventas procesadas")

        # 2. Clientes
        if 'Clientes' in wb.sheetnames:
            print("\n👥 Procesando Clientes...")
            flow_data['clientes'] = parse_clientes(wb['Clientes'])
            print(f"   ✓ {len(flow_data['clientes'])} clientes procesados")

        # 3. Distribuidores y Órdenes de Compra
        if 'Distribuidores' in wb.sheetnames:
            print("\n📦 Procesando Distribuidores y Órdenes de Compra...")
            ordenes, distribuidores = parse_distribuidores(wb['Distribuidores'])
            flow_data['ordenesCompra'] = ordenes
            flow_data['distribuidores'] = distribuidores
            print(f"   ✓ {len(ordenes)} órdenes de compra procesadas")
            print(f"   ✓ {len(distribuidores)} distribuidores procesados")

        # 4. Almacén (enriquecer con datos de OCs y ventas)
        if 'Almacen_Monte' in wb.sheetnames:
            print("\n🏭 Procesando Almacén...")
            flow_data['almacen'] = parse_almacen(
                wb['Almacen_Monte'],
                ordenes_compra=flow_data['ordenesCompra'],
                ventas=flow_data['ventas']
            )
            print(f"   ✓ {len(flow_data['almacen']['entradas'])} entradas procesadas")
            print(f"   ✓ {len(flow_data['almacen']['salidas'])} salidas procesadas")

        # 5. Bancos
        bancos_map = {
            'Bóveda_Monte': 'bovedaMonte',
            'Utilidades': 'utilidades',
            'Flete_Sur': 'fletes',
            'Azteca': 'azteca',
            'Leftie': 'leftie',
            'Profit': 'profit'
        }

        print("\n💰 Procesando Bancos...")
        for sheet_name, key in bancos_map.items():
            if sheet_name in wb.sheetnames:
                flow_data['bancos'][key] = parse_banco(wb[sheet_name], key)
                print(f"   ✓ {key}: {len(flow_data['bancos'][key]['ingresos'])} ingresos, {len(flow_data['bancos'][key]['gastos'])} gastos")

        # Guardar JSON
        output_path = r'C:\Users\xpovo\Documents\premium-ecosystem\public\excel_data.json'
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(flow_data, f, ensure_ascii=False, indent=2)

        print("\n" + "=" * 80)
        print("✅ CONVERSIÓN COMPLETADA EXITOSAMENTE")
        print(f"📁 Archivo generado: {output_path}")
        print("\n📊 RESUMEN:")
        print(f"   • Ventas: {len(flow_data['ventas'])}")
        print(f"   • Clientes: {len(flow_data['clientes'])}")
        print(f"   • Órdenes de Compra: {len(flow_data['ordenesCompra'])}")
        print(f"   • Distribuidores: {len(flow_data['distribuidores'])}")
        print(f"   • Entradas Almacén: {len(flow_data['almacen']['entradas'])}")
        print(f"   • Salidas Almacén: {len(flow_data['almacen']['salidas'])}")

        total_bancos = sum(1 for v in flow_data['bancos'].values() if v is not None)
        print(f"   • Bancos configurados: {total_bancos}")

        print("\n🎯 El archivo JSON está listo para importarse en FlowDistributor")

    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()
