#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Análisis COMPLETO de TODAS las hojas del Excel y sus paneles
"""
import openpyxl
import json
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

# Cargar Excel
wb = openpyxl.load_workbook(r'C:\Users\xpovo\Documents\premium-ecosystem\Administación_General.xlsx', data_only=True)

print('=' * 120)
print('📊 ANÁLISIS COMPLETO: TODAS LAS HOJAS DEL EXCEL')
print('=' * 120)

print(f'\n📄 Total de hojas encontradas: {len(wb.sheetnames)}')
print('\nHojas disponibles:')
for idx, sheet_name in enumerate(wb.sheetnames, 1):
    print(f'   {idx}. {sheet_name}')

# Mapping de hojas a paneles
MAPPING_HOJAS_PANELES = {
    'Distribuidores': {
        'paneles': ['PanelDistribuidores', 'PanelOrdenesCompra'],
        'descripcion': 'Órdenes de compra y consolidado de distribuidores',
        'tablas_esperadas': 2
    },
    'Control_Maestro': {
        'paneles': ['PanelVentas', 'PanelGYA'],
        'descripcion': 'Ventas, Gastos y Abonos',
        'tablas_esperadas': 3
    },
    'Almacen_Monte': {
        'paneles': ['PanelAlmacen'],
        'descripcion': 'Inventario de Almacén Monte',
        'tablas_esperadas': 1
    },
    'Bóveda_Monte': {
        'paneles': ['PanelBovedaMonte'],
        'descripcion': 'Movimientos de Bóveda Monte',
        'tablas_esperadas': 1
    },
    'Bóveda_USA': {
        'paneles': ['PanelBovedaUSA'],
        'descripcion': 'Movimientos de Bóveda USA',
        'tablas_esperadas': 1
    },
    'Flete_Sur': {
        'paneles': ['PanelFleteSur'],
        'descripcion': 'Control de gastos de flete',
        'tablas_esperadas': 1
    },
    'Utilidades': {
        'paneles': ['PanelUtilidades'],
        'descripcion': 'Cálculo de utilidades',
        'tablas_esperadas': 1
    },
    'Azteca': {
        'paneles': ['PanelAzteca'],
        'descripcion': 'Movimientos Banco Azteca',
        'tablas_esperadas': 1
    },
    'Leftie': {
        'paneles': ['PanelLeftie'],
        'descripcion': 'Movimientos Banco Leftie',
        'tablas_esperadas': 1
    },
    'Profit': {
        'paneles': ['PanelProfit'],
        'descripcion': 'Análisis de rentabilidad',
        'tablas_esperadas': 1
    },
    'Clientes': {
        'paneles': ['PanelVentaLocal'],
        'descripcion': 'Clientes y ventas locales',
        'tablas_esperadas': 1
    }
}

def detectar_tablas(ws, max_filas=50):
    """Detecta tablas en una hoja buscando encabezados"""
    tablas = []

    for fila in range(1, min(max_filas, ws.max_row + 1)):
        # Leer toda la fila
        valores_fila = []
        for col in range(1, min(30, ws.max_column + 1)):
            valor = ws.cell(fila, col).value
            if valor and str(valor).strip():
                valores_fila.append({
                    'col': col,
                    'valor': str(valor).strip(),
                    'letra': openpyxl.utils.get_column_letter(col)
                })

        # Si hay al menos 3 celdas con contenido, podría ser encabezado
        if len(valores_fila) >= 3:
            # Verificar si parece encabezado
            es_encabezado = True
            for item in valores_fila:
                # Si tiene números muy grandes, probablemente no es encabezado
                try:
                    num = float(str(item['valor']).replace(',', '').replace('$', ''))
                    if num > 10000:
                        es_encabezado = False
                        break
                except:
                    pass

            if es_encabezado:
                tablas.append({
                    'fila_inicio': fila,
                    'columnas': valores_fila,
                    'num_columnas': len(valores_fila)
                })

    return tablas

def leer_datos_tabla(ws, fila_header, columnas_info, max_registros=200):
    """Lee los datos de una tabla"""
    headers = [col['valor'] for col in columnas_info]
    col_indices = [col['col'] for col in columnas_info]

    registros = []
    fila_actual = fila_header + 1
    filas_vacias = 0

    while filas_vacias < 3 and fila_actual < fila_header + max_registros:
        fila_datos = {}
        tiene_datos = False

        for i, col in enumerate(col_indices):
            valor = ws.cell(fila_actual, col).value
            if valor is not None:
                # Convertir datetime a string
                if isinstance(valor, datetime):
                    valor = valor.strftime('%Y-%m-%d %H:%M:%S')
                fila_datos[headers[i]] = valor
                tiene_datos = True

        if tiene_datos and len(fila_datos) >= len(headers) * 0.3:  # Al menos 30% de campos llenos
            registros.append(fila_datos)
            filas_vacias = 0
        else:
            filas_vacias += 1

        fila_actual += 1

    return registros

# Analizar todas las hojas
resultados_completos = {}

for hoja_nombre in wb.sheetnames:
    if hoja_nombre == 'DATA':
        continue  # Saltar hoja DATA por ahora

    print('\n' + '=' * 120)
    print(f'📄 ANALIZANDO: {hoja_nombre}')
    print('=' * 120)

    ws = wb[hoja_nombre]
    config = MAPPING_HOJAS_PANELES.get(hoja_nombre, {
        'paneles': ['Desconocido'],
        'descripcion': 'Hoja sin panel asignado',
        'tablas_esperadas': 1
    })

    print(f'\n📋 Descripción: {config["descripcion"]}')
    print(f'🎨 Paneles asociados: {", ".join(config["paneles"])}')
    print(f'📊 Tablas esperadas: {config["tablas_esperadas"]}')
    print(f'📐 Dimensiones: {ws.max_row} filas x {ws.max_column} columnas')

    # Detectar tablas
    tablas = detectar_tablas(ws, max_filas=50)
    print(f'\n🔍 Tablas encontradas: {len(tablas)}')

    resultados_hoja = {
        'nombre': hoja_nombre,
        'descripcion': config['descripcion'],
        'paneles_asociados': config['paneles'],
        'tablas_esperadas': config['tablas_esperadas'],
        'tablas_encontradas': len(tablas),
        'dimensiones': {'filas': ws.max_row, 'columnas': ws.max_column},
        'tablas': []
    }

    # Analizar cada tabla
    for idx, tabla in enumerate(tablas, 1):
        print(f'\n  📊 TABLA {idx} (Fila {tabla["fila_inicio"]}):')

        headers = [col['valor'] for col in tabla['columnas']]
        print(f'     Columnas ({len(headers)}): {", ".join(headers)}')

        # Leer datos
        registros = leer_datos_tabla(ws, tabla['fila_inicio'], tabla['columnas'])
        print(f'     Registros: {len(registros)}')

        # Mostrar primer registro
        if registros:
            print(f'\n     📝 Primer registro:')
            for key, val in list(registros[0].items())[:5]:
                print(f'        {key}: {val}')
            if len(registros[0]) > 5:
                print(f'        ... y {len(registros[0]) - 5} campos más')

        tabla_info = {
            'numero': idx,
            'fila_inicio': tabla['fila_inicio'],
            'headers': headers,
            'total_registros': len(registros),
            'columnas_detalle': tabla['columnas'],
            'registros_muestra': registros[:3]  # Solo primeros 3
        }

        resultados_hoja['tablas'].append(tabla_info)

    resultados_completos[hoja_nombre] = resultados_hoja

# Guardar resultados
with open('scripts/analisis-completo-todas-hojas.json', 'w', encoding='utf-8') as f:
    json.dump(resultados_completos, f, ensure_ascii=False, indent=2)

# RESUMEN FINAL
print('\n' + '=' * 120)
print('📊 RESUMEN EJECUTIVO')
print('=' * 120)

print(f'\n📄 Total hojas analizadas: {len(resultados_completos)}')

for hoja, datos in resultados_completos.items():
    estado = '✅' if datos['tablas_encontradas'] == datos['tablas_esperadas'] else '⚠️'
    print(f'\n{estado} {hoja}:')
    print(f'   Tablas esperadas: {datos["tablas_esperadas"]}, Encontradas: {datos["tablas_encontradas"]}')
    print(f'   Paneles: {", ".join(datos["paneles_asociados"])}')

    for tabla in datos['tablas']:
        print(f'   • Tabla {tabla["numero"]}: {len(tabla["headers"])} columnas, {tabla["total_registros"]} registros')

print('\n💾 Resultados guardados en: scripts/analisis-completo-todas-hojas.json')
print('\n' + '=' * 120 + '\n')
