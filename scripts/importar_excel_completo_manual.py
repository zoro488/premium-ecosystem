#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SCRIPT DE IMPORTACIÓN MANUAL Y PRECISO
Importa datos del Excel "Administación_General.xlsx" a FlowDistributor

Autor: Claude Code
Fecha: 2025-10-20
Basado en: MAPEO_COMPLETO_EXCEL_IMPORTACION.md
"""

import openpyxl
import json
import sys
from datetime import datetime
from pathlib import Path

# Configuración de encoding
sys.stdout.reconfigure(encoding='utf-8')

# Rutas
EXCEL_PATH = r'C:\Users\xpovo\Downloads\Copia de Administación_General.xlsx'
OUTPUT_PATH = r'C:\Users\xpovo\Documents\premium-ecosystem\public\excel_data.json'

# Contadores globales para estadísticas
stats = {
    'ventas_pagadas': 0,
    'ventas_pendientes': 0,
    'ordenes_compra': 0,
    'clientes': 0,
    'ingresos_boveda': 0,
    'gastos': 0,
    'entradas_almacen': 0,
    'salidas_almacen': 0
}

def convertir_fecha(fecha_excel):
    """Convierte fecha de Excel a ISO string"""
    if fecha_excel is None:
        return None
    if isinstance(fecha_excel, datetime):
        return fecha_excel.isoformat()
    if isinstance(fecha_excel, str):
        try:
            return datetime.fromisoformat(fecha_excel).isoformat()
        except:
            return None
    return None

def limpiar_valor(valor, tipo='numero'):
    """Limpia y convierte valores del Excel"""
    if valor is None:
        return 0 if tipo == 'numero' else ''
    if tipo == 'numero':
        try:
            return float(valor) if valor != '' else 0
        except:
            return 0
    elif tipo == 'texto':
        return str(valor).strip() if valor != '' else ''
    return valor

print("="*100)
print("IMPORTACIÓN MANUAL Y PRECISA DE EXCEL A FLOWDISTRIBUTOR")
print("="*100)

# Cargar workbook
print(f"\n📂 Cargando archivo: {EXCEL_PATH}")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
print(f"✅ Archivo cargado. Hojas encontradas: {wb.sheetnames}")

# Estructura de datos final
data = {
    "ventas": [],
    "clientes": [],
    "ordenesCompra": [],
    "distribuidores": [],
    "almacen": {
        "stock": [],
        "entradas": [],
        "salidas": []
    },
    "bancos": {
        "bovedaMonte": {
            "nombre": "Bóveda Monte",
            "capitalActual": 0,
            "moneda": "MXN",
            "registros": [],
            "ingresos": [],
            "egresos": [],
            "historico": 0
        },
        "bovedaUSA": {
            "nombre": "Bóveda USA",
            "capitalActual": 0,
            "moneda": "USD",
            "registros": [],
            "ingresos": [],
            "egresos": [],
            "historico": 0
        },
        "utilidades": {
            "nombre": "Utilidades",
            "capitalActual": 0,
            "moneda": "MXN",
            "registros": [],
            "ingresos": [],
            "egresos": [],
            "historico": 0
        },
        "fleteSur": {
            "nombre": "Flete Sur",
            "capitalActual": 0,
            "moneda": "MXN",
            "registros": [],
            "ingresos": [],
            "egresos": [],
            "historico": 0
        },
        "azteca": {
            "nombre": "Azteca",
            "capitalActual": 0,
            "moneda": "MXN",
            "registros": [],
            "ingresos": [],
            "egresos": [],
            "historico": 0
        },
        "leftie": {
            "nombre": "Leftie",
            "capitalActual": 0,
            "moneda": "USD",
            "registros": [],
            "ingresos": [],
            "egresos": [],
            "historico": 0
        },
        "profit": {
            "nombre": "Profit",
            "capitalActual": 0,
            "moneda": "MXN",
            "registros": [],
            "ingresos": [],
            "egresos": [],
            "historico": 0
        }
    },
    "gastosAbonos": []
}

# =============================================================================
# 1. IMPORTAR DISTRIBUIDORES (Órdenes de Compra)
# =============================================================================
print("\n" + "="*100)
print("1️⃣  IMPORTANDO DISTRIBUIDORES (Órdenes de Compra)")
print("="*100)

ws_dist = wb['Distribuidores']

distribuidores_unicos = {}

for row in range(4, 100):
    oc = ws_dist.cell(row, 1).value
    if not oc:
        continue

    fecha = ws_dist.cell(row, 2).value
    origen = limpiar_valor(ws_dist.cell(row, 3).value, 'texto')
    cantidad = limpiar_valor(ws_dist.cell(row, 4).value)
    costo_dist = limpiar_valor(ws_dist.cell(row, 5).value)
    costo_trans = limpiar_valor(ws_dist.cell(row, 6).value)
    costo_unitario = limpiar_valor(ws_dist.cell(row, 7).value)
    costo_total = limpiar_valor(ws_dist.cell(row, 9).value)
    pago_dist = limpiar_valor(ws_dist.cell(row, 10).value)
    deuda = limpiar_valor(ws_dist.cell(row, 11).value)

    orden_compra = {
        "id": oc,
        "fecha": convertir_fecha(fecha),
        "distribuidor": origen,
        "cantidad": cantidad,
        "costoDistribuidor": costo_dist,
        "costoTransporte": costo_trans,
        "costoUnitario": costo_unitario,
        "costoTotal": costo_total,
        "pagado": pago_dist,
        "adeudo": deuda if deuda else (costo_total - pago_dist if pago_dist else costo_total),
        "estado": "pagado" if (deuda == 0 or not deuda) and pago_dist else "pendiente"
    }

    data["ordenesCompra"].append(orden_compra)
    stats['ordenes_compra'] += 1

    # Rastrear distribuidores únicos
    if origen not in distribuidores_unicos:
        distribuidores_unicos[origen] = {
            "nombre": origen,
            "totalComprado": 0,
            "totalPagado": 0,
            "adeudo": 0,
            "ordenes": []
        }

    distribuidores_unicos[origen]["totalComprado"] += costo_total
    distribuidores_unicos[origen]["totalPagado"] += pago_dist
    distribuidores_unicos[origen]["adeudo"] += orden_compra["adeudo"]
    distribuidores_unicos[origen]["ordenes"].append(oc)

    print(f"  ✓ {oc} - {origen} - {cantidad} unidades - ${costo_total:,.0f}")

# Agregar distribuidores únicos
data["distribuidores"] = list(distribuidores_unicos.values())

print(f"\n📊 Órdenes de Compra importadas: {stats['ordenes_compra']}")
print(f"📊 Distribuidores únicos: {len(data['distribuidores'])}")

# =============================================================================
# 2. IMPORTAR ALMACEN - ENTRADAS (desde Órdenes de Compra)
# =============================================================================
print("\n" + "="*100)
print("2️⃣  IMPORTANDO ALMACEN - ENTRADAS")
print("="*100)

ws_almacen = wb['Almacen_Monte']

for row in range(4, 100):
    oc = ws_almacen.cell(row, 1).value
    if not oc:
        continue

    distribuidor = limpiar_valor(ws_almacen.cell(row, 3).value, 'texto')
    cantidad = limpiar_valor(ws_almacen.cell(row, 4).value)
    fecha = ws_almacen.cell(row, 5).value

    entrada = {
        "id": f"ENTRADA-{oc}",
        "tipo": "entrada",
        "fecha": convertir_fecha(fecha),
        "oc": oc,
        "distribuidor": distribuidor,
        "cantidad": cantidad,
        "concepto": f"Entrada de mercancía - {oc}"
    }

    data["almacen"]["entradas"].append(entrada)
    stats['entradas_almacen'] += 1

    print(f"  ✓ Entrada {oc} - {distribuidor} - {cantidad} unidades")

print(f"\n📊 Entradas de almacén importadas: {stats['entradas_almacen']}")

# =============================================================================
# 3. IMPORTAR CLIENTES (Lista básica)
# =============================================================================
print("\n" + "="*100)
print("3️⃣  IMPORTANDO CLIENTES")
print("="*100)

ws_clientes = wb['Clientes']

clientes_map = {}

for row in range(4, 50):
    nombre = limpiar_valor(ws_clientes.cell(row, 5).value, 'texto')
    if not nombre:
        continue

    observaciones = limpiar_valor(ws_clientes.cell(row, 10).value, 'texto')

    cliente = {
        "nombre": nombre,
        "adeudo": 0,  # Se calculará dinámicamente
        "totalComprado": 0,
        "totalAbonado": 0,
        "estado": "activo",
        "observaciones": observaciones,
        "ventas": []
    }

    data["clientes"].append(cliente)
    clientes_map[nombre] = cliente
    stats['clientes'] += 1

    print(f"  ✓ {nombre}")

print(f"\n📊 Clientes importados: {stats['clientes']}")

# =============================================================================
# 4. IMPORTAR VENTAS (Control_Maestro)
# =============================================================================
print("\n" + "="*100)
print("4️⃣  IMPORTANDO VENTAS (Control_Maestro)")
print("="*100)

ws_control = wb['Control_Maestro']

ventas_por_cliente = {}

for row in range(4, 200):
    fecha = ws_control.cell(row, 1).value
    if not fecha:
        continue

    oc_relacionada = limpiar_valor(ws_control.cell(row, 2).value, 'texto')
    cantidad = limpiar_valor(ws_control.cell(row, 3).value)
    cliente = limpiar_valor(ws_control.cell(row, 4).value, 'texto')
    boveda_monte = limpiar_valor(ws_control.cell(row, 5).value)
    precio_venta = limpiar_valor(ws_control.cell(row, 6).value)
    ingreso = limpiar_valor(ws_control.cell(row, 7).value)
    flete = limpiar_valor(ws_control.cell(row, 8).value, 'texto')
    flete_utilidad = limpiar_valor(ws_control.cell(row, 9).value)
    utilidad = limpiar_valor(ws_control.cell(row, 10).value)
    estatus = limpiar_valor(ws_control.cell(row, 11).value, 'texto')
    concepto = limpiar_valor(ws_control.cell(row, 12).value, 'texto')

    venta_id = f"VENTA-{convertir_fecha(fecha)}-{cliente}-{row}"

    venta = {
        "id": venta_id,
        "tipo": "venta",
        "fecha": convertir_fecha(fecha),
        "ocRelacionada": oc_relacionada,
        "cantidad": cantidad,
        "cliente": cliente,
        "productos": [{
            "nombre": "Producto",
            "cantidad": cantidad,
            "precio": precio_venta,
            "subtotal": ingreso
        }],

        # Montos
        "totalVenta": ingreso,
        "totalFletes": flete_utilidad,
        "totalUtilidades": utilidad,

        # Estado de pago (⭐ CRÍTICO)
        "estatus": estatus,
        "estadoPago": "completo" if estatus == "Pagado" else "pendiente",
        "adeudo": 0 if estatus == "Pagado" else boveda_monte,
        "montoPagado": boveda_monte if estatus == "Pagado" else 0,

        # Destino
        "destino": "bovedaMonte",
        "concepto": concepto,

        # Flete
        "aplicaFlete": flete == "Aplica",

        # Campo adicional para rastreo
        "bovedaMonte": boveda_monte
    }

    data["ventas"].append(venta)

    # Actualizar estadísticas de cliente
    if cliente in clientes_map:
        clientes_map[cliente]["totalComprado"] += ingreso
        clientes_map[cliente]["ventas"].append(venta_id)

    # Rastrear ventas por cliente
    if cliente not in ventas_por_cliente:
        ventas_por_cliente[cliente] = []
    ventas_por_cliente[cliente].append(venta)

    if estatus == "Pagado":
        stats['ventas_pagadas'] += 1
        print(f"  ✓ [PAGADO]    Fila {row} - {cliente} - ${boveda_monte:,.0f}")
    else:
        stats['ventas_pendientes'] += 1
        print(f"  ⏳ [PENDIENTE] Fila {row} - {cliente} - ${boveda_monte:,.0f}")

print(f"\n📊 Ventas importadas: {stats['ventas_pagadas'] + stats['ventas_pendientes']}")
print(f"   ├─ Pagadas: {stats['ventas_pagadas']}")
print(f"   └─ Pendientes: {stats['ventas_pendientes']}")

# =============================================================================
# 5. IMPORTAR ALMACEN - SALIDAS (desde Ventas)
# =============================================================================
print("\n" + "="*100)
print("5️⃣  IMPORTANDO ALMACEN - SALIDAS")
print("="*100)

for row in range(4, 200):
    fecha = ws_almacen.cell(row, 7).value
    if not fecha:
        continue

    cliente = limpiar_valor(ws_almacen.cell(row, 8).value, 'texto')
    cantidad = limpiar_valor(ws_almacen.cell(row, 9).value)
    concepto = limpiar_valor(ws_almacen.cell(row, 10).value, 'texto')
    observaciones = limpiar_valor(ws_almacen.cell(row, 11).value, 'texto')

    salida = {
        "id": f"SALIDA-{convertir_fecha(fecha)}-{cliente}-{row}",
        "tipo": "salida",
        "fecha": convertir_fecha(fecha),
        "cliente": cliente,
        "cantidad": cantidad,
        "concepto": concepto,
        "observaciones": observaciones
    }

    data["almacen"]["salidas"].append(salida)
    stats['salidas_almacen'] += 1

    print(f"  ✓ Salida - {cliente} - {cantidad} unidades")

# Calcular stock actual
total_entradas = sum(e["cantidad"] for e in data["almacen"]["entradas"])
total_salidas = sum(s["cantidad"] for s in data["almacen"]["salidas"])
stock_actual = total_entradas - total_salidas

data["almacen"]["stock"] = [{
    "producto": "Producto Principal",
    "cantidad": stock_actual,
    "ultimaActualizacion": datetime.now().isoformat()
}]

print(f"\n📊 Salidas de almacén importadas: {stats['salidas_almacen']}")
print(f"📦 Stock actual: {stock_actual} unidades (Entradas: {total_entradas}, Salidas: {total_salidas})")

# =============================================================================
# 6. IMPORTAR BÓVEDA MONTE - INGRESOS (solo ventas pagadas)
# =============================================================================
print("\n" + "="*100)
print("6️⃣  IMPORTANDO BÓVEDA MONTE - INGRESOS")
print("="*100)

ws_boveda = wb['Bóveda_Monte']

# Obtener RF Actual (capital actual)
rf_actual_boveda = limpiar_valor(ws_boveda.cell(2, 1).value)
data["bancos"]["bovedaMonte"]["capitalActual"] = rf_actual_boveda
data["bancos"]["bovedaMonte"]["historico"] = rf_actual_boveda

print(f"💰 RF Actual Bóveda Monte: ${rf_actual_boveda:,.2f}")

for row in range(4, 200):
    fecha = ws_boveda.cell(row, 1).value
    if not fecha:
        continue

    cliente = limpiar_valor(ws_boveda.cell(row, 2).value, 'texto')
    ingreso = limpiar_valor(ws_boveda.cell(row, 3).value)
    concepto = limpiar_valor(ws_boveda.cell(row, 4).value, 'texto')

    registro_ingreso = {
        "id": f"ING-bovedaMonte-{convertir_fecha(fecha)}-{cliente}-{row}",
        "fecha": convertir_fecha(fecha),
        "tipo": "Ingreso",
        "cliente": cliente,
        "monto": ingreso,
        "concepto": concepto or f"Venta pagada: {cliente}"
    }

    data["bancos"]["bovedaMonte"]["ingresos"].append(registro_ingreso)
    data["bancos"]["bovedaMonte"]["registros"].append(registro_ingreso)
    stats['ingresos_boveda'] += 1

    print(f"  ✓ Ingreso - {cliente} - ${ingreso:,.0f}")

print(f"\n📊 Ingresos Bóveda Monte: {stats['ingresos_boveda']}")

# =============================================================================
# 7. IMPORTAR BÓVEDA MONTE - GASTOS
# =============================================================================
print("\n" + "="*100)
print("7️⃣  IMPORTANDO BÓVEDA MONTE - GASTOS")
print("="*100)

for row in range(4, 200):
    fecha = ws_boveda.cell(row, 7).value
    if not fecha:
        continue

    origen = limpiar_valor(ws_boveda.cell(row, 8).value, 'texto')
    gasto = limpiar_valor(ws_boveda.cell(row, 9).value)
    tc = limpiar_valor(ws_boveda.cell(row, 10).value)
    pesos = limpiar_valor(ws_boveda.cell(row, 11).value)
    destino = limpiar_valor(ws_boveda.cell(row, 12).value, 'texto')
    concepto = limpiar_valor(ws_boveda.cell(row, 13).value, 'texto')
    observaciones = limpiar_valor(ws_boveda.cell(row, 14).value, 'texto')

    registro_gasto = {
        "id": f"EGR-bovedaMonte-{convertir_fecha(fecha)}-{row}",
        "fecha": convertir_fecha(fecha),
        "tipo": "Egreso",
        "origen": origen,
        "monto": gasto,
        "tc": tc,
        "pesos": pesos,
        "destino": destino,
        "concepto": concepto,
        "observaciones": observaciones
    }

    data["bancos"]["bovedaMonte"]["egresos"].append(registro_gasto)
    data["bancos"]["bovedaMonte"]["registros"].append(registro_gasto)

    # Agregar a gastosAbonos para tracking
    data["gastosAbonos"].append({
        "id": registro_gasto["id"],
        "fecha": registro_gasto["fecha"],
        "tipo": "gasto",
        "origenGastoOAbono": origen,
        "valor": gasto,
        "destino": destino,
        "concepto": concepto,
        "observaciones": observaciones
    })

    stats['gastos'] += 1

    print(f"  ✓ Gasto - {origen} → {destino} - ${gasto:,.0f}")

print(f"\n📊 Gastos Bóveda Monte: {stats['gastos']}")

# =============================================================================
# 8. IMPORTAR OTROS BANCOS
# =============================================================================
print("\n" + "="*100)
print("8️⃣  IMPORTANDO OTROS BANCOS")
print("="*100)

bancos_a_importar = [
    ('Bóveda_USA', 'bovedaUSA', 2, 1, 9),
    ('Utilidades', 'utilidades', 2, 1, 5),
    ('Flete_Sur', 'fleteSur', 2, 1, 5),
    ('Azteca', 'azteca', 2, 1, 8),
    ('Leftie', 'leftie', 2, 1, 9),
    ('Profit', 'profit', 2, 1, 9)
]

for hoja_name, banco_key, row_rf, col_rf, col_rf_actual in bancos_a_importar:
    print(f"\n  📁 Importando {hoja_name}...")

    ws_banco = wb[hoja_name]

    # Obtener RF Actual
    rf_actual = limpiar_valor(ws_banco.cell(row_rf, col_rf_actual).value)
    data["bancos"][banco_key]["capitalActual"] = rf_actual
    data["bancos"][banco_key]["historico"] = limpiar_valor(ws_banco.cell(row_rf, col_rf).value)

    print(f"    💰 RF Actual: ${rf_actual:,.2f}")

    # Importar ingresos
    count_ingresos = 0
    for row in range(4, 100):
        fecha = ws_banco.cell(row, 1).value
        if not fecha:
            continue

        cliente = limpiar_valor(ws_banco.cell(row, 2).value, 'texto')
        ingreso = limpiar_valor(ws_banco.cell(row, 3).value)

        registro_ingreso = {
            "id": f"ING-{banco_key}-{convertir_fecha(fecha)}-{row}",
            "fecha": convertir_fecha(fecha),
            "tipo": "Ingreso",
            "cliente": cliente,
            "monto": ingreso,
            "concepto": ""
        }

        data["bancos"][banco_key]["ingresos"].append(registro_ingreso)
        data["bancos"][banco_key]["registros"].append(registro_ingreso)
        count_ingresos += 1

    print(f"    ✓ Ingresos: {count_ingresos}")

print(f"\n📊 Todos los bancos importados correctamente")

# =============================================================================
# 9. IMPORTAR CONTROL_MAESTRO - GYA (Gastos y Abonos)
# =============================================================================
print("\n" + "="*100)
print("9️⃣  IMPORTANDO GYA (Gastos y Abonos)")
print("="*100)

count_gya = 0
for row in range(4, 200):
    fecha_gya = ws_control.cell(row, 15).value
    if not fecha_gya:
        continue

    origen = limpiar_valor(ws_control.cell(row, 16).value, 'texto')
    valor = limpiar_valor(ws_control.cell(row, 17).value)
    tc = limpiar_valor(ws_control.cell(row, 18).value)
    pesos = limpiar_valor(ws_control.cell(row, 19).value)

    # Determinar tipo (abono o gasto)
    tipo = "abono" if origen in clientes_map else "gasto"

    gya_registro = {
        "id": f"GYA-{convertir_fecha(fecha_gya)}-{origen}-{row}",
        "fecha": convertir_fecha(fecha_gya),
        "tipo": tipo,
        "origenGastoOAbono": origen,
        "valor": valor,
        "tc": tc,
        "pesos": pesos,
        "concepto": f"{tipo.capitalize()} - {origen}"
    }

    data["gastosAbonos"].append(gya_registro)

    # Actualizar totalAbonado del cliente
    if tipo == "abono" and origen in clientes_map:
        clientes_map[origen]["totalAbonado"] += valor

    count_gya += 1
    print(f"  ✓ [{tipo.upper()}] {origen} - ${valor:,.0f}")

print(f"\n📊 Registros GYA importados: {count_gya}")

# =============================================================================
# RESUMEN FINAL Y GUARDADO
# =============================================================================
print("\n" + "="*100)
print("📊 RESUMEN FINAL DE IMPORTACIÓN")
print("="*100)

print(f"\n✅ VENTAS:")
print(f"   Total: {len(data['ventas'])}")
print(f"   ├─ Pagadas: {stats['ventas_pagadas']}")
print(f"   └─ Pendientes: {stats['ventas_pendientes']}")

print(f"\n✅ ÓRDENES DE COMPRA:")
print(f"   Total: {len(data['ordenesCompra'])}")

print(f"\n✅ CLIENTES:")
print(f"   Total: {len(data['clientes'])}")

print(f"\n✅ ALMACÉN:")
print(f"   Entradas: {len(data['almacen']['entradas'])}")
print(f"   Salidas: {len(data['almacen']['salidas'])}")
print(f"   Stock Actual: {stock_actual} unidades")

print(f"\n✅ BANCOS:")
for banco_key, banco_data in data['bancos'].items():
    print(f"   {banco_data['nombre']}: ${banco_data['capitalActual']:,.2f} {banco_data['moneda']}")

print(f"\n✅ GASTOS Y ABONOS:")
print(f"   Total: {len(data['gastosAbonos'])}")

# Guardar JSON
print(f"\n💾 Guardando datos en: {OUTPUT_PATH}")
with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f"✅ Archivo guardado exitosamente")

print("\n" + "="*100)
print("🎉 IMPORTACIÓN COMPLETADA")
print("="*100)

print("\n📋 Checklist de Validación:")
print("  [ ] Verificar total de ventas = 96")
print("  [ ] Verificar ventas pagadas = 50")
print("  [ ] Verificar ventas pendientes = 46")
print("  [ ] Verificar Bóveda Monte capital = $5,722,280")
print("  [ ] Verificar stock almacén = 17 unidades")
print("  [ ] Verificar que ventas pendientes tienen adeudo > 0")
print("  [ ] Verificar que ventas pagadas tienen adeudo = 0")

print("\n🚀 Próximo paso: Abrir FlowDistributor y verificar los datos importados")
