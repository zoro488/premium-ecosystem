"""
📦 EXTRACTOR DE DISTRIBUIDORES - DESDE EXCEL
=============================================
Extrae TODOS los distribuidores del Excel sin omitir ninguno
"""

import openpyxl
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r'C:\Users\xpovo\Documents\premium-ecosystem\Administación_General.xlsx'
OUTPUT_JSON = r'C:\Users\xpovo\Documents\premium-ecosystem\src\apps\FlowDistributor\data\distribuidores-excel.json'

def leer_valor(celda):
    """Lee valor de celda"""
    if celda.value is None:
        return None
    if isinstance(celda.value, (int, float)):
        return celda.value
    return str(celda.value).strip()

print('='*80)
print('📦 EXTRACTOR DE DISTRIBUIDORES')
print('='*80)
print()

try:
    print(f'📂 Cargando Excel...')
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Distribuidores']
    print('✅ Excel cargado')
    print()

    print('📋 Extrayendo distribuidores...')
    print('-'*80)

    distribuidores = []
    for row_idx in range(4, ws.max_row + 1):
        oc = leer_valor(ws.cell(row_idx, 1))  # Columna A: OC

        # Si no hay OC, saltar
        if not oc:
            continue

        distribuidor = {
            'id': f'dist-{len(distribuidores) + 1}',
            'oc': oc,
            'fecha': leer_valor(ws.cell(row_idx, 2)),  # B
            'origen': leer_valor(ws.cell(row_idx, 3)),  # C
            'cantidad': leer_valor(ws.cell(row_idx, 4)) or 0,  # D
            'costoDistribuidor': leer_valor(ws.cell(row_idx, 5)) or 0,  # E
            'costoTransporte': leer_valor(ws.cell(row_idx, 6)) or 0,  # F
            'costoPorUnidad': leer_valor(ws.cell(row_idx, 7)) or 0,  # G
            'stockActual': leer_valor(ws.cell(row_idx, 8)) or 0,  # H
            'costoTotal': leer_valor(ws.cell(row_idx, 9)) or 0,  # I
            'pagoDistribuidor': leer_valor(ws.cell(row_idx, 10)) or 0,  # J
            'deuda': leer_valor(ws.cell(row_idx, 11)) or 0,  # K
        }

        distribuidores.append(distribuidor)

        if len(distribuidores) % 100 == 0:
            print(f'  📌 Procesados: {len(distribuidores)} distribuidores...')

    print()
    print('='*80)
    print(f'✅ TOTAL EXTRAÍDO: {len(distribuidores)} distribuidores')
    print('='*80)
    print()

    # Guardar JSON
    print(f'💾 Guardando: {OUTPUT_JSON}')
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(distribuidores, f, ensure_ascii=False, indent=2)

    print('✅ Archivo guardado')
    print()
    print('='*80)
    print('✅ EXTRACCIÓN COMPLETA - SIN OMISIONES')
    print('='*80)

except Exception as e:
    print(f'\n❌ Error: {e}')
    import traceback
    traceback.print_exc()
