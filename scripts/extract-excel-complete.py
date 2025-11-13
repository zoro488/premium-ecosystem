#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
🚀 EXTRACTOR COMPLETO DE DATOS EXCEL → FLOWDISTRIBUTOR
Extrae TODOS los datos de todas las hojas del Excel y genera archivo JSON optimizado
"""

import json
import re
import sys
from datetime import date, datetime
from decimal import Decimal

import openpyxl

# Configuración
EXCEL_PATH = r"C:\Users\xpovo\Downloads\Copia de Administación_General.xlsx"
OUTPUT_JSON = r"src\apps\FlowDistributor\data\flowdistributor_complete_data.json"
OUTPUT_BACKUP = r"backups\flowdistributor_data_backup_{timestamp}.json"

# Colores por banco (mismo esquema que FlowDistributor_2.0)
BANK_COLORS = {
    "Bóveda_Monte": "#10b981",
    "Bóveda_USA": "#3b82f6",
    "Utilidades": "#8b5cf6",
    "Flete_Sur": "#f59e0b",
    "Azteca": "#ef4444",
    "Leftie": "#06b6d4",
    "Profit": "#22c55e",
}

BANK_ICONS = {
    "Bóveda_Monte": "🏦",
    "Bóveda_USA": "🇺🇸",
    "Utilidades": "💰",
    "Flete_Sur": "🚚",
    "Azteca": "🏛️",
    "Leftie": "💳",
    "Profit": "📈",
}


def serialize_value(value):
    """Convierte valores de Excel a JSON-serializable"""
    if value is None:
        return None
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float, str, bool)):
        return value
    if str(value).startswith("="):
        return None  # Ignorar fórmulas
    return str(value)


def clean_number(value):
    """Limpia y convierte valores numéricos"""
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        # Remover símbolos de moneda y espacios
        cleaned = re.sub(r"[^\d.-]", "", value)
        try:
            return float(cleaned) if cleaned else 0
        except:
            return 0
    return 0


def extract_distribuidores(ws):
    """Extrae datos de la hoja Distribuidores"""
    print("📦 Extrayendo Distribuidores...")
    distribuidores = []
    ordenes_compra = []

    # Leer órdenes de compra (filas 4+)
    for row in ws.iter_rows(min_row=4, max_row=100, values_only=True):
        if not row[0]:  # Si no hay OC, terminar
            break

        oc = {
            "id": str(row[0]) if row[0] else None,
            "fecha": serialize_value(row[1]),
            "distribuidor": str(row[2]) if row[2] else None,
            "cantidad": clean_number(row[3]),
            "costoDistribuidor": clean_number(row[4]),
            "costoTransporte": clean_number(row[5]),
            "costoPorUnidad": clean_number(row[6]),
            "stockActual": clean_number(row[7]),
            "costoTotal": clean_number(row[8]),
            "pagoDistribuidor": clean_number(row[9]),
            "deuda": clean_number(row[10]),
        }

        if oc["id"]:
            ordenes_compra.append(oc)

    # Leer resumen de distribuidores (columnas M, N, O, P)
    dist_data = {}
    for row in ws.iter_rows(min_row=4, max_row=20, values_only=True):
        if row[12]:  # Columna M (distribuidor)
            nombre = str(row[12])
            dist_data[nombre] = {
                "nombre": nombre,
                "costoTotal": clean_number(row[13]),
                "abonos": clean_number(row[14]),
                "pendiente": clean_number(row[15]),
            }

    # Convertir a lista
    for nombre, data in dist_data.items():
        distribuidores.append(
            {
                "id": f"DIST-{len(distribuidores)+1:03d}",
                "nombre": nombre,
                "codigo": nombre[:3].upper(),
                "totalComprado": data["costoTotal"],
                "totalPagado": data["abonos"],
                "adeudo": data["pendiente"],
                "estado": "activo" if data["pendiente"] > 0 else "saldado",
                "calificacion": 4.5,
                "diasCredito": 30,
                "contacto": {"telefono": "", "email": "", "direccion": ""},
                "historialPagos": [],
            }
        )

    print(f"  ✅ {len(ordenes_compra)} órdenes de compra")
    print(f"  ✅ {len(distribuidores)} distribuidores")

    return ordenes_compra, distribuidores


def extract_control_maestro(ws):
    """Extrae datos del Control Maestro (ventas)"""
    print("📊 Extrayendo Control Maestro (Ventas)...")
    ventas = []

    for row in ws.iter_rows(min_row=4, max_row=1000, values_only=True):
        if not row[0]:  # Sin fecha, terminar
            break

        venta = {
            "id": f"V-{len(ventas)+1:04d}",
            "fecha": serialize_value(row[0]),
            "ocRelacionada": str(row[1]) if row[1] else None,
            "cantidad": clean_number(row[2]),
            "cliente": str(row[3]) if row[3] else None,
            "bovedaMonte": clean_number(row[4]),
            "precioVenta": clean_number(row[5]),
            "ingreso": clean_number(row[6]),
            "flete": str(row[7]) if row[7] else "No Aplica",
            "fleteUtilidad": clean_number(row[8]),
            "utilidad": clean_number(row[9]),
            "estatus": str(row[10]) if row[10] else "Pendiente",
            "concepto": str(row[11]) if row[11] else None,
        }

        if venta["fecha"]:
            ventas.append(venta)

    print(f"  ✅ {len(ventas)} ventas registradas")
    return ventas


def extract_almacen(ws):
    """Extrae datos del Almacén"""
    print("📦 Extrayendo Almacén Monte...")

    # Leer RF Actual (E2)
    stock_actual = clean_number(ws["E2"].value)
    total_entradas = clean_number(ws["A2"].value)
    total_salidas = clean_number(ws["G2"].value)

    movimientos_entrada = []
    movimientos_salida = []

    # Entradas (columnas A-F)
    for row in ws.iter_rows(min_row=4, max_row=500, values_only=True):
        if not row[0]:
            break

        entrada = {
            "id": f"ENT-{len(movimientos_entrada)+1:04d}",
            "oc": str(row[0]) if row[0] else None,
            "cliente": str(row[1]) if row[1] else None,
            "distribuidor": str(row[2]) if row[2] else None,
            "cantidad": clean_number(row[3]),
            "fecha": serialize_value(row[4]),
            "tipo": "entrada",
        }

        if entrada["oc"]:
            movimientos_entrada.append(entrada)

    # Salidas (columnas G-K)
    for row in ws.iter_rows(min_row=4, max_row=500, values_only=True):
        if not row[6]:  # Columna G
            break

        salida = {
            "id": f"SAL-{len(movimientos_salida)+1:04d}",
            "fecha": serialize_value(row[6]),
            "cliente": str(row[7]) if row[7] else None,
            "cantidad": clean_number(row[8]),
            "concepto": str(row[9]) if row[9] else None,
            "observaciones": str(row[10]) if row[10] else None,
            "tipo": "salida",
        }

        if salida["fecha"]:
            movimientos_salida.append(salida)

    almacen = {
        "stockActual": stock_actual,
        "stockMinimo": 50,
        "stockMaximo": 3000,
        "totalEntradas": total_entradas,
        "totalSalidas": total_salidas,
        "valorInventario": stock_actual * 6300,  # Precio promedio
        "costoPromedio": 6300,
        "movimientos": movimientos_entrada + movimientos_salida,
        "alertas": {
            "stockBajo": stock_actual < 50,
            "reordenar": stock_actual < 100,
            "sobrestock": stock_actual > 2500,
        },
        "ubicaciones": {
            "principal": {"stock": stock_actual, "capacidad": 1000},
            "secundario": {"stock": 0, "capacidad": 500},
            "transito": {"stock": 0, "capacidad": 100},
        },
    }

    print(f"  ✅ Stock actual: {stock_actual} unidades")
    print(f"  ✅ {len(movimientos_entrada)} entradas")
    print(f"  ✅ {len(movimientos_salida)} salidas")

    return almacen


def extract_banco(ws, nombre_banco):
    """Extrae datos de una hoja de banco"""
    print(f"💰 Extrayendo {nombre_banco}...")

    # Leer RF Actual (E2 o I2 dependiendo del banco)
    rf_actual = None
    for cell in ["E2", "I2", "H2"]:
        try:
            val = ws[cell].value
            if val and isinstance(val, (int, float)):
                rf_actual = clean_number(val)
                break
        except:
            continue

    if rf_actual is None:
        rf_actual = 0

    # Total ingresos y gastos
    total_ingresos = clean_number(ws["A2"].value) if ws["A2"].value else 0

    # Buscar celda de gastos (puede variar)
    total_gastos = 0
    for cell in ["G2", "K2", "J2"]:
        try:
            val = ws[cell].value
            if val and isinstance(val, (int, float)):
                total_gastos = clean_number(val)
                break
        except:
            continue

    ingresos = []
    gastos = []

    # Extraer ingresos (primeras 4-6 columnas)
    for row in ws.iter_rows(min_row=4, max_row=500, values_only=True):
        if not row[0]:  # Sin fecha
            break

        ingreso = {
            "id": f"ING-{len(ingresos)+1:04d}",
            "fecha": serialize_value(row[0]),
            "cliente": str(row[1]) if row[1] else None,
            "monto": clean_number(row[2]),
            "tc": clean_number(row[3]) if len(row) > 3 else 0,
            "pesos": clean_number(row[4]) if len(row) > 4 else 0,
            "concepto": str(row[6]) if len(row) > 6 and row[6] else None,
            "observaciones": str(row[7]) if len(row) > 7 and row[7] else None,
            "tipo": "ingreso",
        }

        if ingreso["fecha"] and ingreso["monto"] > 0:
            ingresos.append(ingreso)

    # Extraer gastos (siguientes columnas, típicamente después de columna F o J)
    inicio_gastos = 6  # Ajustar según banco
    for row in ws.iter_rows(min_row=4, max_row=500, values_only=True):
        # Buscar fecha en columnas de gastos
        fecha_gasto = None
        for i in range(inicio_gastos, min(len(row), inicio_gastos + 8)):
            if row[i] and isinstance(row[i], (date, datetime)):
                fecha_gasto = row[i]
                break

        if not fecha_gasto:
            continue

        gasto = {
            "id": f"GAS-{len(gastos)+1:04d}",
            "fecha": serialize_value(fecha_gasto),
            "origen": (
                str(row[inicio_gastos + 1])
                if len(row) > inicio_gastos + 1 and row[inicio_gastos + 1]
                else None
            ),
            "monto": (
                clean_number(row[inicio_gastos + 2])
                if len(row) > inicio_gastos + 2
                else 0
            ),
            "tc": (
                clean_number(row[inicio_gastos + 3])
                if len(row) > inicio_gastos + 3
                else 0
            ),
            "destino": (
                str(row[inicio_gastos + 5])
                if len(row) > inicio_gastos + 5 and row[inicio_gastos + 5]
                else None
            ),
            "concepto": (
                str(row[inicio_gastos + 6])
                if len(row) > inicio_gastos + 6 and row[inicio_gastos + 6]
                else None
            ),
            "observaciones": (
                str(row[inicio_gastos + 7])
                if len(row) > inicio_gastos + 7 and row[inicio_gastos + 7]
                else None
            ),
            "tipo": "gasto",
        }

        if gasto["fecha"] and gasto["monto"] > 0:
            gastos.append(gasto)

    banco_data = {
        "nombre": nombre_banco.replace("_", " "),
        "codigo": nombre_banco[:5].upper(),
        "capitalActual": rf_actual,
        "capitalInicial": rf_actual + total_gastos - total_ingresos,
        "ingresos": ingresos,
        "gastos": gastos,
        "totalIngresos": total_ingresos,
        "totalGastos": total_gastos,
        "estado": "activo" if rf_actual >= 0 else "negativo",
        "color": BANK_COLORS.get(nombre_banco, "#6b7280"),
        "icono": BANK_ICONS.get(nombre_banco, "🏦"),
        "limiteCredito": 10000000,
        "tasaInteres": 0,
    }

    print(f"  ✅ RF Actual: ${rf_actual:,.2f}")
    print(f"  ✅ {len(ingresos)} ingresos, {len(gastos)} gastos")

    return banco_data


def extract_clientes(ws):
    """Extrae datos de la hoja Clientes"""
    print("👥 Extrayendo Clientes...")
    clientes = []

    for row in ws.iter_rows(min_row=4, max_row=200, values_only=True):
        if not row[4]:  # Columna E (nombre cliente)
            break

        cliente = {
            "id": f"CLI-{len(clientes)+1:03d}",
            "nombre": str(row[4]),
            "deuda": clean_number(row[6]),
            "abonos": clean_number(row[7]),
            "pendiente": clean_number(row[8]),
            "totalComprado": clean_number(row[6]),
            "estado": "activo" if clean_number(row[8]) > 0 else "saldado",
            "observaciones": str(row[9]) if row[9] else None,
        }

        if cliente["nombre"] and cliente["nombre"] not in ["Cliente", ""]:
            clientes.append(cliente)

    print(f"  ✅ {len(clientes)} clientes")
    return clientes


def main():
    print("\n" + "=" * 60)
    print("🚀 EXTRACTOR COMPLETO EXCEL → FLOWDISTRIBUTOR")
    print("=" * 60 + "\n")

    try:
        # Cargar Excel
        print(f"📂 Cargando Excel: {EXCEL_PATH}")
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        print(f"✅ Excel cargado. Hojas: {', '.join(wb.sheetnames)}\n")

        # Estructura de datos completa
        data = {
            "config": {
                "empresa": "FLOW DISTRIBUTOR",
                "version": "3.0",
                "moneda": "MXN",
                "tasaCambio": 20.5,
                "actualizado": datetime.now().isoformat(),
                "excelOriginal": EXCEL_PATH,
            },
            "bancos": {},
            "almacen": {},
            "distribuidores": [],
            "ordenesCompra": [],
            "clientes": [],
            "ventas": [],
            "metricas": {},
        }

        # 1. Distribuidores y Órdenes de Compra
        if "Distribuidores" in wb.sheetnames:
            data["ordenesCompra"], data["distribuidores"] = extract_distribuidores(
                wb["Distribuidores"]
            )

        # 2. Control Maestro (Ventas)
        if "Control_Maestro" in wb.sheetnames:
            data["ventas"] = extract_control_maestro(wb["Control_Maestro"])

        # 3. Almacén
        if "Almacen_Monte" in wb.sheetnames:
            data["almacen"] = extract_almacen(wb["Almacen_Monte"])

        # 4. Bancos
        bancos_hojas = [
            "Bóveda_Monte",
            "Bóveda_USA",
            "Utilidades",
            "Flete_Sur",
            "Azteca",
            "Leftie",
            "Profit",
        ]
        for banco in bancos_hojas:
            if banco in wb.sheetnames:
                banco_key = banco.lower().replace("ó", "o").replace("_", "")
                data["bancos"][banco_key] = extract_banco(wb[banco], banco)

        # 5. Clientes
        if "Clientes" in wb.sheetnames:
            data["clientes"] = extract_clientes(wb["Clientes"])

        # 6. Calcular métricas globales
        print("\n📊 Calculando métricas globales...")

        capital_total = sum(b["capitalActual"] for b in data["bancos"].values())
        ventas_totales = sum(v["ingreso"] for v in data["ventas"])
        compras_totales = sum(oc["costoTotal"] for oc in data["ordenesCompra"])
        adeudo_distribuidores = sum(d["adeudo"] for d in data["distribuidores"])
        adeudo_clientes = sum(c["pendiente"] for c in data["clientes"])

        data["metricas"] = {
            "capitalTotal": capital_total,
            "ventasTotales": ventas_totales,
            "comprasTotales": compras_totales,
            "utilidadNeta": ventas_totales - compras_totales,
            "margenPromedio": (
                (ventas_totales - compras_totales) / ventas_totales * 100
                if ventas_totales > 0
                else 0
            ),
            "rotacionInventario": (
                ventas_totales / data["almacen"]["valorInventario"]
                if data["almacen"]["valorInventario"] > 0
                else 0
            ),
            "diasInventario": (
                365 / (ventas_totales / data["almacen"]["valorInventario"])
                if ventas_totales > 0
                else 0
            ),
            "liquidezInmediata": capital_total,
            "apalancamiento": (
                adeudo_distribuidores / capital_total if capital_total > 0 else 0
            ),
            "roi": (
                (ventas_totales - compras_totales) / compras_totales * 100
                if compras_totales > 0
                else 0
            ),
            "adeudoDistribuidores": adeudo_distribuidores,
            "adeudoClientes": adeudo_clientes,
            "totalBancos": len(data["bancos"]),
            "totalVentas": len(data["ventas"]),
            "totalOrdenes": len(data["ordenesCompra"]),
            "totalClientes": len(data["clientes"]),
            "totalDistribuidores": len(data["distribuidores"]),
        }

        # Guardar JSON
        print(f"\n💾 Guardando datos...")

        # Crear directorios si no existen
        import os

        os.makedirs(os.path.dirname(OUTPUT_JSON), exist_ok=True)
        os.makedirs("backups", exist_ok=True)

        # Guardar archivo principal
        with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        print(f"✅ Datos guardados: {OUTPUT_JSON}")

        # Guardar backup
        backup_file = OUTPUT_BACKUP.format(
            timestamp=datetime.now().strftime("%Y%m%d_%H%M%S")
        )
        with open(backup_file, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        print(f"✅ Backup creado: {backup_file}")

        # Resumen
        print("\n" + "=" * 60)
        print("✅ EXTRACCIÓN COMPLETADA CON ÉXITO")
        print("=" * 60)
        print(f"\n📊 RESUMEN DE DATOS EXTRAÍDOS:")
        print(f"   • Bancos: {data['metricas']['totalBancos']}")
        print(f"   • Capital Total: ${data['metricas']['capitalTotal']:,.2f}")
        print(f"   • Ventas: {data['metricas']['totalVentas']}")
        print(f"   • Órdenes de Compra: {data['metricas']['totalOrdenes']}")
        print(f"   • Distribuidores: {data['metricas']['totalDistribuidores']}")
        print(f"   • Clientes: {data['metricas']['totalClientes']}")
        print(f"   • Stock Almacén: {data['almacen']['stockActual']} unidades")
        print(f"   • Utilidad Neta: ${data['metricas']['utilidadNeta']:,.2f}")
        print(f"   • ROI: {data['metricas']['roi']:.2f}%")

        print("\n🎯 SIGUIENTE PASO:")
        print(f"   npm run dev")
        print(f"   Abrir: http://localhost:5173/flowdistributor\n")

        return 0

    except FileNotFoundError:
        print(f"\n❌ ERROR: No se encontró el archivo Excel:")
        print(f"   {EXCEL_PATH}")
        print(f"\n💡 Verifica que el archivo exista en esa ubicación.\n")
        return 1

    except Exception as e:
        print(f"\n❌ ERROR durante la extracción:")
        print(f"   {str(e)}")
        import traceback

        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
