"""
ANÁLISIS CORRECTO DEL EXCEL
Separando las 3 tablas por hoja correctamente
Detectando columnas vacías como separadores
"""

import openpyxl
import json
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r'C:\Users\xpovo\Downloads\Copia de Administación_General.xlsx'

print('=' * 100)
print('ANÁLISIS CORRECTO - SEPARANDO TABLAS POR COLUMNAS VACÍAS')
print('=' * 100)
print(f'\nArchivo: {excel_path}')

wb = openpyxl.load_workbook(excel_path, data_only=True)

def obtener_valor(celda):
    """Obtiene el valor de una celda"""
    if celda is None:
        return None
    if isinstance(celda, datetime):
        return celda.strftime('%Y-%m-%d')
    return celda

def extraer_tabla_separada(ws, col_inicio, fila_header, max_filas=500):
    """
    Extrae UNA tabla hasta encontrar columna vacía o fin de columnas
    """
    # Leer headers hasta encontrar columna vacía
    headers = []
    col = col_inicio

    while col <= ws.max_column:
        header = ws.cell(fila_header, col).value

        # Si el header está vacío, terminamos esta tabla
        if header is None or str(header).strip() == '':
            break

        headers.append(str(header))
        col += 1

    num_columnas = len(headers)
    if num_columnas == 0:
        return [], [], col

    # Leer datos
    datos = []
    for fila in range(fila_header + 1, fila_header + max_filas):
        fila_vacia = True
        fila_datos = {}

        for i, header in enumerate(headers):
            valor = ws.cell(fila, col_inicio + i).value
            if valor is not None and str(valor).strip() != '':
                fila_vacia = False
            fila_datos[header] = obtener_valor(valor)

        if fila_vacia:
            break

        datos.append(fila_datos)

    # Retornar también la siguiente columna a revisar
    return headers, datos, col_inicio + num_columnas

resultado = {
    'metadata': {
        'archivo': excel_path,
        'fecha_analisis': datetime.now().isoformat(),
        'total_hojas': len(wb.sheetnames)
    },
    'hojas': {}
}

print('\n' + '=' * 100)
print('EXTRACCIÓN TABLA POR TABLA (SEPARADAS CORRECTAMENTE)')
print('=' * 100)

for sheet_name in wb.sheetnames:
    print(f'\n{"=" * 100}')
    print(f'📄 HOJA: {sheet_name}')
    print(f'{"=" * 100}')

    ws = wb[sheet_name]
    resultado['hojas'][sheet_name] = {'tablas': []}

    # Buscar tablas separadas por columnas vacías
    col_actual = 1
    numero_tabla = 0

    while col_actual <= ws.max_column:
        # Verificar si hay un header en esta columna
        header_test = ws.cell(3, col_actual).value

        if header_test and str(header_test).strip() != '':
            # Extraer esta tabla
            headers, datos, siguiente_col = extraer_tabla_separada(ws, col_actual, 3)

            if len(datos) > 0:
                numero_tabla += 1

                # Contar ingresos y gastos si aplica
                ingresos_count = len([d for d in datos if d.get('Ingreso')])
                gastos_count = len([d for d in datos if d.get('Gasto')])

                tabla_info = {
                    'numero': numero_tabla,
                    'columna_inicio': col_actual,
                    'columna_fin': siguiente_col - 1,
                    'headers': headers,
                    'num_columnas': len(headers),
                    'num_registros': len(datos),
                    'datos': datos
                }

                resultado['hojas'][sheet_name]['tablas'].append(tabla_info)

                print(f'\n📊 TABLA {numero_tabla}:')
                print(f'   Rango: Columnas {col_actual} a {siguiente_col - 1}')
                print(f'   Headers ({len(headers)}): {", ".join(headers)}')
                print(f'   Total registros: {len(datos)}')

                if ingresos_count > 0 or gastos_count > 0:
                    print(f'   → Ingresos: {ingresos_count}')
                    print(f'   → Gastos: {gastos_count}')

                print(f'\n   Primeros 2 registros:')
                for i, registro in enumerate(datos[:2], 1):
                    print(f'   {i}. {registro}')

                if len(datos) > 2:
                    print(f'   ... y {len(datos) - 2} registros más')

            # Saltar a la siguiente columna después de esta tabla
            col_actual = siguiente_col + 1
        else:
            # Columna vacía, saltar
            col_actual += 1

    print(f'\n✅ Total de tablas encontradas en {sheet_name}: {numero_tabla}')
    resultado['hojas'][sheet_name]['total_tablas'] = numero_tabla

# RESUMEN
print('\n' + '=' * 100)
print('📊 RESUMEN GENERAL POR HOJA')
print('=' * 100)

total_tablas = 0
total_registros = 0

for sheet_name, sheet_data in resultado['hojas'].items():
    num_tablas = sheet_data['total_tablas']
    num_registros = sum(tabla['num_registros'] for tabla in sheet_data['tablas'])
    total_tablas += num_tablas
    total_registros += num_registros

    print(f'\n📄 {sheet_name}:')
    print(f'   → Tablas: {num_tablas}')
    print(f'   → Registros totales: {num_registros}')

    for tabla in sheet_data['tablas']:
        cols_str = f"cols {tabla['columna_inicio']}-{tabla['columna_fin']}"
        headers_preview = ", ".join(tabla['headers'][:3])
        if len(tabla['headers']) > 3:
            headers_preview += f"... (+{len(tabla['headers'])-3})"
        print(f'      • Tabla {tabla["numero"]}: {tabla["num_registros"]} registros ({cols_str}) - {headers_preview}')

print(f'\n{"=" * 100}')
print(f'📈 TOTALES GENERALES:')
print(f'   Total hojas: {len(resultado["hojas"])}')
print(f'   Total tablas: {total_tablas}')
print(f'   Total registros: {total_registros}')
print(f'{"=" * 100}')

# Guardar
output_json = 'scripts/analisis_excel_separado_correcto.json'
with open(output_json, 'w', encoding='utf-8') as f:
    json.dump(resultado, f, ensure_ascii=False, indent=2, default=str)

print(f'\n✅ Análisis guardado en: {output_json}')
