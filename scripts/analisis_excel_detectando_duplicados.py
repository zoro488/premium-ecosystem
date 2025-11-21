"""
ANÁLISIS DETECTANDO TABLAS POR HEADERS DUPLICADOS
Cuando un header se repite, indica inicio de nueva tabla
"""

import openpyxl
import json
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r'C:\Users\xpovo\Downloads\Copia de Administación_General.xlsx'

print('=' * 100)
print('ANÁLISIS DETECTANDO TABLAS POR HEADERS DUPLICADOS')
print('=' * 100)

wb = openpyxl.load_workbook(excel_path, data_only=True)

def obtener_valor(celda):
    if celda is None:
        return None
    if isinstance(celda, datetime):
        return celda.strftime('%Y-%m-%d')
    return celda

def detectar_tablas_por_headers(ws, fila_header=3):
    """
    Lee todos los headers y los separa cuando encuentra duplicados
    """
    # Leer TODOS los headers de la fila 3
    todos_headers = []
    col = 1
    while col <= ws.max_column:
        header = ws.cell(fila_header, col).value
        if header and str(header).strip() != '':
            todos_headers.append((col, str(header)))
        col += 1

    if len(todos_headers) == 0:
        return []

    # Detectar tablas separando cuando hay headers duplicados
    tablas = []
    tabla_actual = []
    headers_vistos = set()

    for col, header in todos_headers:
        # Si ya vimos este header, empezamos nueva tabla
        if header in headers_vistos and len(tabla_actual) > 0:
            tablas.append(tabla_actual)
            tabla_actual = [(col, header)]
            headers_vistos = {header}
        else:
            tabla_actual.append((col, header))
            headers_vistos.add(header)

    # Agregar última tabla
    if len(tabla_actual) > 0:
        tablas.append(tabla_actual)

    return tablas

def extraer_datos_tabla(ws, columnas_tabla, fila_inicio=4, max_filas=500):
    """
    Extrae datos de una tabla dadas sus columnas
    columnas_tabla: lista de (col_num, header_name)
    """
    headers = [h for _, h in columnas_tabla]
    columnas = [c for c, _ in columnas_tabla]

    datos = []
    for fila in range(fila_inicio, fila_inicio + max_filas):
        fila_vacia = True
        fila_datos = {}

        for i, (col, header) in enumerate(columnas_tabla):
            valor = ws.cell(fila, col).value
            if valor is not None and str(valor).strip() != '':
                fila_vacia = False
            fila_datos[header] = obtener_valor(valor)

        if fila_vacia:
            break

        datos.append(fila_datos)

    return headers, datos

resultado = {
    'metadata': {
        'archivo': excel_path,
        'fecha_analisis': datetime.now().isoformat()
    },
    'hojas': {}
}

print('\n' + '=' * 100)
print('EXTRACCIÓN POR DETECCIÓN DE HEADERS DUPLICADOS')
print('=' * 100)

for sheet_name in wb.sheetnames:
    print(f'\n{"=" * 100}')
    print(f'📄 HOJA: {sheet_name}')
    print(f'{"=" * 100}')

    ws = wb[sheet_name]
    resultado['hojas'][sheet_name] = {'tablas': []}

    # Detectar tablas por headers duplicados
    tablas_detectadas = detectar_tablas_por_headers(ws)

    print(f'\n🔍 Headers detectados en fila 3:')

    # Mostrar TODOS los headers antes de separar
    col = 1
    all_headers = []
    while col <= 30:  # Revisar primeras 30 columnas
        header = ws.cell(3, col).value
        if header:
            all_headers.append(f'Col{col}:{header}')
        col += 1

    print(f'   {" | ".join(all_headers[:20])}')
    if len(all_headers) > 20:
        print(f'   ... y {len(all_headers) - 20} más')

    print(f'\n📊 Tablas detectadas: {len(tablas_detectadas)}')

    for i, columnas_tabla in enumerate(tablas_detectadas, 1):
        headers, datos = extraer_datos_tabla(ws, columnas_tabla)

        if len(datos) > 0:
            col_inicio = columnas_tabla[0][0]
            col_fin = columnas_tabla[-1][0]

            # Contar ingresos/gastos
            ingresos_count = len([d for d in datos if d.get('Ingreso')])
            gastos_count = len([d for d in datos if d.get('Gasto')])

            tabla_info = {
                'numero': i,
                'columna_inicio': col_inicio,
                'columna_fin': col_fin,
                'headers': headers,
                'num_registros': len(datos),
                'datos': datos[:5]  # Solo guardar primeros 5 para no saturar el JSON
            }

            resultado['hojas'][sheet_name]['tablas'].append(tabla_info)

            print(f'\n   📊 TABLA {i}:')
            print(f'      Columnas: {col_inicio} a {col_fin}')
            print(f'      Headers ({len(headers)}): {", ".join(headers)}')
            print(f'      Registros: {len(datos)}')

            if ingresos_count > 0 or gastos_count > 0:
                print(f'      → Ingresos: {ingresos_count}')
                print(f'      → Gastos: {gastos_count}')

            print(f'\n      Primeros 2 registros:')
            for j, reg in enumerate(datos[:2], 1):
                print(f'      {j}. {reg}')

    resultado['hojas'][sheet_name]['total_tablas'] = len(tablas_detectadas)

# RESUMEN
print('\n' + '=' * 100)
print('📊 RESUMEN FINAL')
print('=' * 100)

for sheet_name, sheet_data in resultado['hojas'].items():
    num_tablas = len(sheet_data['tablas'])
    total_regs = sum(t['num_registros'] for t in sheet_data['tablas'])

    print(f'\n📄 {sheet_name}: {num_tablas} tablas | {total_regs} registros')

    for tabla in sheet_data['tablas']:
        print(f'   • Tabla {tabla["numero"]}: {tabla["num_registros"]} regs - {", ".join(tabla["headers"][:4])}...')

print(f'\n{"=" * 100}')

# Guardar
output_json = 'scripts/analisis_excel_detectando_duplicados.json'
with open(output_json, 'w', encoding='utf-8') as f:
    json.dump(resultado, f, ensure_ascii=False, indent=2, default=str)

print(f'\n✅ Guardado en: {output_json}')
