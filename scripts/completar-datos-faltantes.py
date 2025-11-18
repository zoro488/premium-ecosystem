"""
📦 COMPLETAR DATOS FALTANTES DEL EXCEL
=======================================
Extrae ORDENES_COMPRA y ALMACEN_MONTE completos desde Excel
"""

import openpyxl
import json
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r'C:\Users\xpovo\Documents\premium-ecosystem\Administación_General.xlsx'
OUTPUT_DIR = r'C:\Users\xpovo\Documents\premium-ecosystem\src\apps\FlowDistributor\data'

def leer_valor(celda):
    """Lee valor de celda"""
    if celda.value is None:
        return None
    if isinstance(celda.value, (int, float)):
        return celda.value
    if isinstance(celda.value, datetime):
        return celda.value.strftime('%Y-%m-%d')
    return str(celda.value).strip()

print('='*80)
print('📦 COMPLETAR DATOS FALTANTES - ORDENES_COMPRA Y ALMACEN_MONTE')
print('='*80)
print()

try:
    print('📂 Cargando Excel...')
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print('✅ Excel cargado')
    print()

    # ========================================================================
    # ORDENES_COMPRA desde Control_Maestro
    # ========================================================================
    print('📋 Extrayendo ORDENES_COMPRA desde Control_Maestro...')
    print('-'*80)

    ws = wb['Control_Maestro']
    ordenesCompra = []

    for row_idx in range(4, ws.max_row + 1):
        oc = leer_valor(ws.cell(row_idx, 2))  # Columna B: OC Relacionada
        cantidad = leer_valor(ws.cell(row_idx, 3))  # Columna C: Cantidad
        cliente = leer_valor(ws.cell(row_idx, 4))  # Columna D: Cliente

        # Solo registros con OC, Cantidad y Cliente
        if not oc or not cantidad or not cliente:
            continue

        orden = {
            'id': oc,
            'numeroOC': oc,
            'fecha': leer_valor(ws.cell(row_idx, 1)) or leer_valor(ws.cell(row_idx, 15)),
            'cliente': cliente,
            'cantidad': cantidad,
            'precioVenta': leer_valor(ws.cell(row_idx, 6)) or 0,
            'ingreso': leer_valor(ws.cell(row_idx, 7)) or 0,
            'flete': leer_valor(ws.cell(row_idx, 8)) or 0,
            'utilidad': leer_valor(ws.cell(row_idx, 10)) or 0,
            'estatus': leer_valor(ws.cell(row_idx, 11)) or 'Pendiente',
            'observaciones': leer_valor(ws.cell(row_idx, 22)) or '',
        }

        ordenesCompra.append(orden)

        if len(ordenesCompra) % 20 == 0:
            print(f'  📌 {len(ordenesCompra)} órdenes procesadas...')

    print()
    print(f'✅ ORDENES_COMPRA: {len(ordenesCompra)} registros extraídos')
    print()

    # Guardar
    output_file = f'{OUTPUT_DIR}/ordenes-compra-excel.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(ordenesCompra, f, ensure_ascii=False, indent=2)
    print(f'💾 Guardado: {output_file}')
    print()

    # ========================================================================
    # ALMACEN_MONTE
    # ========================================================================
    print('📦 Extrayendo ALMACEN_MONTE...')
    print('-'*80)

    ws = wb['Almacen_Monte']
    almacen = []

    for row_idx in range(4, ws.max_row + 1):
        cliente = leer_valor(ws.cell(row_idx, 2))  # Columna B: Cliente

        # Solo registros con Cliente
        if not cliente:
            continue

        registro = {
            'id': f'alm-{len(almacen) + 1}',
            'oc': leer_valor(ws.cell(row_idx, 1)) or '',
            'cliente': cliente,
            'distribuidor': leer_valor(ws.cell(row_idx, 3)) or '',
            'cantidad': leer_valor(ws.cell(row_idx, 4)) or 0,
            'fechaEntrada': leer_valor(ws.cell(row_idx, 5)),
            'corte': leer_valor(ws.cell(row_idx, 6)) or '',
            'fechaSalida': leer_valor(ws.cell(row_idx, 7)),
            'clienteSalida': leer_valor(ws.cell(row_idx, 8)) or cliente,
            'cantidadSalida': leer_valor(ws.cell(row_idx, 9)) or 0,
            'concepto': leer_valor(ws.cell(row_idx, 10)) or '',
            'observaciones': leer_valor(ws.cell(row_idx, 11)) or '',
        }

        almacen.append(registro)

        if len(almacen) % 20 == 0:
            print(f'  📌 {len(almacen)} registros procesados...')

    print()
    print(f'✅ ALMACEN_MONTE: {len(almacen)} registros extraídos')
    print()

    # Guardar
    output_file = f'{OUTPUT_DIR}/almacen-monte-excel.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(almacen, f, ensure_ascii=False, indent=2)
    print(f'💾 Guardado: {output_file}')
    print()

    print('='*80)
    print('✅ EXTRACCIÓN COMPLETA')
    print('='*80)
    print(f'ORDENES_COMPRA: {len(ordenesCompra)} registros')
    print(f'ALMACEN_MONTE: {len(almacen)} registros')
    print(f'TOTAL: {len(ordenesCompra) + len(almacen)} registros')
    print('='*80)

except Exception as e:
    print(f'\n❌ Error: {e}')
    import traceback
    traceback.print_exc()
