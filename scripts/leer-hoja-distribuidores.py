#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Análisis detallado de la hoja Distribuidores del Excel
"""
import openpyxl
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')

# Cargar Excel
wb = openpyxl.load_workbook(r'C:\Users\xpovo\Documents\premium-ecosystem\Administación_General.xlsx', data_only=True)

print('=' * 100)
print('📊 ANÁLISIS COMPLETO: HOJA DISTRIBUIDORES')
print('=' * 100)

# Verificar nombre exacto de la hoja
print('\n📄 Hojas disponibles en el Excel:')
for idx, sheet_name in enumerate(wb.sheetnames, 1):
    print(f'   {idx}. {sheet_name}')

# Buscar la hoja de distribuidores
hoja_distribuidores = None
for name in wb.sheetnames:
    if 'distribuidor' in name.lower():
        hoja_distribuidores = name
        break

if not hoja_distribuidores:
    print('\n❌ ERROR: No se encontró hoja de Distribuidores')
    sys.exit(1)

ws = wb[hoja_distribuidores]

print(f'\n✅ Hoja encontrada: "{hoja_distribuidores}"')
print(f'   Dimensiones: {ws.max_row} filas x {ws.max_column} columnas')

# Buscar tablas en la hoja
print('\n' + '=' * 100)
print('🔍 BUSCANDO TABLAS EN LA HOJA...')
print('=' * 100)

# Función para detectar tablas
def encontrar_tablas(ws, max_filas=100):
    """Busca encabezados de tablas en la hoja"""
    tablas = []

    for fila in range(1, min(max_filas, ws.max_row + 1)):
        # Leer toda la fila
        valores_fila = []
        for col in range(1, min(20, ws.max_column + 1)):
            valor = ws.cell(fila, col).value
            if valor:
                valores_fila.append((col, str(valor)))

        # Si hay al menos 3 celdas con contenido, podría ser un encabezado
        if len(valores_fila) >= 3:
            # Verificar si parece encabezado (texto corto, sin números grandes)
            es_encabezado = True
            for col, valor in valores_fila:
                # Si tiene números muy grandes, probablemente no es encabezado
                try:
                    num = float(valor.replace(',', ''))
                    if num > 10000:
                        es_encabezado = False
                        break
                except:
                    pass

            if es_encabezado:
                tablas.append({
                    'fila_inicio': fila,
                    'encabezados': valores_fila,
                    'num_columnas': len(valores_fila)
                })

    return tablas

tablas_encontradas = encontrar_tablas(ws)

print(f'\n📋 Se encontraron {len(tablas_encontradas)} posibles tablas:\n')

for idx, tabla in enumerate(tablas_encontradas, 1):
    print(f'{idx}. TABLA en Fila {tabla["fila_inicio"]} ({tabla["num_columnas"]} columnas):')
    headers = [h[1] for h in tabla['encabezados']]
    print(f'   Encabezados: {" | ".join(headers)}')
    print('')

# Analizar cada tabla en detalle
print('\n' + '=' * 100)
print('📊 ANÁLISIS DETALLADO DE CADA TABLA')
print('=' * 100)

resultados = {
    'hoja': hoja_distribuidores,
    'tablas': []
}

for idx, tabla in enumerate(tablas_encontradas, 1):
    fila_header = tabla['fila_inicio']

    print(f'\n{"=" * 100}')
    print(f'TABLA {idx}: Fila {fila_header}')
    print('=' * 100)

    # Leer encabezados
    headers = []
    col_indices = []
    for col, valor in tabla['encabezados']:
        headers.append(valor)
        col_indices.append(col)

    print(f'\n📋 ENCABEZADOS ({len(headers)}):')
    for i, h in enumerate(headers, 1):
        print(f'   {i}. {h}')

    # Leer datos (hasta 50 filas después del header)
    registros = []
    fila_actual = fila_header + 1
    filas_vacias = 0

    while filas_vacias < 3 and fila_actual < fila_header + 100:
        # Leer fila
        fila_datos = {}
        tiene_datos = False

        for i, col in enumerate(col_indices):
            valor = ws.cell(fila_actual, col).value
            if valor is not None and str(valor).strip() != '':
                fila_datos[headers[i]] = valor
                tiene_datos = True

        if tiene_datos:
            registros.append(fila_datos)
            filas_vacias = 0
        else:
            filas_vacias += 1

        fila_actual += 1

    print(f'\n📊 REGISTROS: {len(registros)}')

    if registros:
        print(f'\n💾 PRIMEROS 3 REGISTROS:')
        for i, reg in enumerate(registros[:3], 1):
            print(f'\n   Registro {i}:')
            for key, val in reg.items():
                print(f'      {key}: {val}')

    # Guardar en resultados
    tabla_info = {
        'nombre': f'Tabla_{idx}',
        'fila_inicio': fila_header,
        'encabezados': headers,
        'total_registros': len(registros),
        'registros': registros
    }

    # Intentar identificar el tipo de tabla
    headers_lower = [h.lower() for h in headers]
    if any('orden' in h or 'oc' in h for h in headers_lower):
        tabla_info['tipo_identificado'] = 'ORDENES_COMPRA'
        tabla_info['descripcion'] = 'Tabla de Órdenes de Compra'
    elif any('distribuidor' in h and 'costo' in h for h in headers_lower):
        tabla_info['tipo_identificado'] = 'DISTRIBUIDORES_CONSOLIDADO'
        tabla_info['descripcion'] = 'Tabla consolidada de Distribuidores'
    else:
        tabla_info['tipo_identificado'] = 'DESCONOCIDO'
        tabla_info['descripcion'] = 'Tabla no identificada'

    print(f'\n🏷️  TIPO IDENTIFICADO: {tabla_info["tipo_identificado"]}')
    print(f'   {tabla_info["descripcion"]}')

    resultados['tablas'].append(tabla_info)

# Guardar resultados
with open('scripts/analisis-hoja-distribuidores.json', 'w', encoding='utf-8') as f:
    json.dump(resultados, f, ensure_ascii=False, indent=2)

print('\n' + '=' * 100)
print('📊 RESUMEN FINAL')
print('=' * 100)
print(f'\nHoja analizada: {hoja_distribuidores}')
print(f'Total de tablas encontradas: {len(tablas_encontradas)}')

for tabla in resultados['tablas']:
    print(f'\n• {tabla["tipo_identificado"]}:')
    print(f'  - Fila: {tabla["fila_inicio"]}')
    print(f'  - Columnas: {len(tabla["encabezados"])}')
    print(f'  - Registros: {tabla["total_registros"]}')
    print(f'  - Encabezados: {", ".join(tabla["encabezados"])}')

print('\n💾 Resultados guardados en: scripts/analisis-hoja-distribuidores.json')
print('\n' + '=' * 100)
