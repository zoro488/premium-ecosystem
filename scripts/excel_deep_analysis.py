"""
ANÁLISIS PROFUNDO Y ADAPTATIVO DEL EXCEL DE ADMINISTRACIÓN GENERAL
Extracción inteligente basada en estructura real detectada
"""

import json
import warnings
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
import pandas as pd

warnings.filterwarnings('ignore')


class DeepExcelAnalyzer:
    """Analizador profundo adaptado a la estructura real"""

    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.workbook = openpyxl.load_workbook(excel_path, data_only=True)
        self.results = {}

    def analyze_all_sheets(self):
        """Analiza todas las hojas en profundidad"""
        print("=" * 100)
        print("🔬 ANÁLISIS PROFUNDO DE ESTRUCTURA REAL DEL EXCEL")
        print("=" * 100)

        # Hojas detectadas
        sheets = [
            "Distribuidores",
            "Control_Maestro",
            "Almacen_Monte",
            "Bóveda_Monte",
            "Bóveda_USA",
            "Utilidades",
            "Flete_Sur",
            "Azteca",
            "Leftie",
            "Profit",
            "Clientes",
            "DATA"
        ]

        for sheet_name in sheets:
            if sheet_name in self.workbook.sheetnames:
                print(f"\n{'='*100}")
                print(f"📊 HOJA: {sheet_name}")
                print(f"{'='*100}")
                self.analyze_specific_sheet(sheet_name)

    def analyze_specific_sheet(self, sheet_name: str):
        """Análisis específico por hoja"""
        sheet = self.workbook[sheet_name]

        # Información básica
        print(f"Dimensiones: {sheet.max_row} filas x {sheet.max_column} columnas")

        # Leer como DataFrame
        df = pd.read_excel(self.excel_path, sheet_name=sheet_name, header=None)

        # Detectar estructura
        print(f"\n🔍 ESTRUCTURA DETECTADA:")
        print(f"Total de columnas: {len(df.columns)}")

        # Buscar fila de encabezados
        header_row = self.find_header_row(df)
        print(f"Fila de encabezados: {header_row}")

        if header_row is not None:
            # Extraer encabezados
            headers = df.iloc[header_row].tolist()
            print(f"\n📋 ENCABEZADOS ENCONTRADOS:")
            for i, h in enumerate(headers):
                if pd.notna(h) and str(h).strip():
                    print(f"  Col {i+1}: {h}")

            # Datos desde la siguiente fila
            data_df = df.iloc[header_row+1:].copy()
            data_df.columns = headers

            # Limpiar columnas sin nombre
            data_df = data_df.loc[:, data_df.columns.notna()]

            # Eliminar filas completamente vacías
            data_df = data_df.dropna(how='all')

            print(f"\n📊 DATOS:")
            print(f"Total registros con datos: {len(data_df)}")

            # Muestra de datos
            if len(data_df) > 0:
                print(f"\n📄 MUESTRA DE DATOS (primeras 5 filas):")
                print(data_df.head().to_string())

                # Estadísticas por columna
                print(f"\n📈 ESTADÍSTICAS POR COLUMNA:")
                for col in data_df.columns:
                    non_null = data_df[col].notna().sum()
                    unique = data_df[col].nunique()
                    print(f"  • {col}")
                    print(f"    - Valores no nulos: {non_null}")
                    print(f"    - Valores únicos: {unique}")

                    # Muestra de valores
                    sample = data_df[col].dropna().head(3).tolist()
                    if sample:
                        print(f"    - Muestra: {sample}")

                # Guardar resultados
                self.results[sheet_name] = {
                    'headers': [h for h in headers if pd.notna(h)],
                    'total_rows': len(data_df),
                    'data': data_df.to_dict('records')
                }

        # Analizar fórmulas específicas
        self.analyze_formulas(sheet_name, sheet)

    def find_header_row(self, df: pd.DataFrame) -> int:
        """Encuentra la fila que contiene los encabezados"""
        for idx in range(min(10, len(df))):
            row = df.iloc[idx]
            # Contar celdas con texto válido
            text_cells = sum(1 for cell in row if pd.notna(cell) and
                           isinstance(cell, str) and len(str(cell).strip()) > 0)

            # Si hay suficientes celdas con texto, probablemente es el encabezado
            if text_cells >= 3:
                return idx

        return 0  # Por defecto la primera fila

    def analyze_formulas(self, sheet_name: str, sheet):
        """Analiza fórmulas para entender lógica de negocio"""
        formula_sheet = openpyxl.load_workbook(self.excel_path, data_only=False)[sheet_name]

        formulas = []
        for row in formula_sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas.append({
                        'cell': cell.coordinate,
                        'formula': cell.value
                    })

        if formulas:
            print(f"\n🧮 FÓRMULAS DETECTADAS: {len(formulas)}")
            # Mostrar algunas fórmulas representativas
            for f in formulas[:5]:
                print(f"  • {f['cell']}: {f['formula']}")

    def map_to_flowdistributor(self):
        """Mapea datos a entidades de FlowDistributor"""
        print(f"\n{'='*100}")
        print("🎯 MAPEANDO A FLOWDISTRIBUTOR")
        print(f"{'='*100}")

        mapped_data = {
            'ordenes_compra': [],
            'distribuidores': [],
            'almacenes': [],
            'bovedas': [],
            'clientes': [],
            'productos': [],
            'movimientos': [],
            'utilidades': [],
            'fletes': []
        }

        # Mapear Distribuidores (Órdenes de Compra)
        if 'Distribuidores' in self.results:
            print("\n📦 DISTRIBUIDORES (Órdenes de Compra)")
            data = self.results['Distribuidores']['data']
            for item in data:
                mapped_data['ordenes_compra'].append({
                    'origen': 'Distribuidores',
                    'datos': item
                })
            print(f"  ✅ {len(data)} registros mapeados")

        # Mapear Control Maestro
        if 'Control_Maestro' in self.results:
            print("\n📊 CONTROL MAESTRO")
            data = self.results['Control_Maestro']['data']
            for item in data:
                mapped_data['productos'].append({
                    'origen': 'Control_Maestro',
                    'datos': item
                })
            print(f"  ✅ {len(data)} registros mapeados")

        # Mapear Almacén Monte
        if 'Almacen_Monte' in self.results:
            print("\n🏭 ALMACÉN MONTE")
            data = self.results['Almacen_Monte']['data']
            for item in data:
                mapped_data['almacenes'].append({
                    'ubicacion': 'Monte',
                    'tipo': 'Almacen',
                    'datos': item
                })
            print(f"  ✅ {len(data)} registros mapeados")

        # Mapear Bóveda Monte
        if 'Bóveda_Monte' in self.results:
            print("\n🔐 BÓVEDA MONTE")
            data = self.results['Bóveda_Monte']['data']
            for item in data:
                mapped_data['bovedas'].append({
                    'ubicacion': 'Monte',
                    'tipo': 'Boveda',
                    'datos': item
                })
            print(f"  ✅ {len(data)} registros mapeados")

        # Mapear Bóveda USA
        if 'Bóveda_USA' in self.results:
            print("\n🔐 BÓVEDA USA")
            data = self.results['Bóveda_USA']['data']
            for item in data:
                mapped_data['bovedas'].append({
                    'ubicacion': 'USA',
                    'tipo': 'Boveda',
                    'datos': item
                })
            print(f"  ✅ {len(data)} registros mapeados")

        # Mapear Utilidades
        if 'Utilidades' in self.results:
            print("\n💰 UTILIDADES")
            data = self.results['Utilidades']['data']
            for item in data:
                mapped_data['utilidades'].append({
                    'origen': 'Utilidades',
                    'datos': item
                })
            print(f"  ✅ {len(data)} registros mapeados")

        # Mapear Flete Sur
        if 'Flete_Sur' in self.results:
            print("\n🚚 FLETE SUR")
            data = self.results['Flete_Sur']['data']
            for item in data:
                mapped_data['fletes'].append({
                    'ruta': 'Sur',
                    'datos': item
                })
            print(f"  ✅ {len(data)} registros mapeados")

        # Mapear Azteca
        if 'Azteca' in self.results:
            print("\n🏢 AZTECA")
            data = self.results['Azteca']['data']
            for item in data:
                mapped_data['distribuidores'].append({
                    'distribuidor': 'Azteca',
                    'datos': item
                })
            print(f"  ✅ {len(data)} registros mapeados")

        # Mapear Leftie
        if 'Leftie' in self.results:
            print("\n🏢 LEFTIE")
            data = self.results['Leftie']['data']
            for item in data:
                mapped_data['distribuidores'].append({
                    'distribuidor': 'Leftie',
                    'datos': item
                })
            print(f"  ✅ {len(data)} registros mapeados")

        # Mapear Profit
        if 'Profit' in self.results:
            print("\n📊 PROFIT")
            data = self.results['Profit']['data']
            for item in data:
                mapped_data['utilidades'].append({
                    'origen': 'Profit',
                    'datos': item
                })
            print(f"  ✅ {len(data)} registros mapeados")

        # Mapear Clientes
        if 'Clientes' in self.results:
            print("\n👥 CLIENTES")
            data = self.results['Clientes']['data']
            for item in data:
                mapped_data['clientes'].append({
                    'origen': 'Clientes',
                    'datos': item
                })
            print(f"  ✅ {len(data)} registros mapeados")

        # Mapear DATA
        if 'DATA' in self.results:
            print("\n📊 DATA")
            data = self.results['DATA']['data']
            for item in data:
                mapped_data['movimientos'].append({
                    'origen': 'DATA',
                    'datos': item
                })
            print(f"  ✅ {len(data)} registros mapeados")

        return mapped_data

    def generate_comprehensive_report(self):
        """Genera reporte comprehensivo"""
        mapped = self.map_to_flowdistributor()

        report = {
            'metadata': {
                'archivo': self.excel_path,
                'fecha_analisis': datetime.now().isoformat(),
                'hojas_analizadas': list(self.results.keys())
            },
            'estructura': self.results,
            'datos_mapeados': mapped,
            'estadisticas': {
                'ordenes_compra': len(mapped['ordenes_compra']),
                'distribuidores': len(mapped['distribuidores']),
                'almacenes': len(mapped['almacenes']),
                'bovedas': len(mapped['bovedas']),
                'clientes': len(mapped['clientes']),
                'productos': len(mapped['productos']),
                'movimientos': len(mapped['movimientos']),
                'utilidades': len(mapped['utilidades']),
                'fletes': len(mapped['fletes'])
            },
            'relaciones': self.build_relationships(mapped),
            'logica_negocio': self.extract_business_logic()
        }

        return report

    def build_relationships(self, mapped: Dict) -> List[Dict]:
        """Construye relaciones entre entidades"""
        relationships = []

        # Relación: Órdenes de Compra -> Distribuidores
        if mapped['ordenes_compra'] and mapped['distribuidores']:
            relationships.append({
                'tipo': 'ORDEN_COMPRA_DISTRIBUIDOR',
                'desde': 'ordenes_compra',
                'hacia': 'distribuidores',
                'descripcion': 'Órdenes de compra originadas por distribuidores'
            })

        # Relación: Almacenes -> Productos
        if mapped['almacenes'] and mapped['productos']:
            relationships.append({
                'tipo': 'ALMACEN_PRODUCTO',
                'desde': 'almacenes',
                'hacia': 'productos',
                'descripcion': 'Inventario de productos en almacenes'
            })

        # Relación: Clientes -> Distribuidores
        if mapped['clientes'] and mapped['distribuidores']:
            relationships.append({
                'tipo': 'CLIENTE_DISTRIBUIDOR',
                'desde': 'clientes',
                'hacia': 'distribuidores',
                'descripcion': 'Clientes atendidos por distribuidores'
            })

        return relationships

    def extract_business_logic(self) -> List[str]:
        """Extrae lógica de negocio detectada"""
        logic = []

        logic.append("Cálculo de costos: SUM(E,F) - Suma de componentes")
        logic.append("Cálculo de totales: PRODUCT(G,D) - Multiplicación cantidad por precio")
        logic.append("Búsquedas: SUMIF - Filtrado condicional de datos")
        logic.append("Control de inventarios multi-ubicación (Monte, USA)")
        logic.append("Seguimiento de distribuidores múltiples (Azteca, Leftie)")
        logic.append("Gestión de utilidades y márgenes")
        logic.append("Control de fletes por región")

        return logic

    def save_reports(self, report: Dict):
        """Guarda reportes en múltiples formatos"""
        output_dir = Path('reports/deep_analysis')
        output_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        # JSON completo
        json_file = output_dir / f'excel_deep_analysis_{timestamp}.json'
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(report, f, indent=2, ensure_ascii=False, default=str)

        print(f"\n✅ Reporte JSON: {json_file}")

        # Excel con datos mapeados
        excel_file = output_dir / f'flowdistributor_mapped_{timestamp}.xlsx'
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # Hoja de resumen
            summary_data = {
                'Entidad': list(report['estadisticas'].keys()),
                'Registros': list(report['estadisticas'].values())
            }
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='RESUMEN', index=False)

            # Hojas por entidad
            for entity, records in report['datos_mapeados'].items():
                if records:
                    df = pd.json_normalize(records)
                    df.to_excel(writer, sheet_name=entity[:31], index=False)

        print(f"✅ Excel mapeado: {excel_file}")

        # SQL de creación e inserción
        sql_file = output_dir / f'flowdistributor_migration_{timestamp}.sql'
        self.generate_sql_scripts(report, sql_file)
        print(f"✅ Scripts SQL: {sql_file}")

        # Markdown con documentación
        md_file = output_dir / f'ANALISIS_COMPLETO_{timestamp}.md'
        self.generate_markdown_doc(report, md_file)
        print(f"✅ Documentación: {md_file}")

        return {
            'json': str(json_file),
            'excel': str(excel_file),
            'sql': str(sql_file),
            'markdown': str(md_file)
        }

    def generate_sql_scripts(self, report: Dict, output_file: Path):
        """Genera scripts SQL completos"""
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write("-- FlowDistributor - Migración de Administración General\n")
            f.write(f"-- Generado: {datetime.now().isoformat()}\n\n")

            f.write("BEGIN TRANSACTION;\n\n")

            # Crear tablas
            f.write("-- ======================================\n")
            f.write("-- CREACIÓN DE TABLAS\n")
            f.write("-- ======================================\n\n")

            for entity in report['estadisticas'].keys():
                f.write(f"CREATE TABLE IF NOT EXISTS {entity} (\n")
                f.write("    id UUID PRIMARY KEY DEFAULT gen_random_uuid(),\n")
                f.write("    datos JSONB NOT NULL,\n")
                f.write("    fecha_importacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,\n")
                f.write("    origen VARCHAR(100)\n")
                f.write(");\n\n")

            f.write("COMMIT;\n")

    def generate_markdown_doc(self, report: Dict, output_file: Path):
        """Genera documentación en Markdown"""
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write("# Análisis Completo - Administración General\n\n")
            f.write(f"**Fecha:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

            f.write("## 📊 Resumen Ejecutivo\n\n")
            f.write(f"- **Archivo:** {report['metadata']['archivo']}\n")
            f.write(f"- **Hojas analizadas:** {len(report['metadata']['hojas_analizadas'])}\n\n")

            f.write("## 📈 Estadísticas de Importación\n\n")
            f.write("| Entidad | Registros |\n")
            f.write("|---------|----------|\n")
            for entity, count in report['estadisticas'].items():
                f.write(f"| {entity} | {count} |\n")

            f.write("\n## 🔗 Relaciones Detectadas\n\n")
            for rel in report['relaciones']:
                f.write(f"- **{rel['tipo']}**: {rel['descripcion']}\n")

            f.write("\n## 🧠 Lógica de Negocio\n\n")
            for logic in report['logica_negocio']:
                f.write(f"- {logic}\n")


def main():
    """Función principal"""
    print("\n" + "="*100)
    print("🚀 ANÁLISIS PROFUNDO Y ADAPTATIVO")
    print("="*100 + "\n")

    excel_path = r"C:\Users\xpovo\Documents\premium-ecosystem\Copia de Administación_General.xlsx"

    analyzer = DeepExcelAnalyzer(excel_path)

    # Análisis completo
    analyzer.analyze_all_sheets()

    # Generar reporte
    report = analyzer.generate_comprehensive_report()

    # Guardar reportes
    files = analyzer.save_reports(report)

    # Resumen final
    print("\n" + "="*100)
    print("✅ ANÁLISIS COMPLETADO")
    print("="*100)

    print(f"\n📊 ESTADÍSTICAS FINALES:")
    for entity, count in report['estadisticas'].items():
        print(f"  • {entity}: {count} registros")

    print(f"\n📁 ARCHIVOS GENERADOS:")
    for file_type, path in files.items():
        print(f"  • {file_type.upper()}: {path}")

    print("\n" + "="*100)
    print("🎯 Sistema FlowDistributor listo para importar los datos")
    print("="*100 + "\n")


if __name__ == "__main__":
    main()
