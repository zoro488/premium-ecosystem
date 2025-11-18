#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para extraer TODAS las hojas del Excel de Administración General
Autor: GitHub Copilot
Fecha: 2025-11-03
"""

import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

# Ruta del Excel descargado
EXCEL_PATH = (
    Path(__file__).parent.parent
    / "Zoom Glitch Logo - Square"
    / "Copia de Administación_General.xlsx"
)
OUTPUT_PATH = Path(__file__).parent.parent / "BASE_DATOS_excel_data_COMPLETO.json"


def serialize_value(value):
    """Serializa valores para JSON incluyendo datetime"""
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def has_meaningful_data(row):
    """Verifica si una fila tiene datos significativos (no solo None/vacío)"""
    if not row:
        return False
    # Verificar si hay al menos un valor que no sea None, vacío o espacios
    return any(
        value is not None
        and str(value).strip() != ""
        and str(value).strip().lower() != "none"
        for value in row
    )


def extract_sheet_data(sheet, sheet_name):
    """Extrae todos los datos de una hoja con headers"""
    print(f"📊 Extrayendo hoja: {sheet_name}")

    # Obtener todas las filas
    rows = list(sheet.values)

    if not rows:
        print(f"⚠️  Hoja '{sheet_name}' está vacía")
        return []

    # Primera fila como headers
    headers = rows[0]

    # Verificar si los headers son válidos
    if not has_meaningful_data(headers):
        print(f"⚠️  Hoja '{sheet_name}' no tiene headers válidos")
        return []

    data = []
    empty_rows = 0

    # Procesar cada fila de datos
    for row_idx, row in enumerate(rows[1:], start=2):
        # Ignorar SOLO si la fila está completamente vacía
        if not has_meaningful_data(row):
            empty_rows += 1
            continue

        # Si la fila tiene al menos un dato, procesarla completa
        row_data = {}

        for col_idx, (header, value) in enumerate(zip(headers, row)):
            if header:  # Solo si hay header
                # Serializar valor (mantener None/null para campos vacíos)
                serialized = serialize_value(value)

                # Incluir TODOS los campos, incluso vacíos/None
                # porque representan "no aplica" para este registro
                row_data[str(header)] = serialized

        # Agregar la fila completa (con todos sus campos)
        if row_data:
            data.append(row_data)

    if empty_rows > 0:
        print(f"   ℹ️  {empty_rows} filas vacías ignoradas")
    print(f"   ✅ {len(data)} registros con datos extraídos")
    return data


def extract_all_sheets():
    """Extrae todas las hojas del Excel"""

    print("=" * 80)
    print("🚀 EXTRACCIÓN COMPLETA DE EXCEL - ADMINISTRACIÓN GENERAL")
    print("=" * 80)
    print(f"📁 Archivo: {EXCEL_PATH}")
    print(f"📂 Output: {OUTPUT_PATH}")
    print()

    # Verificar que existe el archivo
    if not EXCEL_PATH.exists():
        print(f"❌ ERROR: No se encuentra el archivo {EXCEL_PATH}")
        sys.exit(1)

    # Cargar el workbook
    print("📖 Cargando Excel...")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    except Exception as e:
        print(f"❌ ERROR al cargar Excel: {e}")
        sys.exit(1)

    # Determinar hojas a procesar (excluir 'DATA')
    sheetnames_to_process = [
        s for s in wb.sheetnames if s and s.strip().lower() != "data"
    ]

    print(
        f"✅ Excel cargado - {len(wb.sheetnames)} hojas encontradas ({len(sheetnames_to_process)} a procesar)"
    )
    print()

    # Estructura del JSON completo (solo hojas procesadas)
    complete_data = {
        "metadata": {
            "archivo": "Copia de Administación_General.xlsx",
            "fecha_extraccion": datetime.now().isoformat(),
            "total_hojas": len(sheetnames_to_process),
            "hojas": sheetnames_to_process,
            "hojas_excluidas": [
                s for s in wb.sheetnames if s and s.strip().lower() == "data"
            ],
        },
        "datos": {},
    }

    # Extraer cada hoja (ignorando 'DATA')
    for sheet_name in sheetnames_to_process:
        sheet = wb[sheet_name]

        # Normalizar nombre para usar como key
        key_name = sheet_name.lower().replace(" ", "_")
        key_name = key_name.replace("ó", "o").replace("á", "a")

        # Extraer datos
        sheet_data = extract_sheet_data(sheet, sheet_name)

        # Guardar en estructura
        complete_data["datos"][key_name] = {
            "nombre_original": sheet_name,
            "total_registros": len(sheet_data),
            "registros": sheet_data,
        }

    # Guardar JSON
    print()
    print("💾 Guardando JSON completo...")
    try:
        with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
            json.dump(complete_data, f, ensure_ascii=False, indent=2)
        print(f"✅ JSON guardado exitosamente: {OUTPUT_PATH}")
    except Exception as e:
        print(f"❌ ERROR al guardar JSON: {e}")
        sys.exit(1)

    # Resumen final
    print()
    print("=" * 80)
    print("📊 RESUMEN DE EXTRACCIÓN")
    print("=" * 80)

    total_records = 0
    for key, sheet_data in complete_data["datos"].items():
        records = sheet_data["total_registros"]
        total_records += records
        print(f"   {sheet_data['nombre_original']:30s} → {records:5d} registros")

    print("=" * 80)
    print(
        f"   TOTAL: {total_records} registros extraídos de {len(wb.sheetnames)} hojas"
    )
    print("=" * 80)
    print()
    print("✅ EXTRACCIÓN COMPLETADA EXITOSAMENTE")
    print()

    # Crear mapeo a paneles de FlowDistributor
    panel_mapping = {
        "ordenes_de_compra": "PanelOrdenesCompra",
        "profit": "PanelProfit",
        "leftie": "PanelLeftie",
        "azteca": "PanelAzteca",
        "boveda_monte": "PanelBovedaMonte",
        "boveda_usa": "PanelBovedaUSA",
        "flete_sur": "PanelFleteSur",
        "control_maestro": "PanelControlMaestro",
        "gya": "PanelGYA",
        "distribuidores": "PanelDistribuidores",
        "ventas_locales": "PanelVentasLocales",
        "clientes": "PanelClientes",
    }

    print("🗺️  MAPEO DE HOJAS A PANELES:")
    print("-" * 80)
    for sheet_key, panel_name in panel_mapping.items():
        if sheet_key in complete_data["datos"]:
            records = complete_data["datos"][sheet_key]["total_registros"]
            print(f"   {sheet_key:30s} → {panel_name:30s} ({records} registros)")
        else:
            print(f"   {sheet_key:30s} → {panel_name:30s} (NO ENCONTRADA)")

    print("-" * 80)

    return complete_data


if __name__ == "__main__":
    try:
        complete_data = extract_all_sheets()

        print()
        print("🎉 PROCESO COMPLETADO EXITOSAMENTE")
        print()
        print("📝 Siguiente paso:")
        print("   Ejecutar: python scripts/comparar-datos-json.py")
        print("   Para comparar con BASE_DATOS_excel_data.json actual")

    except Exception as e:
        print(f"\n❌ ERROR CRÍTICO: {e}")
        import traceback

        traceback.print_exc()
        sys.exit(1)
        sys.exit(1)
