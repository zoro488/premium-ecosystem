#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Análisis DETALLADO de Control_Maestro
Identifica columnas exactas para Ventas y Gastos/Abonos
"""
import openpyxl
import json
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\xpovo\Documents\premium-ecosystem\Administación_General.xlsx', data_only=True)
ws = wb['Control_Maestro']

print('=' * 120)
print('📊 ANÁLISIS DETALLADO: CONTROL_MAESTRO')
print('=' * 120)

print(f'\n📐 Dimensiones: {ws.max_row} filas x {ws.max_column} columnas')

# Leer fila de encabezados (fila 3)
print('\n📋 ANALIZANDO HEADERS (Fila 3):')
print('-' * 120)

headers = []
for col in range(1, ws.max_column + 1):
    valor = ws.cell(3, col).value
    letra = openpyxl.utils.get_column_letter(col)
    headers.append({
        'col_numero': col,
        'col_letra': letra,
        'nombre': valor if valor else f'[Vacía_{letra}]'
    })
    if col <= 25:  # Mostrar primeras 25 columnas
        print(f'{col:2d}. {letra:3s} → {valor if valor else "[Vacía]"}')

# Identificar secciones
print('\n' + '=' * 120)
print('🔍 IDENTIFICACIÓN DE SECCIONES')
print('=' * 120)

# Buscar columnas con "Panel" o separadores
separadores = []
for h in headers:
    if h['nombre'] and ('panel' in str(h['nombre']).lower() or
                        'rf actual' in str(h['nombre']).lower() or
                        str(h['nombre']).strip() == ''):
        separadores.append(h)

print('\n📌 Separadores encontrados:')
for sep in separadores:
    print(f"   Col {sep['col_numero']} ({sep['col_letra']}): {sep['nombre']}")

# Dividir en secciones
print('\n📊 DIVISIÓN DE SECCIONES:')

# SECCIÓN 1: VENTAS (hasta encontrar separador)
ventas_end = 12  # Asumiendo que Ventas termina en col 12
for sep in separadores:
    if sep['col_numero'] < 15:
        ventas_end = sep['col_numero'] - 1
        break

ventas_headers = headers[0:ventas_end]
print(f'\n✅ SECCIÓN VENTAS (Columnas 1-{ventas_end}):')
for h in ventas_headers:
    print(f"   {h['col_numero']:2d}. {h['col_letra']:3s} → {h['nombre']}")

# SECCIÓN 2: GASTOS Y ABONOS (después del separador)
gya_start = ventas_end + 1
# Buscar hasta dónde hay headers válidos
gya_end = ws.max_column
for col in range(gya_start, ws.max_column + 1):
    if not ws.cell(3, col).value:
        gya_end = col - 1
        break

gya_headers = headers[gya_start-1:gya_end]
print(f'\n✅ SECCIÓN GASTOS Y ABONOS (Columnas {gya_start}-{gya_end}):')
for h in gya_headers:
    print(f"   {h['col_numero']:2d}. {h['col_letra']:3s} → {h['nombre']}")

# Leer datos de ejemplo (primeras 5 filas)
print('\n' + '=' * 120)
print('📝 DATOS DE EJEMPLO (Primeras 3 ventas)')
print('=' * 120)

ventas_ejemplo = []
for fila in range(4, 7):  # Filas 4-6
    venta = {}
    for h in ventas_headers:
        valor = ws.cell(fila, h['col_numero']).value
        if isinstance(valor, datetime):
            valor = valor.strftime('%Y-%m-%d')
        venta[h['nombre']] = valor
    ventas_ejemplo.append(venta)

    print(f'\nVenta {fila - 3}:')
    for key, val in venta.items():
        if val is not None:
            print(f'   {key}: {val}')

print('\n' + '=' * 120)
print('📝 DATOS DE EJEMPLO (Primeros 3 gastos/abonos)')
print('=' * 120)

gya_ejemplo = []
for fila in range(4, 7):
    gya = {}
    for h in gya_headers:
        valor = ws.cell(fila, h['col_numero']).value
        if isinstance(valor, datetime):
            valor = valor.strftime('%Y-%m-%d')
        gya[h['nombre']] = valor
    gya_ejemplo.append(gya)

    print(f'\nGasto/Abono {fila - 3}:')
    for key, val in gya.items():
        if val is not None:
            print(f'   {key}: {val}')

# Contar registros reales
total_ventas = 0
total_gya = 0

for fila in range(4, min(300, ws.max_row + 1)):
    # Verificar si hay venta (col 1 tiene fecha)
    if ws.cell(fila, 1).value:
        total_ventas += 1

    # Verificar si hay gasto/abono (col gya_start tiene fecha)
    if ws.cell(fila, gya_start).value:
        total_gya += 1

print('\n' + '=' * 120)
print('📊 TOTALES')
print('=' * 120)
print(f'\nTotal VENTAS: {total_ventas}')
print(f'Total GASTOS/ABONOS: {total_gya}')

# Guardar resultados
resultados = {
    'ventas': {
        'columnas': [{'numero': h['col_numero'], 'letra': h['col_letra'], 'nombre': h['nombre']} for h in ventas_headers],
        'total_columnas': len(ventas_headers),
        'total_registros': total_ventas,
        'ejemplo': ventas_ejemplo
    },
    'gastos_abonos': {
        'columnas': [{'numero': h['col_numero'], 'letra': h['col_letra'], 'nombre': h['nombre']} for h in gya_headers],
        'total_columnas': len(gya_headers),
        'total_registros': total_gya,
        'ejemplo': gya_ejemplo
    }
}

with open('scripts/analisis-control-maestro-detallado.json', 'w', encoding='utf-8') as f:
    json.dump(resultados, f, ensure_ascii=False, indent=2)

print('\n💾 Resultados guardados en: scripts/analisis-control-maestro-detallado.json')
print('\n' + '=' * 120 + '\n')
