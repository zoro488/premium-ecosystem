"""
Análisis COMPLETO del Excel - Todas las hojas, todas las tablas, todos los registros
"""
import openpyxl
import sys
import json
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

# Excel a analizar
excel_path = r'C:\Users\xpovo\Downloads\Copia de Administación_General.xlsx'

print('=' * 100)
print('ANÁLISIS COMPLETO DEL EXCEL - TODAS LAS HOJAS Y REGISTROS')
print('=' * 100)
print(f'\nArchivo: {excel_path}')
print(f'Fecha de análisis: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
print()

# Cargar Excel
wb = openpyxl.load_workbook(excel_path, data_only=True)

print(f'Total de hojas: {len(wb.sheetnames)}')
print('\nHojas encontradas:')
for i, name in enumerate(wb.sheetnames, 1):
    print(f'  {i}. {name}')

# Diccionario para almacenar todos los datos
datos_completos = {}

# ============================================================================
# FUNCIÓN HELPER: Extraer todos los datos de una tabla
# ============================================================================
def extraer_tabla(ws, col_inicio, fila_inicio, num_columnas, nombre_tabla):
    """Extrae todos los datos de una tabla"""
    datos = []

    # Leer headers
    headers = []
    for col in range(col_inicio, col_inicio + num_columnas):
        header = ws.cell(fila_inicio, col).value
        headers.append(header if header else f'Col{col}')

    # Leer datos
    fila = fila_inicio + 1
    while True:
        # Verificar si la fila está vacía
        primera_celda = ws.cell(fila, col_inicio).value
        if not primera_celda or str(primera_celda).strip() == '':
            break

        # Extraer datos de la fila
        registro = {}
        for i, col in enumerate(range(col_inicio, col_inicio + num_columnas)):
            valor = ws.cell(fila, col).value
            registro[headers[i]] = valor

        datos.append(registro)
        fila += 1

        # Límite de seguridad
        if fila > 500:
            break

    return {
        'headers': headers,
        'registros': datos,
        'total': len(datos)
    }

# ============================================================================
# 1. HOJA: Distribuidores (Órdenes de Compra)
# ============================================================================
print('\n' + '=' * 100)
print('HOJA 1: DISTRIBUIDORES (ÓRDENES DE COMPRA)')
print('=' * 100)

ws = wb['Distribuidores']
tabla_oc = extraer_tabla(ws, 1, 3, 11, 'Órdenes de Compra')

print(f'\n📊 Total de registros: {tabla_oc["total"]}')
print(f'\n📋 Columnas: {", ".join(tabla_oc["headers"])}')
print('\n📄 TODOS LOS REGISTROS:')
print('-' * 100)

for i, reg in enumerate(tabla_oc['registros'], 1):
    print(f'\n[{i}] {reg.get("OC", "N/A")}')
    for key, val in reg.items():
        print(f'    {key}: {val}')

datos_completos['distribuidores_oc'] = tabla_oc

# ============================================================================
# 2. HOJA: Control_Maestro (Ventas)
# ============================================================================
print('\n' + '=' * 100)
print('HOJA 2: CONTROL_MAESTRO (VENTAS)')
print('=' * 100)

ws = wb['Control_Maestro']
tabla_ventas = extraer_tabla(ws, 1, 3, 12, 'Ventas')

print(f'\n📊 Total de registros: {tabla_ventas["total"]}')
print(f'\n📋 Columnas: {", ".join(tabla_ventas["headers"])}')
print('\n📄 PRIMERAS 10 VENTAS:')
print('-' * 100)

for i, reg in enumerate(tabla_ventas['registros'][:10], 1):
    print(f'\n[{i}] Fecha: {reg.get("Fecha", "N/A")} | Cliente: {reg.get("Cliente", "N/A")}')
    for key, val in reg.items():
        if key not in ['Fecha', 'Cliente']:
            print(f'    {key}: {val}')

print(f'\n... y {tabla_ventas["total"] - 10} ventas más.')

datos_completos['control_maestro_ventas'] = tabla_ventas

# ============================================================================
# 3. HOJA: Almacen_Monte
# ============================================================================
print('\n' + '=' * 100)
print('HOJA 3: ALMACEN_MONTE')
print('=' * 100)

ws = wb['Almacen_Monte']

# Tabla 1: Ingresos (columnas A-D)
print('\n📦 TABLA 1: INGRESOS (Órdenes de Compra)')
print('-' * 100)
tabla_almacen_ingresos = extraer_tabla(ws, 1, 3, 4, 'Almacén Ingresos')
print(f'Total registros: {tabla_almacen_ingresos["total"]}')
print(f'Columnas: {", ".join(tabla_almacen_ingresos["headers"])}')
for i, reg in enumerate(tabla_almacen_ingresos['registros'][:5], 1):
    print(f'{i}. {reg}')
if tabla_almacen_ingresos["total"] > 5:
    print(f'... y {tabla_almacen_ingresos["total"] - 5} más.')

# Tabla 2: Salidas (columnas F-J)
print('\n📤 TABLA 2: SALIDAS')
print('-' * 100)
tabla_almacen_salidas = extraer_tabla(ws, 6, 3, 5, 'Almacén Salidas')
print(f'Total registros: {tabla_almacen_salidas["total"]}')
print(f'Columnas: {", ".join(tabla_almacen_salidas["headers"])}')
for i, reg in enumerate(tabla_almacen_salidas['registros'][:5], 1):
    print(f'{i}. {reg}')
if tabla_almacen_salidas["total"] > 5:
    print(f'... y {tabla_almacen_salidas["total"] - 5} más.')

# Tabla 3: RF Actual Cortes (columnas L-M)
print('\n📊 TABLA 3: RF ACTUAL CORTES')
print('-' * 100)
tabla_almacen_rf = extraer_tabla(ws, 12, 3, 2, 'RF Actual Cortes')
print(f'Total registros: {tabla_almacen_rf["total"]}')
print(f'Columnas: {", ".join(tabla_almacen_rf["headers"])}')
for i, reg in enumerate(tabla_almacen_rf['registros'], 1):
    print(f'{i}. {reg}')

datos_completos['almacen_monte'] = {
    'ingresos': tabla_almacen_ingresos,
    'salidas': tabla_almacen_salidas,
    'rf_cortes': tabla_almacen_rf
}

# ============================================================================
# 4. HOJA: Bóveda_Monte
# ============================================================================
print('\n' + '=' * 100)
print('HOJA 4: BÓVEDA_MONTE')
print('=' * 100)

ws = wb['Bóveda_Monte']

# Tabla 1: Ingresos (columnas A-D)
print('\n💰 TABLA 1: INGRESOS')
print('-' * 100)
tabla_boveda_ingresos = extraer_tabla(ws, 1, 3, 4, 'Bóveda Monte Ingresos')
print(f'Total registros: {tabla_boveda_ingresos["total"]}')
print(f'Columnas: {", ".join(tabla_boveda_ingresos["headers"])}')
for i, reg in enumerate(tabla_boveda_ingresos['registros'][:10], 1):
    print(f'{i}. {reg}')
if tabla_boveda_ingresos["total"] > 10:
    print(f'... y {tabla_boveda_ingresos["total"] - 10} más.')

# Tabla 2: Gastos (columnas F-I)
print('\n💸 TABLA 2: GASTOS')
print('-' * 100)
tabla_boveda_gastos = extraer_tabla(ws, 6, 3, 4, 'Bóveda Monte Gastos')
print(f'Total registros: {tabla_boveda_gastos["total"]}')
print(f'Columnas: {", ".join(tabla_boveda_gastos["headers"])}')
for i, reg in enumerate(tabla_boveda_gastos['registros'], 1):
    print(f'{i}. {reg}')

datos_completos['boveda_monte'] = {
    'ingresos': tabla_boveda_ingresos,
    'gastos': tabla_boveda_gastos
}

# ============================================================================
# 5. HOJA: Bóveda_USA
# ============================================================================
print('\n' + '=' * 100)
print('HOJA 5: BÓVEDA_USA')
print('=' * 100)

ws = wb['Bóveda_USA']

# Tabla 1: Ingresos
print('\n💰 TABLA 1: INGRESOS')
print('-' * 100)
tabla_usa_ingresos = extraer_tabla(ws, 1, 3, 4, 'Bóveda USA Ingresos')
print(f'Total registros: {tabla_usa_ingresos["total"]}')
print(f'Columnas: {", ".join(tabla_usa_ingresos["headers"])}')
for i, reg in enumerate(tabla_usa_ingresos['registros'], 1):
    print(f'{i}. {reg}')

# Tabla 2: Gastos
print('\n💸 TABLA 2: GASTOS')
print('-' * 100)
tabla_usa_gastos = extraer_tabla(ws, 6, 3, 4, 'Bóveda USA Gastos')
print(f'Total registros: {tabla_usa_gastos["total"]}')
print(f'Columnas: {", ".join(tabla_usa_gastos["headers"])}')
for i, reg in enumerate(tabla_usa_gastos['registros'], 1):
    print(f'{i}. {reg}')

datos_completos['boveda_usa'] = {
    'ingresos': tabla_usa_ingresos,
    'gastos': tabla_usa_gastos
}

# ============================================================================
# 6. HOJA: Flete_Sur
# ============================================================================
print('\n' + '=' * 100)
print('HOJA 6: FLETE_SUR')
print('=' * 100)

ws = wb['Flete_Sur']

print('\n💰 TABLA 1: INGRESOS')
print('-' * 100)
tabla_flete_ingresos = extraer_tabla(ws, 1, 3, 4, 'Flete Sur Ingresos')
print(f'Total registros: {tabla_flete_ingresos["total"]}')
print(f'Columnas: {", ".join(tabla_flete_ingresos["headers"])}')
for i, reg in enumerate(tabla_flete_ingresos['registros'], 1):
    print(f'{i}. {reg}')

datos_completos['flete_sur'] = {
    'ingresos': tabla_flete_ingresos
}

# ============================================================================
# 7. HOJA: Azteca
# ============================================================================
print('\n' + '=' * 100)
print('HOJA 7: AZTECA')
print('=' * 100)

ws = wb['Azteca']

print('\n💰 TABLA 1: INGRESOS')
print('-' * 100)
tabla_azteca_ingresos = extraer_tabla(ws, 1, 3, 4, 'Azteca Ingresos')
print(f'Total registros: {tabla_azteca_ingresos["total"]}')
if tabla_azteca_ingresos["total"] > 0:
    print(f'Columnas: {", ".join(tabla_azteca_ingresos["headers"])}')
    for i, reg in enumerate(tabla_azteca_ingresos['registros'], 1):
        print(f'{i}. {reg}')
else:
    print('⚠️ NO HAY REGISTROS en esta tabla')

print('\n💸 TABLA 2: GASTOS')
print('-' * 100)
tabla_azteca_gastos = extraer_tabla(ws, 6, 3, 4, 'Azteca Gastos')
print(f'Total registros: {tabla_azteca_gastos["total"]}')
if tabla_azteca_gastos["total"] > 0:
    print(f'Columnas: {", ".join(tabla_azteca_gastos["headers"])}')
    for i, reg in enumerate(tabla_azteca_gastos['registros'], 1):
        print(f'{i}. {reg}')
else:
    print('⚠️ NO HAY REGISTROS en esta tabla')

datos_completos['azteca'] = {
    'ingresos': tabla_azteca_ingresos,
    'gastos': tabla_azteca_gastos
}

# ============================================================================
# 8. HOJA: Leftie
# ============================================================================
print('\n' + '=' * 100)
print('HOJA 8: LEFTIE')
print('=' * 100)

ws = wb['Leftie']

print('\n💰 TABLA 1: INGRESOS')
print('-' * 100)
tabla_leftie_ingresos = extraer_tabla(ws, 1, 3, 4, 'Leftie Ingresos')
print(f'Total registros: {tabla_leftie_ingresos["total"]}')
print(f'Columnas: {", ".join(tabla_leftie_ingresos["headers"])}')
for i, reg in enumerate(tabla_leftie_ingresos['registros'], 1):
    print(f'{i}. {reg}')

print('\n💸 TABLA 2: GASTOS')
print('-' * 100)
tabla_leftie_gastos = extraer_tabla(ws, 6, 3, 4, 'Leftie Gastos')
print(f'Total registros: {tabla_leftie_gastos["total"]}')
print(f'Columnas: {", ".join(tabla_leftie_gastos["headers"])}')
for i, reg in enumerate(tabla_leftie_gastos['registros'], 1):
    print(f'{i}. {reg}')

datos_completos['leftie'] = {
    'ingresos': tabla_leftie_ingresos,
    'gastos': tabla_leftie_gastos
}

# ============================================================================
# 9. HOJA: Profit
# ============================================================================
print('\n' + '=' * 100)
print('HOJA 9: PROFIT')
print('=' * 100)

ws = wb['Profit']

print('\n💰 TABLA 1: INGRESOS')
print('-' * 100)
tabla_profit_ingresos = extraer_tabla(ws, 1, 3, 4, 'Profit Ingresos')
print(f'Total registros: {tabla_profit_ingresos["total"]}')
print(f'Columnas: {", ".join(tabla_profit_ingresos["headers"])}')
for i, reg in enumerate(tabla_profit_ingresos['registros'][:10], 1):
    print(f'{i}. {reg}')
if tabla_profit_ingresos["total"] > 10:
    print(f'... y {tabla_profit_ingresos["total"] - 10} más.')

datos_completos['profit'] = {
    'ingresos': tabla_profit_ingresos
}

# ============================================================================
# 10. HOJA: Clientes
# ============================================================================
print('\n' + '=' * 100)
print('HOJA 10: CLIENTES')
print('=' * 100)

ws = wb['Clientes']

tabla_clientes = extraer_tabla(ws, 1, 3, 7, 'Clientes')
print(f'\n📊 Total de registros: {tabla_clientes["total"]}')

if tabla_clientes["total"] > 0:
    print(f'\n📋 Columnas: {", ".join(tabla_clientes["headers"])}')
    print('\n📄 TODOS LOS CLIENTES:')
    print('-' * 100)
    for i, reg in enumerate(tabla_clientes['registros'], 1):
        print(f'\n[{i}] {reg.get("Cliente", "N/A")}')
        for key, val in reg.items():
            if key != 'Cliente':
                print(f'    {key}: {val}')
else:
    print('⚠️ NO HAY REGISTROS DE CLIENTES en esta hoja')
    print('(La hoja puede estar vacía o los datos pueden estar en otra ubicación)')

datos_completos['clientes'] = tabla_clientes

# ============================================================================
# 11. HOJA: Utilidades
# ============================================================================
print('\n' + '=' * 100)
print('HOJA 11: UTILIDADES')
print('=' * 100)

ws = wb['Utilidades']

print('\n💰 INGRESOS')
print('-' * 100)
tabla_util_ingresos = extraer_tabla(ws, 1, 3, 4, 'Utilidades Ingresos')
print(f'Total registros: {tabla_util_ingresos["total"]}')
if tabla_util_ingresos["total"] > 0:
    for i, reg in enumerate(tabla_util_ingresos['registros'][:5], 1):
        print(f'{i}. {reg}')

datos_completos['utilidades'] = {
    'ingresos': tabla_util_ingresos
}

# ============================================================================
# RESUMEN FINAL
# ============================================================================
print('\n' + '=' * 100)
print('RESUMEN FINAL - CONTEO DE REGISTROS POR HOJA')
print('=' * 100)

print('\n📊 TOTALES:')
print(f'  1. Distribuidores (OC):        {datos_completos["distribuidores_oc"]["total"]} registros')
print(f'  2. Control Maestro (Ventas):   {datos_completos["control_maestro_ventas"]["total"]} registros')
print(f'  3. Almacén Monte:')
print(f'     - Ingresos:                 {datos_completos["almacen_monte"]["ingresos"]["total"]} registros')
print(f'     - Salidas:                  {datos_completos["almacen_monte"]["salidas"]["total"]} registros')
print(f'     - RF Cortes:                {datos_completos["almacen_monte"]["rf_cortes"]["total"]} registros')
print(f'  4. Bóveda Monte:')
print(f'     - Ingresos:                 {datos_completos["boveda_monte"]["ingresos"]["total"]} registros')
print(f'     - Gastos:                   {datos_completos["boveda_monte"]["gastos"]["total"]} registros')
print(f'  5. Bóveda USA:')
print(f'     - Ingresos:                 {datos_completos["boveda_usa"]["ingresos"]["total"]} registros')
print(f'     - Gastos:                   {datos_completos["boveda_usa"]["gastos"]["total"]} registros')
print(f'  6. Flete Sur:')
print(f'     - Ingresos:                 {datos_completos["flete_sur"]["ingresos"]["total"]} registros')
print(f'  7. Azteca:')
print(f'     - Ingresos:                 {datos_completos["azteca"]["ingresos"]["total"]} registros')
print(f'     - Gastos:                   {datos_completos["azteca"]["gastos"]["total"]} registros')
print(f'  8. Leftie:')
print(f'     - Ingresos:                 {datos_completos["leftie"]["ingresos"]["total"]} registros')
print(f'     - Gastos:                   {datos_completos["leftie"]["gastos"]["total"]} registros')
print(f'  9. Profit:')
print(f'     - Ingresos:                 {datos_completos["profit"]["ingresos"]["total"]} registros')
print(f' 10. Clientes:                   {datos_completos["clientes"]["total"]} registros')

# Guardar en JSON
output_file = r'C:\Users\xpovo\Documents\premium-ecosystem\scripts\analisis_excel_completo.json'
with open(output_file, 'w', encoding='utf-8') as f:
    json.dump(datos_completos, f, ensure_ascii=False, indent=2, default=str)

print(f'\n💾 Datos guardados en: {output_file}')

print('\n' + '=' * 100)
print('ANÁLISIS COMPLETADO')
print('=' * 100)
