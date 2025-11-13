#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de diagnóstico para analizar la hoja Distribuidores
"""

import openpyxl
from pathlib import Path

EXCEL_PATH = Path(__file__).parent.parent / "Zoom Glitch Logo - Square" / "Copia de Administación_General.xlsx"

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# Analizar hoja Distribuidores
sheet = wb["Distribuidores"]

print("=" * 80)
print("🔍 DIAGNÓSTICO DE HOJA: Distribuidores")
print("=" * 80)
print()

# Obtener dimensiones
print(f"📏 Dimensiones de la hoja:")
print(f"   max_row: {sheet.max_row}")
print(f"   max_column: {sheet.max_column}")
print()

# Analizar filas
rows = list(sheet.values)
print(f"📊 Total de filas en sheet.values: {len(rows)}")
print()

# Mostrar headers
print("📋 Headers (fila 1):")
headers = rows[0] if rows else []
print(f"   {headers}")
print()

# Analizar filas con datos
print("🔍 Análisis de filas:")
filas_con_datos = 0
filas_completamente_vacias = 0
filas_con_algunos_datos = 0

for idx, row in enumerate(rows[1:], start=2):
    # Contar valores no vacíos
    valores_no_vacios = sum(1 for v in row if v is not None and str(v).strip() != '')

    if valores_no_vacios == 0:
        filas_completamente_vacias += 1
    elif valores_no_vacios == len(row):
        filas_con_datos += 1
    else:
        filas_con_algunos_datos += 1

    # Mostrar primeras 10 filas con datos
    if valores_no_vacios > 0 and filas_con_datos + filas_con_algunos_datos <= 10:
        print(f"   Fila {idx}: {row}")

print()
print("📊 RESUMEN:")
print(f"   Filas con TODOS los campos llenos: {filas_con_datos}")
print(f"   Filas con ALGUNOS campos llenos: {filas_con_algunos_datos}")
print(f"   Filas COMPLETAMENTE vacías: {filas_completamente_vacias}")
print(f"   TOTAL filas procesadas: {len(rows) - 1}")
print()

# Datos reales esperados
print("=" * 80)
print("✅ DATOS REALES ESPERADOS (según tu input):")
print("=" * 80)
print("   Distribuidores: 2 registros")
print("   - PACMAN: $6,142,500.00")
print("   - Q-MAYA: $6,098,400.00")
print()

print("=" * 80)
print("❌ ERROR IDENTIFICADO:")
print("=" * 80)
print(f"   El script extrajo: {filas_con_datos + filas_con_algunos_datos} registros")
print(f"   Debería extraer: 2 registros")
print(f"   Diferencia: {(filas_con_datos + filas_con_algunos_datos) - 2} registros de más")
print()
