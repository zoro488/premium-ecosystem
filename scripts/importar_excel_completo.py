"""
============================================
IMPORTADOR COMPLETO EXCEL -> JSON
Sistema de importación quirúrgica con reseteo total
============================================
"""
import json
import openpyxl
import sys
from datetime import datetime
from collections import defaultdict

# Configurar encoding para Windows
sys.stdout.reconfigure(encoding='utf-8')

class ImportadorExcelCompleto:
    def __init__(self, excel_path, json_output_path):
        self.excel_path = excel_path
        self.json_output_path = json_output_path
        self.wb = None
        self.data = {
            "ventas": [],
            "compras": [],
            "distribuidores": [],
            "clientes": [],
            "gastosAbonos": [],
            "bancos": [],
            "almacen": [],
            "movimientos": [],
            "metricasFinancieras": {
                "capitalTotal": 0,
                "inventarioActual": 0,
                "carteraPorCobrar": 0,
                "cuentasPorPagar": 0,
                "utilidadTotal": 0,
                "costoTotalInventario": 0,
                "ventasTotales": 0,
                "comprasTotales": 0
            },
            "resumen": {
                "ordenesCompra": 0,
                "distribuidores": 0,
                "ventasLocales": 0,
                "gastosAbonos": 0,
                "clientes": 0,
                "bancos": 0,
                "ingresosAlmacen": 0,
                "salidasAlmacen": 0,
                "movimientosBancarios": 0
            },
            "ultimaActualizacion": None,
            "version": "2.0.0",
            "estado": "importado"
        }
        self.stats = {
            "procesados": 0,
            "exitosos": 0,
            "errores": 0,
            "warnings": []
        }

    def log(self, mensaje, tipo="INFO"):
        prefijos = {"INFO": "ℹ️", "SUCCESS": "✅", "WARNING": "⚠️", "ERROR": "❌"}
        print(f"{prefijos.get(tipo, 'ℹ️')} {mensaje}")

    def safe_get_cell(self, ws, row, col, default=None):
        """Obtiene valor de celda de forma segura"""
        try:
            val = ws.cell(row, col).value
            return val if val is not None else default
        except:
            return default

    def safe_float(self, val, default=0.0):
        """Convierte a float de forma segura"""
        if val is None or val == '':
            return default
        try:
            return float(val)
        except:
            return default

    def safe_int(self, val, default=0):
        """Convierte a int de forma segura"""
        if val is None or val == '':
            return default
        try:
            return int(float(val))
        except:
            return default

    def safe_date(self, val):
        """Convierte fecha a string ISO"""
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.isoformat()
        try:
            return str(val)
        except:
            return None

    def cargar_excel(self):
        """Carga el archivo Excel"""
        self.log(f"Cargando Excel: {self.excel_path}")
        try:
            self.wb = openpyxl.load_workbook(self.excel_path, data_only=True)
            self.log(f"Excel cargado - Hojas: {', '.join(self.wb.sheetnames)}", "SUCCESS")
            return True
        except Exception as e:
            self.log(f"Error al cargar Excel: {e}", "ERROR")
            return False

    def procesar_distribuidores(self):
        """Procesa hoja de Distribuidores (Órdenes de Compra)"""
        self.log("Procesando Distribuidores...")

        if 'Distribuidores' not in self.wb.sheetnames:
            self.log("Hoja 'Distribuidores' no encontrada", "WARNING")
            return

        ws = self.wb['Distribuidores']
        distribuidores_set = set()

        # Headers en fila 3 (índice 3)
        # Columnas: A=OC, B=Fecha, C=Origen, D=Cantidad, E=Costo Dist, F=Costo Trans,
        #           G=Costo/Unidad, H=Stock Actual, I=Costo Total, J=Pago Dist, K=Deuda

        for row_idx in range(4, ws.max_row + 1):
            oc = self.safe_get_cell(ws, row_idx, 1)  # Columna A

            if not oc or str(oc).strip() == '':
                continue

            fecha = self.safe_date(self.safe_get_cell(ws, row_idx, 2))  # B
            origen = self.safe_get_cell(ws, row_idx, 3, "Sin Origen")  # C
            cantidad = self.safe_int(self.safe_get_cell(ws, row_idx, 4))  # D
            costo_distribuidor = self.safe_float(self.safe_get_cell(ws, row_idx, 5))  # E
            costo_transporte = self.safe_float(self.safe_get_cell(ws, row_idx, 6))  # F
            costo_por_unidad = self.safe_float(self.safe_get_cell(ws, row_idx, 7))  # G
            stock_actual = self.safe_int(self.safe_get_cell(ws, row_idx, 8))  # H
            costo_total = self.safe_float(self.safe_get_cell(ws, row_idx, 9))  # I
            pago_distribuidor = self.safe_float(self.safe_get_cell(ws, row_idx, 10))  # J
            deuda = self.safe_float(self.safe_get_cell(ws, row_idx, 11))  # K

            # Crear orden de compra
            compra = {
                "id": str(oc),
                "tipo": "compra",
                "fecha": fecha,
                "distribuidor": origen,
                "cantidad": cantidad,
                "costoDistribuidor": costo_distribuidor,
                "costoTransporte": costo_transporte,
                "costoPorUnidad": costo_por_unidad,
                "stockActual": stock_actual,
                "costoTotal": costo_total,
                "pagoDistribuidor": pago_distribuidor,
                "deuda": deuda,
                "estatus": "completada" if deuda <= 0 else "pendiente"
            }

            self.data["compras"].append(compra)
            distribuidores_set.add(origen)
            self.stats["procesados"] += 1
            self.stats["exitosos"] += 1

        # Crear registros de distribuidores únicos
        for dist_nombre in distribuidores_set:
            # Calcular totales para este distribuidor
            compras_dist = [c for c in self.data["compras"] if c["distribuidor"] == dist_nombre]
            total_comprado = sum(c["costoTotal"] for c in compras_dist)
            total_pagado = sum(c["pagoDistribuidor"] for c in compras_dist)
            adeudo = sum(c["deuda"] for c in compras_dist)

            self.data["distribuidores"].append({
                "id": f"DIST-{dist_nombre.upper().replace(' ', '-')}",
                "nombre": dist_nombre,
                "totalComprado": total_comprado,
                "totalPagado": total_pagado,
                "adeudo": adeudo,
                "ordenesCompra": len(compras_dist),
                "estado": "activo"
            })

        self.data["resumen"]["ordenesCompra"] = len(self.data["compras"])
        self.data["resumen"]["distribuidores"] = len(self.data["distribuidores"])
        self.log(f"✅ {len(self.data['compras'])} órdenes de compra procesadas", "SUCCESS")
        self.log(f"✅ {len(self.data['distribuidores'])} distribuidores únicos", "SUCCESS")

    def procesar_control_maestro(self):
        """Procesa hoja de Control Maestro (Ventas Locales)"""
        self.log("Procesando Control Maestro (Ventas)...")

        if 'Control_Maestro' not in self.wb.sheetnames:
            self.log("Hoja 'Control_Maestro' no encontrada", "WARNING")
            return

        ws = self.wb['Control_Maestro']

        # Headers en fila 3, datos desde fila 4
        # A=Fecha, B=OC Relacionada, C=Cantidad, D=Cliente, E=Bóveda Monte (destino)
        # F=Precio De Venta, G=Ingreso, H=Flete, I=Flete Utilidad, J=Utilidad, K=Estatus, L=Concepto

        for row_idx in range(4, ws.max_row + 1):
            fecha = self.safe_date(self.safe_get_cell(ws, row_idx, 1))  # A

            # Si no hay fecha, saltar
            if not fecha:
                continue

            oc = self.safe_get_cell(ws, row_idx, 2, "")  # B
            cantidad = self.safe_int(self.safe_get_cell(ws, row_idx, 3))  # C
            cliente = self.safe_get_cell(ws, row_idx, 4, "Cliente Genérico")  # D
            destino_monto = self.safe_float(self.safe_get_cell(ws, row_idx, 5))  # E
            precio = self.safe_float(self.safe_get_cell(ws, row_idx, 6))  # F
            ingreso_total = self.safe_float(self.safe_get_cell(ws, row_idx, 7))  # G
            flete_texto = self.safe_get_cell(ws, row_idx, 8, "")  # H
            flete_utilidad = self.safe_float(self.safe_get_cell(ws, row_idx, 9))  # I
            utilidades = self.safe_float(self.safe_get_cell(ws, row_idx, 10))  # J
            estatus = self.safe_get_cell(ws, row_idx, 11, "Pendiente")  # K
            concepto = self.safe_get_cell(ws, row_idx, 12, "")  # L

            # Generar ID único
            venta_id = f"VENTA-{fecha}-{cliente}-{row_idx}"

            # Determinar si aplica flete
            aplica_flete = flete_texto == "Aplica" if isinstance(flete_texto, str) else False

            venta = {
                "id": venta_id,
                "tipo": "venta",
                "fecha": fecha,
                "ocRelacionada": str(oc) if oc else "",
                "cantidad": cantidad,
                "cliente": str(cliente),
                "productos": [{
                    "nombre": "Producto",
                    "cantidad": cantidad,
                    "precio": precio,
                    "subtotal": ingreso_total
                }],
                "totalVenta": ingreso_total,
                "totalFletes": flete_utilidad if aplica_flete else 0,
                "totalUtilidades": utilidades,
                "estatus": str(estatus) if estatus else "Pendiente",
                "estadoPago": "pendiente",  # Por defecto pendiente
                "adeudo": 0,
                "montoPagado": 0,
                "destino": "bovedaMonte",
                "concepto": str(concepto) if concepto else "",
                "aplicaFlete": aplica_flete,
                "bovedaMonte": destino_monto
            }

            self.data["ventas"].append(venta)
            self.stats["procesados"] += 1
            self.stats["exitosos"] += 1

        self.data["resumen"]["ventasLocales"] = len(self.data["ventas"])
        self.log(f"✅ {len(self.data['ventas'])} ventas procesadas", "SUCCESS")

    def procesar_clientes(self):
        """Procesa hoja de Clientes"""
        self.log("Procesando Clientes...")

        if 'Clientes' not in self.wb.sheetnames:
            self.log("Hoja 'Clientes' no encontrada", "WARNING")
            return

        ws = self.wb['Clientes']
        clientes_dict = defaultdict(lambda: {
            "totalComprado": 0,
            "totalAbonado": 0,
            "adeudo": 0,
            "ventas": []
        })

        # Headers en fila 4, columna E empieza "Clientes"
        for row_idx in range(5, ws.max_row + 1):
            cliente = self.safe_get_cell(ws, row_idx, 5)  # E - Clientes

            if not cliente or str(cliente).strip() == '':
                continue

            compra = self.safe_float(self.safe_get_cell(ws, row_idx, 6))  # F
            abono = self.safe_float(self.safe_get_cell(ws, row_idx, 7))  # G
            adeudo = self.safe_float(self.safe_get_cell(ws, row_idx, 8))  # H

            clientes_dict[cliente]["totalComprado"] += compra
            clientes_dict[cliente]["totalAbonado"] += abono
            clientes_dict[cliente]["adeudo"] = adeudo  # Usar el último adeudo

        # Convertir a lista
        for nombre, datos in clientes_dict.items():
            nombre_str = str(nombre)
            self.data["clientes"].append({
                "id": f"CLI-{nombre_str.upper().replace(' ', '-')}",
                "nombre": nombre_str,
                "totalComprado": datos["totalComprado"],
                "totalAbonado": datos["totalAbonado"],
                "adeudo": datos["adeudo"],
                "estado": "activo",
                "observaciones": "",
                "ventas": []
            })

        self.data["resumen"]["clientes"] = len(self.data["clientes"])
        self.log(f"✅ {len(self.data['clientes'])} clientes procesados", "SUCCESS")

    def procesar_bancos(self):
        """Procesa hojas de bancos (Azteca, Leftie, Profit)"""
        self.log("Procesando Bancos...")

        bancos_hojas = {
            'Azteca': 'Banco Azteca',
            'Leftie': 'Banco Leftie',
            'Profit': 'Banco Profit'
        }

        for hoja, nombre_banco in bancos_hojas.items():
            if hoja not in self.wb.sheetnames:
                continue

            ws = self.wb[hoja]
            movimientos = []
            saldo_actual = 0

            # Headers en fila 2
            # A=Ingresos, columnas 8-9=RF Actual, columna 10-11=Gastos
            for row_idx in range(3, min(ws.max_row + 1, 100)):  # Limitar a 100 filas
                ingreso = self.safe_float(self.safe_get_cell(ws, row_idx, 1))  # A
                fecha_ingreso = self.safe_date(self.safe_get_cell(ws, row_idx, 2))  # B

                # RF Actual (saldo)
                rf_actual = self.safe_float(self.safe_get_cell(ws, row_idx, 8))  # H
                if rf_actual > 0:
                    saldo_actual = rf_actual

                # Gastos
                gasto = self.safe_float(self.safe_get_cell(ws, row_idx, 10))  # J
                fecha_gasto = self.safe_date(self.safe_get_cell(ws, row_idx, 11))  # K

                if ingreso > 0:
                    movimientos.append({
                        "id": f"MOV-{hoja}-ING-{row_idx}",
                        "tipo": "ingreso",
                        "monto": ingreso,
                        "fecha": fecha_ingreso,
                        "concepto": "Ingreso",
                        "banco": nombre_banco
                    })

                if gasto > 0:
                    movimientos.append({
                        "id": f"MOV-{hoja}-GAS-{row_idx}",
                        "tipo": "gasto",
                        "monto": gasto,
                        "fecha": fecha_gasto,
                        "concepto": "Gasto",
                        "banco": nombre_banco
                    })

            # Crear banco
            banco = {
                "id": f"BANCO-{hoja.upper()}",
                "nombre": nombre_banco,
                "saldoActual": saldo_actual,
                "totalIngresos": sum(m["monto"] for m in movimientos if m["tipo"] == "ingreso"),
                "totalGastos": sum(m["monto"] for m in movimientos if m["tipo"] == "gasto"),
                "movimientos": len(movimientos),
                "estado": "activo"
            }

            self.data["bancos"].append(banco)
            self.data["movimientos"].extend(movimientos)
            self.log(f"  ✅ {nombre_banco}: {len(movimientos)} movimientos, Saldo: ${saldo_actual:,.2f}")

        self.data["resumen"]["bancos"] = len(self.data["bancos"])
        self.data["resumen"]["movimientosBancarios"] = len(self.data["movimientos"])
        self.log(f"✅ {len(self.data['bancos'])} bancos procesados", "SUCCESS")

    def procesar_almacen(self):
        """Procesa hoja de Almacén Monte"""
        self.log("Procesando Almacén Monte...")

        if 'Almacen_Monte' not in self.wb.sheetnames:
            self.log("Hoja 'Almacen_Monte' no encontrada", "WARNING")
            return

        ws = self.wb['Almacen_Monte']
        ingresos = 0
        salidas = 0
        stock_actual = 0

        # Headers en fila 1: A=Ingresos, E=RF Actual, G=Salida
        for row_idx in range(2, min(ws.max_row + 1, 100)):
            ingreso = self.safe_int(self.safe_get_cell(ws, row_idx, 1))  # A
            rf_actual = self.safe_int(self.safe_get_cell(ws, row_idx, 5))  # E
            salida = self.safe_int(self.safe_get_cell(ws, row_idx, 7))  # G

            if ingreso > 0:
                ingresos += ingreso
            if salida > 0:
                salidas += salida
            if rf_actual > 0:
                stock_actual = rf_actual

        self.data["almacen"].append({
            "id": "ALMACEN-MONTE",
            "nombre": "Almacén Monte",
            "stockActual": stock_actual,
            "totalIngresos": ingresos,
            "totalSalidas": salidas,
            "ubicacion": "Monte",
            "estado": "activo"
        })

        self.data["resumen"]["ingresosAlmacen"] = ingresos
        self.data["resumen"]["salidasAlmacen"] = salidas
        self.log(f"✅ Almacén: {ingresos} ingresos, {salidas} salidas, Stock: {stock_actual}", "SUCCESS")

    def calcular_metricas(self):
        """Calcula métricas financieras finales"""
        self.log("Calculando métricas financieras...")

        # Capital total (suma de saldos bancarios)
        capital_total = sum(b["saldoActual"] for b in self.data["bancos"])

        # Inventario actual (suma de stocks)
        inventario_actual = sum(a["stockActual"] for a in self.data["almacen"])

        # Cartera por cobrar (adeudos de ventas)
        cartera_por_cobrar = sum(v["adeudo"] for v in self.data["ventas"])

        # Cuentas por pagar (deudas a distribuidores)
        cuentas_por_pagar = sum(c["deuda"] for c in self.data["compras"])

        # Utilidad total
        utilidad_total = sum(v["totalUtilidades"] for v in self.data["ventas"])

        # Costo total inventario
        costo_total_inventario = sum(c["costoTotal"] for c in self.data["compras"])

        # Ventas totales
        ventas_totales = sum(v["totalVenta"] for v in self.data["ventas"])

        # Compras totales
        compras_totales = sum(c["costoTotal"] for c in self.data["compras"])

        self.data["metricasFinancieras"] = {
            "capitalTotal": capital_total,
            "inventarioActual": inventario_actual,
            "carteraPorCobrar": cartera_por_cobrar,
            "cuentasPorPagar": cuentas_por_pagar,
            "utilidadTotal": utilidad_total,
            "costoTotalInventario": costo_total_inventario,
            "ventasTotales": ventas_totales,
            "comprasTotales": compras_totales
        }

        self.log(f"💰 Capital Total: ${capital_total:,.2f}", "SUCCESS")
        self.log(f"📦 Inventario: {inventario_actual} unidades", "SUCCESS")
        self.log(f"💵 Por Cobrar: ${cartera_por_cobrar:,.2f}", "SUCCESS")
        self.log(f"💳 Por Pagar: ${cuentas_por_pagar:,.2f}", "SUCCESS")

    def guardar_json(self):
        """Guarda datos en archivo JSON"""
        self.log(f"Guardando datos en {self.json_output_path}...")

        self.data["ultimaActualizacion"] = datetime.now().isoformat()

        try:
            with open(self.json_output_path, 'w', encoding='utf-8') as f:
                json.dump(self.data, f, ensure_ascii=False, indent=2)
            self.log("✅ JSON guardado exitosamente", "SUCCESS")
            return True
        except Exception as e:
            self.log(f"Error al guardar JSON: {e}", "ERROR")
            return False

    def ejecutar(self):
        """Ejecuta el proceso completo de importación"""
        self.log("=" * 80)
        self.log("INICIANDO IMPORTACIÓN COMPLETA DEL EXCEL", "SUCCESS")
        self.log("=" * 80)

        if not self.cargar_excel():
            return False

        self.procesar_distribuidores()
        self.procesar_control_maestro()
        self.procesar_clientes()
        self.procesar_bancos()
        self.procesar_almacen()
        self.calcular_metricas()

        if not self.guardar_json():
            return False

        self.log("=" * 80)
        self.log("IMPORTACIÓN COMPLETADA EXITOSAMENTE", "SUCCESS")
        self.log("=" * 80)
        self.log(f"📊 Registros procesados: {self.stats['procesados']}")
        self.log(f"✅ Exitosos: {self.stats['exitosos']}")
        self.log(f"❌ Errores: {self.stats['errores']}")
        self.log(f"⚠️  Warnings: {len(self.stats['warnings'])}")

        return True

if __name__ == "__main__":
    excel_path = r'c:\Users\xpovo\Documents\premium-ecosystem\Copia de Administación_General.xlsx'
    json_path = r'c:\Users\xpovo\Documents\premium-ecosystem\public\excel_data.json'

    importador = ImportadorExcelCompleto(excel_path, json_path)
    importador.ejecutar()
