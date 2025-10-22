import openpyxl
import json
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r'C:\Users\xpovo\Documents\premium-ecosystem\Copia de Administación_General.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

print('EXTRACCIÓN COMPLETA DEL EXCEL')
print('='*80)

def limpiar_valor(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, (int, float)):
        return val
    return str(val).strip()

datos_completos = {}

# ==============================================================================
# 1. BANCOS - EXTRAER TODOS CON SUS MOVIMIENTOS
# ==============================================================================
print('\n💰 EXTRAYENDO BANCOS...')

bancos_config = {
    'Bóveda_Monte': {
        'nombre': 'bovedaMonte',
        'col_ingreso_fecha': 1,
        'col_ingreso_cliente': 2,
        'col_ingreso_monto': 3,
        'col_ingreso_concepto': 4,
        'col_gasto_fecha': 7,
        'col_gasto_origen': 8,
        'col_gasto_monto': 9,
        'col_gasto_concepto': 13,
        'col_rf_actual': 5,
        'fila_rf': 2
    },
    'Bóveda_USA': {
        'nombre': 'bovedaUsa',
        'col_ingreso_fecha': 1,
        'col_ingreso_cliente': 2,
        'col_ingreso_monto': 3,
        'col_ingreso_concepto': 7,
        'col_gasto_fecha': 11,
        'col_gasto_origen': 12,
        'col_gasto_monto': 13,
        'col_gasto_concepto': 17,
        'col_rf_actual': 9,
        'fila_rf': 2
    },
    'Utilidades': {
        'nombre': 'utilidades',
        'col_ingreso_fecha': 1,
        'col_ingreso_cliente': 2,
        'col_ingreso_monto': 3,
        'col_ingreso_concepto': 4,
        'col_gasto_fecha': 7,
        'col_gasto_origen': 8,
        'col_gasto_monto': 9,
        'col_gasto_concepto': 12,
        'col_rf_actual': 5,
        'fila_rf': 2
    },
    'Flete_Sur': {
        'nombre': 'fleteSur',
        'col_ingreso_fecha': 1,
        'col_ingreso_cliente': 2,
        'col_ingreso_monto': 3,
        'col_ingreso_concepto': 4,
        'col_gasto_fecha': 7,
        'col_gasto_origen': 8,
        'col_gasto_monto': 9,
        'col_gasto_concepto': 13,
        'col_rf_actual': 5,
        'fila_rf': 2
    },
    'Azteca': {
        'nombre': 'azteca',
        'col_ingreso_fecha': 1,
        'col_ingreso_cliente': 2,
        'col_ingreso_monto': 3,
        'col_ingreso_concepto': 7,
        'col_gasto_fecha': 10,
        'col_gasto_origen': 11,
        'col_gasto_monto': 12,
        'col_gasto_concepto': 16,
        'col_rf_actual': 8,
        'fila_rf': 2
    },
    'Leftie': {
        'nombre': 'leftie',
        'col_ingreso_fecha': 1,
        'col_ingreso_cliente': 2,
        'col_ingreso_monto': 3,
        'col_ingreso_concepto': 7,
        'col_gasto_fecha': 11,
        'col_gasto_origen': 12,
        'col_gasto_monto': 13,
        'col_gasto_concepto': 17,
        'col_rf_actual': 9,
        'fila_rf': 2
    },
    'Profit': {
        'nombre': 'profit',
        'col_ingreso_fecha': 1,
        'col_ingreso_cliente': 2,
        'col_ingreso_monto': 3,
        'col_ingreso_concepto': 7,
        'col_gasto_fecha': 11,
        'col_gasto_origen': 12,
        'col_gasto_monto': 15,
        'col_gasto_concepto': 7,
        'col_rf_actual': 9,
        'fila_rf': 2
    }
}

bancos = {}

for hoja_nombre, config in bancos_config.items():
    ws = wb[hoja_nombre]

    # Obtener RF Actual (saldo actual)
    rf_actual = ws.cell(config['fila_rf'], config['col_rf_actual']).value or 0

    # Extraer ingresos
    ingresos = []
    for row in range(4, min(ws.max_row + 1, 200)):  # Limitar a 200 filas
        fecha = ws.cell(row, config['col_ingreso_fecha']).value
        monto = ws.cell(row, config['col_ingreso_monto']).value

        if fecha and monto and monto > 0:
            ingreso = {
                'fecha': limpiar_valor(fecha),
                'cliente': limpiar_valor(ws.cell(row, config['col_ingreso_cliente']).value),
                'monto': limpiar_valor(monto),
                'concepto': limpiar_valor(ws.cell(row, config['col_ingreso_concepto']).value) or ''
            }
            ingresos.append(ingreso)

    # Extraer gastos
    gastos = []
    for row in range(4, min(ws.max_row + 1, 200)):
        fecha = ws.cell(row, config['col_gasto_fecha']).value
        monto = ws.cell(row, config['col_gasto_monto']).value

        if fecha and monto and monto > 0:
            gasto = {
                'fecha': limpiar_valor(fecha),
                'origen': limpiar_valor(ws.cell(row, config['col_gasto_origen']).value),
                'monto': limpiar_valor(monto),
                'concepto': limpiar_valor(ws.cell(row, config['col_gasto_concepto']).value) or ''
            }
            gastos.append(gasto)

    bancos[config['nombre']] = {
        'nombre': config['nombre'],
        'saldoActual': rf_actual,
        'ingresos': ingresos,
        'gastos': gastos,
        'totalIngresos': sum(i['monto'] for i in ingresos),
        'totalGastos': sum(g['monto'] for g in gastos)
    }

    print(f'  ✓ {hoja_nombre}: Saldo ${rf_actual:,.0f} | {len(ingresos)} ingresos | {len(gastos)} gastos')

datos_completos['bancos'] = bancos

# ==============================================================================
# 2. ALMACÉN MONTE - INVENTARIO
# ==============================================================================
print('\n📦 EXTRAYENDO ALMACÉN MONTE (Inventario)...')

ws_almacen = wb['Almacen_Monte']
rf_almacen = ws_almacen.cell(2, 5).value or 0

# Entradas
entradas = []
for row in range(4, min(ws_almacen.max_row + 1, 100)):
    oc = ws_almacen.cell(row, 1).value
    if oc and str(oc).startswith('OC'):
        entrada = {
            'oc': limpiar_valor(oc),
            'fecha': limpiar_valor(ws_almacen.cell(row, 2).value),
            'distribuidor': limpiar_valor(ws_almacen.cell(row, 3).value),
            'cantidad': limpiar_valor(ws_almacen.cell(row, 4).value)
        }
        entradas.append(entrada)

# Salidas
salidas = []
for row in range(4, min(ws_almacen.max_row + 1, 150)):
    cliente = ws_almacen.cell(row, 8).value
    cantidad = ws_almacen.cell(row, 9).value
    if cliente and cantidad:
        salida = {
            'fecha': limpiar_valor(ws_almacen.cell(row, 7).value),
            'cliente': limpiar_valor(cliente),
            'cantidad': limpiar_valor(cantidad)
        }
        salidas.append(salida)

datos_completos['almacen'] = {
    'stockActual': rf_almacen,
    'entradas': entradas,
    'salidas': salidas,
    'totalEntradas': sum(e['cantidad'] for e in entradas if e['cantidad']),
    'totalSalidas': sum(s['cantidad'] for s in salidas if s['cantidad'])
}

print(f'  ✓ Stock actual: {rf_almacen} unidades')
print(f'  ✓ {len(entradas)} entradas | {len(salidas)} salidas')

# ==============================================================================
# 3. DISTRIBUIDORES - DEUDAS REALES
# ==============================================================================
print('\n📋 EXTRAYENDO DISTRIBUIDORES CON DEUDAS...')

ws_dist = wb['Distribuidores']
distribuidores = []

# Leer columnas 13-16 (Distribuidores, Costo total, Abonos, Pendiente)
for row in range(4, min(ws_dist.max_row + 1, 50)):
    nombre = ws_dist.cell(row, 13).value  # Columna M
    costo_total = ws_dist.cell(row, 14).value  # Columna N
    abonos = ws_dist.cell(row, 15).value  # Columna O
    pendiente = ws_dist.cell(row, 16).value  # Columna P

    if nombre and costo_total:
        distribuidor = {
            'nombre': limpiar_valor(nombre),
            'costoTotal': limpiar_valor(costo_total),
            'abonos': limpiar_valor(abonos) if abonos else 0,
            'deuda': limpiar_valor(pendiente) if pendiente else limpiar_valor(costo_total)
        }
        distribuidores.append(distribuidor)
        print(f'  ✓ {nombre}: Total ${costo_total:,.0f} | Deuda ${distribuidor["deuda"]:,.0f}')

# También extraer las OCs
ocs = []
for row in range(4, min(ws_dist.max_row + 1, 50)):
    oc = ws_dist.cell(row, 1).value
    if oc and str(oc).startswith('OC'):
        oc_data = {
            'oc': limpiar_valor(oc),
            'fecha': limpiar_valor(ws_dist.cell(row, 2).value),
            'origen': limpiar_valor(ws_dist.cell(row, 3).value),
            'cantidad': limpiar_valor(ws_dist.cell(row, 4).value),
            'costoDistribuidor': limpiar_valor(ws_dist.cell(row, 5).value),
            'costoTransporte': limpiar_valor(ws_dist.cell(row, 6).value),
            'costoPorUnidad': limpiar_valor(ws_dist.cell(row, 7).value),
            'costoTotal': limpiar_valor(ws_dist.cell(row, 9).value)
        }
        ocs.append(oc_data)

datos_completos['distribuidores'] = distribuidores
datos_completos['ordenesCompra'] = ocs

# ==============================================================================
# 4. CLIENTES
# ==============================================================================
print('\n👥 EXTRAYENDO CLIENTES...')

ws_clientes = wb['Clientes']
clientes = []

for row in range(4, min(ws_clientes.max_row + 1, 100)):
    nombre = ws_clientes.cell(row, 5).value
    if nombre and str(nombre).strip() and nombre != 'Primo':
        cliente = {
            'nombre': limpiar_valor(nombre),
            'actual': limpiar_valor(ws_clientes.cell(row, 6).value),
            'deuda': limpiar_valor(ws_clientes.cell(row, 7).value) or 0,
            'abonos': limpiar_valor(ws_clientes.cell(row, 8).value) or 0,
            'pendiente': limpiar_valor(ws_clientes.cell(row, 9).value) or 0,
            'observaciones': limpiar_valor(ws_clientes.cell(row, 10).value) or ''
        }
        clientes.append(cliente)

datos_completos['clientes'] = clientes
print(f'  ✓ {len(clientes)} clientes extraídos')

# ==============================================================================
# 5. VENTAS (Control Maestro)
# ==============================================================================
print('\n📊 EXTRAYENDO VENTAS...')

ws_ventas = wb['Control_Maestro']
ventas = []

for row in range(4, min(ws_ventas.max_row + 1, 150)):
    fecha = ws_ventas.cell(row, 1).value
    if fecha:
        venta = {
            'fecha': limpiar_valor(fecha),
            'ocRelacionada': limpiar_valor(ws_ventas.cell(row, 2).value),
            'cantidad': limpiar_valor(ws_ventas.cell(row, 3).value),
            'cliente': limpiar_valor(ws_ventas.cell(row, 4).value),
            'bovedaMonte': limpiar_valor(ws_ventas.cell(row, 5).value),
            'precioVenta': limpiar_valor(ws_ventas.cell(row, 6).value),
            'ingreso': limpiar_valor(ws_ventas.cell(row, 7).value),
            'flete': limpiar_valor(ws_ventas.cell(row, 8).value),
            'fleteUtilidad': limpiar_valor(ws_ventas.cell(row, 9).value),
            'utilidad': limpiar_valor(ws_ventas.cell(row, 10).value),
            'estatus': limpiar_valor(ws_ventas.cell(row, 11).value)
        }
        ventas.append(venta)

datos_completos['ventas'] = ventas
print(f'  ✓ {len(ventas)} ventas extraídas')

# ==============================================================================
# GUARDAR
# ==============================================================================
with open('datos_excel_completos.json', 'w', encoding='utf-8') as f:
    json.dump(datos_completos, f, ensure_ascii=False, indent=2)

print(f'\n\n✅ EXTRACCIÓN COMPLETA EXITOSA')
print(f'\n📊 RESUMEN:')
print(f'   💰 {len(bancos)} Bancos con todos sus movimientos')
print(f'   📦 Almacén: {rf_almacen} unidades en stock')
print(f'   📋 {len(distribuidores)} Distribuidores con deudas')
print(f'   🛒 {len(ocs)} Órdenes de Compra')
print(f'   👥 {len(clientes)} Clientes')
print(f'   📊 {len(ventas)} Ventas')
print(f'\n📄 Datos guardados en: datos_excel_completos.json')
